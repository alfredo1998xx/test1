# app.py  – Hotel Labor Tool
# ───────────────────────────────────────────────────────────────
import os
import streamlit as st
import pandas as pd
import numpy as np
import datetime
import datetime as dt_mod
#import pdfkit
from jinja2 import Template
from datetime import date, timedelta
from st_aggrid import JsCode
from sqlalchemy import func
from dateutil.relativedelta import relativedelta, MO
from sqlalchemy.orm import scoped_session
from sqlalchemy import or_, func
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session as SQLASession
from db import ENGINE
def build_ot_risk_exports(export_df, hotel_name, username, week_start, week_end, sched_week_start, sched_week_end):
      """
      Returns:
            excel_bytes, pdf_bytes
      This function builds both Excel and PDF versions of the OT Risk Report.
      """

      excel_bytes = build_ot_risk_excel(export_df, hotel_name, username, week_start, week_end, sched_week_start, sched_week_end)
      pdf_bytes = build_ot_risk_pdf(export_df, hotel_name, username, week_start, week_end, sched_week_start, sched_week_end)

      return excel_bytes, pdf_bytes
def get_scoped_session():
    return SQLASession(bind=ENGINE)
# ENGINE is imported from db.py (uses DATABASE_URL / PostgreSQL)


def _current_role() -> str:
      # Try a few places; fall back to empty (which means “show everything”)
      user = (st.session_state.get("user") or {})
      return (user.get("role") or st.session_state.get("role") or "").strip().lower()



# ==================================
# OT RISK REPORT FUNCTION
# ==================================
def run_ot_risk_report(
      session,
      week_start,
      week_end,
      sched_week_start,
      sched_week_end,
      dept,
      pos
):
      """
      Executes the FULL OT Risk report logic and returns the resulting DataFrame.
      """

      # === QUERY ACTUALS ===
      q = (
            session.query(
                  db.Actual.emp_id.label("Number"),
                  db.Actual.date.label("Business Date"),
                  (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
                  db.Position.name.label("Position"),
                  db.Department.name.label("Department")
            )
            .join(db.Position, db.Actual.position_id == db.Position.id)
            .join(db.Department, db.Position.department_id == db.Department.id)
            .filter(db.Actual.date.between(week_start, week_end))
            .filter(or_(db.Actual.hours != 0, db.Actual.ot_hours != 0))
      )

      if dept != "(All)":
            q = q.filter(db.Department.name == dept)
      if pos != "(All)":
            q = q.filter(db.Position.name == pos)

      raw = pd.DataFrame(q.all(), columns=["Number", "Business Date", "Hours", "Position", "Department"])
      if raw.empty:
            return None

      # ============================================
      # Normalize Number + Merge Employee Names
      # ============================================
      raw["Number"] = (
            pd.to_numeric(raw["Number"], errors="coerce")
            .fillna(0).astype(int).astype(str).str.zfill(5)
      )

      emp_df = refresh(db.Employee).copy()
      parts = emp_df["name"].astype(str).str.extract(
            r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
      )
      emp_df["ID"] = parts["ID"].fillna("").astype(str).str.strip().str.zfill(5)
      emp_df["First Name"] = parts["First_Name"].str.strip()
      emp_df["Last Name"] = parts["Last_Name"].str.strip()
      emp_df["match_ID"] = emp_df["ID"].astype(str).str.lstrip("0")

      raw["match_ID"] = raw["Number"].astype(str).str.lstrip("0")

      merged = raw.merge(
            emp_df[["match_ID", "First Name", "Last Name"]],
            on="match_ID",
            how="left"
      )

      # ============================================
      # Group Totals (Hours + Days Worked)
      # ============================================
      agg = merged.groupby(["Number", "First Name", "Last Name"]).agg(
            total_hours=("Hours", "sum"),
            days_worked=("Business Date", pd.Series.nunique)
      ).reset_index()

      # ============================================
      # QUERY SCHEDULE WEEK (Mon–Sun)
      # ============================================
      sched_rows = (
            session.query(
                  db.Employee.name,
                  db.Schedule.day,
                  db.Schedule.shift_type
            )
            .join(db.Employee, db.Employee.id == db.Schedule.emp_id)
            .filter(db.Schedule.day.between(sched_week_start, sched_week_end))
            .all()
      )
      sched_df = pd.DataFrame(sched_rows, columns=["name", "day", "shift_type"])

      if not sched_df.empty:

            sched_df["shift_type"] = sched_df["shift_type"].fillna("").astype(str).str.upper().str.strip()
            sched_df = sched_df[sched_df["shift_type"] != "OFF"]

            sched_df["Number"] = sched_df["name"].str.extract(r"(\d+)$")[0].fillna("").str.zfill(5)
            sched_df["day"] = pd.to_datetime(sched_df["day"])

            valid_numbers = set(agg["Number"].astype(str).unique())
            sched_df = sched_df[sched_df["Number"].isin(valid_numbers)]

            merged["Business Date"] = pd.to_datetime(merged["Business Date"])
            last_worked = (
                  merged.groupby("Number")["Business Date"]
                  .max().reset_index().rename(columns={"Business Date": "last_worked"})
            )

            sched_df = sched_df.merge(last_worked, on="Number", how="left")

            sched_df["last_worked"] = sched_df["last_worked"].fillna(
                  pd.to_datetime(week_start) - pd.Timedelta(days=1)
            )

            sched_df["after_work"] = sched_df["day"] > sched_df["last_worked"]

            # Count days
            sched_counts = (
                  sched_df.groupby("Number")["day"]
                  .nunique().reset_index(name="Days Scheduled")
            )

            sched_future = sched_df[sched_df["after_work"]].copy()
            days_remaining = (
                  sched_future.groupby("Number")["day"]
                  .nunique().reset_index(name="Days Remaining")
            )

            # Convert shift text → hours
            def parse_shift_to_hours(shift_str):
                  try:
                        start, end = shift_str.split("-")
                        start_dt = pd.to_datetime(start, format="%H:%M")
                        end_dt = pd.to_datetime(end, format="%H:%M")
                        hours = (end_dt - start_dt).total_seconds() / 3600
                        if hours < 0:
                              hours += 24
                        return max(0, hours - 0.5)
                  except:
                        return 0

            sched_future["shift_hours"] = sched_future["shift_type"].apply(parse_shift_to_hours)

            future_hours = (
                  sched_future.groupby("Number")["shift_hours"]
                  .sum().reset_index(name="Future Scheduled Hrs")
            )

            # merge into agg
            agg = agg.merge(sched_counts, on="Number", how="left")
            agg = agg.merge(days_remaining, on="Number", how="left")
            agg = agg.merge(future_hours, on="Number", how="left")

      else:
            agg["Days Scheduled"] = 0
            agg["Days Remaining"] = 0
            agg["Future Scheduled Hrs"] = 0

      # ============================================
      # FINAL CALCULATIONS
      # ============================================
      agg["Days Scheduled"] = agg["Days Scheduled"].fillna(0).astype(int)
      agg["Days Remaining"] = agg["Days Remaining"].fillna(0).astype(int)
      agg["Future Scheduled Hrs"] = agg["Future Scheduled Hrs"].fillna(0)

      agg["Total Hrs Worked + Schedule"] = (
            agg["total_hours"] + agg["Future Scheduled Hrs"]
      ).round(2)

      # OT Risk Classification
      def classify_ot_risk(row):
            total = row["Total Hrs Worked + Schedule"]
            remaining = row["Days Remaining"]
            if total <= 40:
                  return "No Risk"
            elif remaining > 0:
                  return "At Risk"
            else:
                  return "OT"

      agg["OT Risk"] = agg.apply(classify_ot_risk, axis=1)

      # Risk Percentage
      def estimate_risk_percent(row):
            if row["OT Risk"] == "No Risk":
                  return "0%"
            if row["Days Remaining"] == 0:
                  return "100%"
            elif row["Days Remaining"] == 1:
                  return "80%"
            elif row["Days Remaining"] == 2:
                  return "60%"
            elif row["Days Remaining"] == 3:
                  return "40%"
            return "20%"

      agg["OT Risk %"] = agg.apply(estimate_risk_percent, axis=1)

      # Add OT Cost (1.5 × rate)
      emp_df["rate"] = emp_df.get("hourly_rate", 0).fillna(0)
      agg = agg.merge(emp_df[["ID", "rate"]], left_on="Number", right_on="ID", how="left")
      agg["rate"] = agg["rate"].fillna(0)

      agg["Projected OT"] = agg["Total Hrs Worked + Schedule"].apply(
            lambda h: max(round(h - 40, 2), 0)
      )
      agg["OT Cost"] = (agg["Projected OT"] * agg["rate"] * 1.5).round(2)

      return agg

# ───── ONE-TIME RESET FOR OT RISK ─────
with ENGINE.connect() as conn:
    conn.execute(text("DROP TABLE IF EXISTS ot_risk_all"))
    print("✅ Dropped old ot_risk_all table with missing columns")
from st_aggrid import AgGrid, GridOptionsBuilder
from st_aggrid import AgGrid, GridOptionsBuilder
from db import Session, LaborStandard, RoomForecast
import pandas as pd
import math
import os, json, math, pandas as pd
from datetime import date, timedelta, datetime, time
from db import Position, LaborStandard
from db import RoomForecast
from datetime import date, timedelta
from collections import defaultdict
from db import Actual, Schedule, RoomActual, RoomForecast, Position, Department, ShiftTime
from db import Schedule, Employee, Position, ShiftTime, RoomForecast, Actual
import db                                # local ORM layer

import streamlit as st
import requests
from db import current_hotel_context  # ✅ Import global hotel context

st.set_page_config(page_title="Hotel Labor Tool", layout="wide", initial_sidebar_state="expanded")
# =====================================================
# MANAGER SCOPE FILTER (GLOBAL) — SAFE HERE
# =====================================================
def apply_manager_scope(df):
    user = st.session_state.user

    if user["role"].strip().lower() in ("admin", "super user", "night audit", "asset manager"):
        return df

    scopes = user.get("scope", [])
    if not scopes:
        return df.iloc[0:0]

    mask = False
    for s in scopes:
        mask |= (
            (df["department"] == s["department"]) &
            (df["position"] == s["position"])
        )

    return df[mask]
# ───────── GLOBAL UI + WIDGET STYLING (SINGLE BLOCK — NO GAPS + HOVER ANIMATION) ─────────
st.markdown("""
<style>

/* ───── REMOVE STREAMLIT CHROME ───── */
header[data-testid="stHeader"] { display: none !important; }
div[data-testid="stToolbar"]  { display: none !important; }
#MainMenu, footer             { display: none !important; }

/* ───── MAIN CONTENT BACKGROUND ───── */
[data-testid="stAppViewContainer"] {
    background-color: #fdfdfd !important;
}
[data-testid="stAppViewContainer"] .block-container {
    background-color: #fdfdfd !important;
}

/* ───── ZERO ALL TOP MARGINS / PADDING ───── */
html, body { margin-top:0 !important; padding-top:0 !important; }
[data-testid="stAppViewContainer"] { margin-top:0 !important; padding-top:0 !important; }
[data-testid="stAppViewContainer"] > .main { margin-top:0 !important; padding-top:0 !important; }
[data-testid="stAppViewContainer"] .block-container { margin-top:0 !important; padding-top:0 !important; }
.block-container > :first-child { margin-top:0 !important; }

/* ───── REMOVE MARKDOWN-INJECTED GAPS ───── */
section[data-testid="stMain"] .stMarkdown {
    margin-top: 0 !important;
    margin-bottom: 0 !important;
}

/* ─────────────────────────────────────────────
   UNIFIED FORM STYLING (ALL INPUTS MATCH)
   ───────────────────────────────────────────── */

/* ===== SELECTBOX ===== */
section[data-testid="stMain"] .stSelectbox div[data-baseweb="select"] {
    background-color: #ffffff !important;
    border: 1.5px solid #d1d5db !important;
    border-radius: 10px !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.10) !important;
}
section[data-testid="stMain"] .stSelectbox div[data-baseweb="select"] > div {
    background-color: #ffffff !important;
}

/* ===== RADIO ===== */
section[data-testid="stMain"] .stRadio > div {
    background-color: #ffffff !important;
    border: 1.5px solid #d1d5db !important;
    border-radius: 10px !important;
    padding: 8px 14px !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.10) !important;
}

/* ===== TEXT INPUT ===== */
section[data-testid="stMain"] .stTextInput div[data-baseweb="input"] {
    background-color: #ffffff !important;
    border: 1.5px solid #d1d5db !important;
    border-radius: 10px !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.10) !important;
}
section[data-testid="stMain"] .stTextInput input {
    background-color: #ffffff !important;
    border: none !important;
    padding: 8px 12px !important;
}

/* ===== NUMBER INPUT ===== */
section[data-testid="stMain"] .stNumberInput div[data-baseweb="input"] {
    background-color: #ffffff !important;
    border: 1.5px solid #d1d5db !important;
    border-radius: 10px !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.10) !important;
}
section[data-testid="stMain"] .stNumberInput input {
    background-color: #ffffff !important;
    border: none !important;
    padding: 8px 12px !important;
}

/* ===== DATE INPUT ===== */
section[data-testid="stMain"] .stDateInput div[data-baseweb="datepicker"],
section[data-testid="stMain"] .stDateInput div[data-baseweb="input"] {
    background-color: #ffffff !important;
    border: 1.5px solid #d1d5db !important;
    border-radius: 10px !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.10) !important;
    margin: 0 !important;
    padding: 0 !important;
    min-height: unset !important;
}
section[data-testid="stMain"] .stDateInput input {
    background-color: #ffffff !important;
    border: none !important;
    margin: 0 !important;
    padding: 8px 12px !important;
}

/* ─────────────────────────────────────────────
   PROFESSIONAL LIFT HOVER (NO ZOOM, NO SHIFT)
   ───────────────────────────────────────────── */

section[data-testid="stMain"] .stSelectbox div[data-baseweb="select"],
section[data-testid="stMain"] .stTextInput div[data-baseweb="input"],
section[data-testid="stMain"] .stNumberInput div[data-baseweb="input"],
section[data-testid="stMain"] .stDateInput div[data-baseweb="datepicker"],
section[data-testid="stMain"] .stDateInput div[data-baseweb="input"],
section[data-testid="stMain"] .stRadio > div {
    transition:
        box-shadow 0.2s ease,
        border-color 0.2s ease,
        transform 0.2s ease;
}

/* Hover = gentle lift */
section[data-testid="stMain"] .stSelectbox div[data-baseweb="select"]:hover,
section[data-testid="stMain"] .stTextInput div[data-baseweb="input"]:hover,
section[data-testid="stMain"] .stNumberInput div[data-baseweb="input"]:hover,
section[data-testid="stMain"] .stDateInput div[data-baseweb="datepicker"]:hover,
section[data-testid="stMain"] .stDateInput div[data-baseweb="input"]:hover,
section[data-testid="stMain"] .stRadio > div:hover {
    transform: translateY(-2px);
    box-shadow: 0 10px 22px rgba(0,0,0,0.16);
    border-color: #9ca3af !important;
}

/* ───── HAND CURSOR ON HOVER (ALL FORM INPUTS) ───── */
section[data-testid="stMain"] .stSelectbox div[data-baseweb="select"],
section[data-testid="stMain"] .stTextInput div[data-baseweb="input"],
section[data-testid="stMain"] .stNumberInput div[data-baseweb="input"],
section[data-testid="stMain"] .stDateInput div[data-baseweb="datepicker"],
section[data-testid="stMain"] .stDateInput div[data-baseweb="input"],
section[data-testid="stMain"] .stRadio > div {
    cursor: pointer !important;
}

/* Keep normal text cursor INSIDE the actual input */
section[data-testid="stMain"] .stTextInput input,
section[data-testid="stMain"] .stNumberInput input,
section[data-testid="stMain"] .stDateInput input {
    cursor: text !important;
}
/* ─────────────────────────────────────────────
   ✅ GLOBAL PROFESSIONAL FONT (ONE-TIME CHANGE)
   ───────────────────────────────────────────── */

* {
    font-family: "Inter", -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif !important;
}

</style>
""", unsafe_allow_html=True)

# --- handle logout via query param ---
if st.query_params.get("logout") == "1":
    st.session_state.clear()
    # Hard browser redirect to clean URL — removes ?logout=1 from address bar
    st.markdown('<meta http-equiv="refresh" content="0;url=/">', unsafe_allow_html=True)
    st.stop()

# ---------- decide if the user is logged in ----------
# Pick the key your app sets after successful login:
# e.g., st.session_state["user"] or ["jwt"] or ["logged_in"] = True
is_logged_in = bool(st.session_state.get("user")) or bool(st.session_state.get("logged_in")) or bool(st.session_state.get("jwt"))

# =========================================================
# LATEST ACTUAL HOURS QUERY
# =========================================================
def get_latest_actual_hours_date():
    session = get_scoped_session()

    row = (
        session.query(func.max(Actual.date))
        .filter(
            Actual.source.in_(["manual", "contract"]),
            (Actual.hours > 0) | (Actual.ot_hours > 0)
        )
        .first()
    )

    session.close()

    return row[0] if row and row[0] else None
# =========================================================
# POPUP SHOULD SHOW ONLY ON LOGIN
# =========================================================
if "show_login_popup" not in st.session_state:
    st.session_state.show_login_popup = True
# =========================================================
# STREAMLIT NATIVE POPUP (MODAL)
# =========================================================
@st.dialog("Welcome Back!")
def show_login_popup():
    # ---------- CSS for popup ONLY ----------
    st.markdown("""
    <style>
    /* Red button styling */
    button[kind="primary"] {
        background-color: #e74c3c !important;
        color: #ffffff !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
    }
    button[kind="primary"]:hover {
        background-color: #c0392b !important;
    }

    /* Border + Animation on dialog container */
    .stDialog .st-emotion-cache-1wqg5s7 {
        border: 2px solid #E0E0E0 !important;
        border-radius: 14px !important;
        animation: popupFadeZoom 0.22s ease-out !important;
    }

    /* Fade + Zoom animation */
    @keyframes popupFadeZoom {
        0%   { opacity:0; transform: scale(0.95); }
        100% { opacity:1; transform: scale(1); }
    }
    </style>
    """, unsafe_allow_html=True)
    # CSS just for this modal
    st.markdown("""
    <style>
    button[kind="primary"] {
        background-color: #e74c3c !important;
        color: #ffffff !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
    }
    button[kind="primary"]:hover {
        background-color: #c0392b !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ---------- 1. Get latest actual hours date ----------
    latest_actual_date = st.session_state.get("latest_actual_hours_date")

    if latest_actual_date:
        latest_text = latest_actual_date.strftime("%B %d, %Y")
    else:
        latest_text = "No uploads yet"

    # ---------- 2. Latest Hours Section ----------
    st.markdown(f"""
    <div style="
        background:#EAF3FF;
        padding:14px 16px;
        border-radius:10px;
        border:1px solid #D5E7FF;
        margin-bottom:12px;
    ">
        <div style="font-size:18px; font-weight:600;">📅 Latest Actual Hours Uploaded:</div>
        <div style="font-size:16px; margin-top:4px; font-weight:700;">{latest_text}</div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    # ---------- 3. Reminder Section ----------
    st.markdown("""
    ### 📝 **Before You Start**
    Please make sure to:

    - Upload the latest **Employee Roster**  
    - Upload the latest **Hourly Employee Pay** report  
    - Verify all **Department & Position Mapping**
    """)


    # ---------- 4. Close button ----------
    if st.button("Got it", use_container_width=True, type="primary"):
        st.session_state.show_login_popup = False
        st.rerun()

# =========================================================
# TRIGGER POPUP ONLY IF LOGGED IN
# =========================================================
if is_logged_in:

    # Always update the latest hours date
    st.session_state["latest_actual_hours_date"] = get_latest_actual_hours_date()

    # Show popup once
    if st.session_state.show_login_popup:
        show_login_popup()
# ---------- only show sticky top bar when logged in ----------
if is_logged_in:

    hotel_name = (st.session_state.user or {}).get("hotel_name", "") or ""

    st.markdown(f"""
    <style>
    /* ===== FIXED top bar ===== */
    .topbar {{
      position: fixed;
      top: 0; left: 0; right: 0;
      height: 60px;
      background: #ffffff;
      border-bottom: 1px solid #eee;
      z-index: 10000;
      display: flex; align-items: center; gap: 12px;
      padding: 10px 16px;

      /* subtle drop shadow */
      box-shadow: 0 2px 6px rgba(0,0,0,0.07);
    }}

    /* Push content down ONLY when topbar exists */
    div[data-testid="stAppViewContainer"] .main .block-container {{
      padding-top: 80px;
    }}

    .topbar-title {{ font-weight: 800; font-size: 22px; }}
    .topbar-spacer {{ flex: 1; }}

    /* ----- centered hotel name WITHOUT affecting layout ----- */
    .topbar-center {{
        position: absolute;
        left: 50%;
        transform: translateX(-50%);
        pointer-events: none;
        font-weight: 800;
        font-size: 22px;
        white-space: nowrap;
    }}

    /* ✅ username next to logout — vertically centered */
    .topbar-user {{
        margin-right: 10px;
        font-size: 14px;
        font-weight: 600;
        opacity: 0.75;
        white-space: nowrap;

        display: inline-flex;
        align-items: center;
        height: 100%;
    }}

    /* Logout button style */
    .topbar-logout {{
      display: inline-block;
      background: #ff5a57;
      color: #ffffff !important;
      font-weight: 600;
      font-size: 14px;
      line-height: 1;
      padding: 10px 16px;
      border-radius: 25px;
      text-decoration: none !important;
      border: none;
      box-shadow: 0 2px 3px rgba(0,0,0,0.08);
      transition: all 0.15s ease-in-out;
      user-select: none;
    }}
    .topbar-logout:hover {{
      filter: brightness(0.97);
      transform: translateY(-1px);
      box-shadow: 0 4px 6px rgba(0,0,0,0.12);
    }}
    .topbar-logout:active {{
      transform: translateY(1px);
      box-shadow: 0 1px 2px rgba(0,0,0,0.10);
    }}
    </style>

    <div class="topbar">

      <!-- CENTERED HOTEL NAME -->
      <div class="topbar-center">{hotel_name}</div>

      <!-- SPACER + USER + LOGOUT BUTTON -->
      <div class="topbar-spacer"></div>

      <span class="topbar-user">
          {st.session_state.user.get("username","")}
      </span>
      <a class="topbar-logout" href="?logout=1" target="_self" rel="noopener">
          Logout
      </a>

    </div>
    """,
    unsafe_allow_html=True
)


# Cosmetic only: hide Streamlit toolbar/header/footer
st.markdown("""
<style>
div[data-testid="stToolbar"] { display: none !important; }
#MainMenu { visibility: hidden; }
header { visibility: hidden; }
footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


API_URL = "http://localhost:8000"

# ─────────────────────────────────────────
# PASSWORD RESET PAGE (MATCH LOGIN STYLE – NO WHITE BOX)
# ─────────────────────────────────────────
query_params = st.query_params
reset_token = query_params.get("token")

if reset_token:

    # Center layout (same as login)
    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:

        hotel_label = st.session_state.get("user", {}).get("hotel_name", "Hotel")

        # Title (same style as login)
        st.title(f"Labor Pilot | Reset Password | {hotel_label}")



        # Reset form
        with st.form("reset_password_form"):
            new_pass_1 = st.text_input("New Password", type="password")
            new_pass_2 = st.text_input("Confirm Password", type="password")
            submit_reset = st.form_submit_button("Reset Password")

        if submit_reset:

            if not new_pass_1 or not new_pass_2:
                st.error("❌ Please complete both password fields.")
                st.stop()

            if new_pass_1 != new_pass_2:
                st.error("❌ Passwords do not match.")
                st.stop()

            # Call FastAPI reset endpoint
            try:
                resp = requests.post(
                    f"{API_URL}/reset-password",
                    params={
                        "token": reset_token,
                        "new_password": new_pass_1
                    }
                )

                if resp.status_code == 200:
                    st.success("✅ Password reset successful! Redirecting to login...")

                    st.markdown(
                        "<meta http-equiv='refresh' content='2;url=/' />",
                        unsafe_allow_html=True
                    )
                    st.stop()

                else:
                    st.error(f"❌ {resp.json().get('detail','Reset failed')}")

            except Exception as e:
                st.error(f"❌ {e}")

    st.stop()



# ─────────────────────────────────────────
# 🔄 RESET PASSWORD REQUEST PAGE
# ─────────────────────────────────────────
if st.query_params.get("reset_request") == "true":

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:

        st.title("Forgot Password")
        st.markdown("Enter your email address to receive a password reset link.")

        with st.form("forgot_password_form"):
            email_input = st.text_input("Email Address")
            submit_reset = st.form_submit_button("Send Reset Link")

        if submit_reset:
            resp = requests.post(
                f"{API_URL}/forgot-password",
                params={"email": email_input}
            )

            if resp.status_code == 200:
                st.success("📩 A password reset link has been sent if this email is registered.")
            else:
                st.error("❌ Unable to process password reset.")

    st.stop()



# ─────────────────────────────────────────
# 🔐 LOGIN PAGE
# ─────────────────────────────────────────
if "token" not in st.session_state:

    col1, col2, col3 = st.columns([1, 2, 1])  # Center the login form

    with col2:

        st.title("Hotel Login")

        _form_ver = st.session_state.get("_logout_count", 0)
        with st.form(f"login_form_{_form_ver}"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submit = st.form_submit_button("Login", type="primary")
            # NEW — Streamlit-native Forgot Password button
            forgot = st.form_submit_button("Forgot Password?", type="secondary")



        # Process Forgot Password click
        if forgot:
            st.query_params["reset_request"] = "true"
            st.rerun()

        # Process login
        if submit:
            response = requests.post(f"{API_URL}/login", json={
                "username": username,
                "password": password
            })

            if response.status_code == 200:

                st.success("Login successful!")

                st.session_state.token = response.json()["access_token"]

                # Fetch hotel/user info
                user = requests.get(
                    f"{API_URL}/me",
                    headers={"Authorization": f"Bearer {st.session_state.token}"}
                )

                if user.status_code == 200:
                    data = user.json()

                    st.session_state.user = {
                        "username": data["username"],
                        "role": data["role"],
                        "hotel_name": data["hotel_name"]
                    }

                    # -------------------------------
                    # LOAD MANAGER SCOPE (ON LOGIN)
                    # -------------------------------
                    scope_resp = requests.get(
                        f"{API_URL}/my-scope",
                        headers={"Authorization": f"Bearer {st.session_state.token}"}
                    )

                    st.session_state.user["scope"] = (
                        scope_resp.json() if scope_resp.status_code == 200 else []
                    )

                    st.session_state.hotel_name = data["hotel_name"]
                    current_hotel_context["hotel_name"] = data["hotel_name"]

                    st.rerun()
                else:
                    st.error("Could not retrieve user info.")

            else:
                st.error("Invalid username or password.")

        # Demo Request Section
        with st.expander("**Request a Demo**"):
            with st.form("demo_request_form"):
                st.write("We'll reach out to set up a personalized demo.")
                name = st.text_input("Full Name")
                email = st.text_input("Email")
                hotel = st.text_input("Hotel Name")
                notes = st.text_area("Message (optional)")
                send = st.form_submit_button("Submit")

            if send:
                st.success("✅ Your demo request has been submitted!")

    st.stop()


# After login
if "hotel_name" in st.session_state:
    current_hotel_context["hotel_name"] = st.session_state.hotel_name




# ─────────────── Utility Functions ───────────────

def load_acl_once():
      """Fetch (department, position) pairs + hourly-flag for the logged-in user and cache in session."""
      if st.session_state.get("acl_loaded"):
            return
      headers = {"Authorization": f"Bearer {st.session_state.token}"}

      try:
            # Use /my/access — works for all roles (Admin, Manager, Employee, Night Audit)
            r = requests.get(f"{API_URL}/my/access", headers=headers, timeout=10)
            rows = r.json() if r.status_code == 200 else []
      except Exception:
            rows = []

      pairs = {(row["department"], row["position"]) for row in rows}
      pos_by_dept = {}
      for d, p in pairs:
            pos_by_dept.setdefault(d, []).append(p)
      for d in pos_by_dept:
            pos_by_dept[d] = sorted(pos_by_dept[d])

      st.session_state.acl = {
            "pairs": pairs,
            "departments": sorted({d for d, _ in pairs}),
            "positions_by_dept": pos_by_dept,
            "can_view_hourly": any(row.get("can_view_hourly_rate") for row in rows),
      }
      st.session_state.acl_loaded = True

def apply_acl_to_df(df: pd.DataFrame) -> pd.DataFrame:
      """Return only rows the manager is allowed to see. Also hides hourly_rate if not allowed."""
      role = st.session_state.user["role"].strip().lower()
      if role in ("admin", "super user"):
            return df  # no restriction
      acl = st.session_state.get("acl", {"pairs": set(), "can_view_hourly": False})
      pairs = acl["pairs"]
      if not pairs or df.empty:
            return df.iloc[0:0]  # nothing allowed

      # Keep rows where (department, role/position) is allowed
      # Your data sometimes uses 'role' for the position field.
      pos_col = "role" if "role" in df.columns else "position"
      mask = df.apply(lambda r: (r.get("department"), r.get(pos_col)) in pairs, axis=1)
      out = df[mask].copy()

      if not acl["can_view_hourly"] and "hourly_rate" in out.columns:
            out["hourly_rate"] = None
      return out

def get_week_start(any_date=None):
      if any_date is None:
            any_date = date.today()
      return any_date - timedelta(days=any_date.weekday())  # Week starts on Monday

def today():
      return date.today()

def generate_ot_risk_data(week_start, week_end, sel_dept, sel_pos):
    from sqlalchemy import or_

    # ───── Pull actual hours with department & position ─────
    q = (
        session.query(
            db.Actual.emp_id.label("Number"),
            db.Actual.date.label("Business Date"),
            (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
            db.Position.name.label("Position"),
            db.Department.name.label("Department")
        )
        .join(db.Position, db.Actual.position_id == db.Position.id)
        .join(db.Department, db.Position.department_id == db.Department.id)
        .filter(db.Actual.date.between(week_start, week_end))
        .filter(or_(db.Actual.hours != 0, db.Actual.ot_hours != 0))
    )

    if sel_dept:
        q = q.filter(db.Department.name == sel_dept)
    if sel_pos:
        q = q.filter(db.Position.name == sel_pos)

    raw = pd.DataFrame(q.all(), columns=["Number", "Business Date", "Hours", "Position", "Department"])

    if raw.empty:
        return pd.DataFrame()

    # ───── Match employee names ─────
    emp_df = refresh(db.Employee).copy()
    parts = emp_df["name"].astype(str).str.extract(
        r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
    )
    emp_df["ID"] = parts["ID"].fillna("").astype(str).str.zfill(5)
    emp_df["First Name"] = parts["First_Name"].str.strip()
    emp_df["Last Name"] = parts["Last_Name"].str.strip()
    emp_df["match_ID"] = emp_df["ID"].astype(str).str.lstrip("0")
    raw["match_ID"] = raw["Number"].astype(str).str.lstrip("0")

    merged = raw.merge(emp_df[["match_ID", "First Name", "Last Name"]], on="match_ID", how="left")

    # ───── Aggregate actuals ─────
    agg = merged.groupby(["Number", "First Name", "Last Name"]).agg(
        total_hours=("Hours", "sum"),
        days_worked=("Business Date", pd.Series.nunique)
    ).reset_index()

    # Ensure 'Number' is string before merging
    agg["Number"] = agg["Number"].astype(str)

    # ───── Fill placeholders for shift logic ─────
    agg["Days Scheduled"] = 0
    agg["Days Remaining"] = 0
    agg["Future Scheduled Hrs"] = 0
    agg["Total Hrs Worked + Schedule"] = agg["total_hours"].round(2)

    def classify_ot_risk(row):
        if row["Total Hrs Worked + Schedule"] <= 40:
            return "No Risk"
        return "OT"

    agg["OT Risk"] = agg.apply(classify_ot_risk, axis=1)
    agg["OT Risk %"] = agg["total_hours"].apply(
        lambda h: "0%" if h <= 40 else f"{round(((h - 40)/40)*100)}%"
    )
    agg["Projected OT"] = agg["total_hours"].apply(lambda h: max(round(h - 40, 2), 0))

    # ───── Merge hourly rate ─────
    emp_df["ID"] = emp_df["ID"].astype(str)  # Coerce to match
    if "hourly_rate" in emp_df.columns:
        emp_df["rate"] = emp_df["hourly_rate"].fillna(0)
    else:
        emp_df["rate"] = 0.00

    agg = agg.merge(emp_df[["ID", "rate"]], left_on="Number", right_on="ID", how="left")
    agg["rate"] = agg["rate"].fillna(0)
    agg["OT Cost"] = (agg["Projected OT"] * agg["rate"] * 1.5).round(2)

    # ───── Add week, department, position for filtering ─────
    agg["department"] = sel_dept
    agg["position"] = sel_pos

    return agg
def save_ot_risk_to_db(week_start, week_end, sel_dept, sel_pos):
    from sqlalchemy import create_engine
    # ENGINE is imported from db.py (uses DATABASE_URL / PostgreSQL)

    df = generate_ot_risk_data(week_start, week_end, sel_dept, sel_pos)
    if df.empty:
        return False

    df["Business Date"] = pd.to_datetime(week_start)  # Required for filtering later
    with ENGINE.connect() as conn:
        try:
            df.to_sql("ot_risk_all", conn, if_exists="append", index=False)
            return True
        except Exception as e:
            print("❌ Failed to save OT Risk:", e)
            return False

session = scoped_session(db.Session)

# ---------- helpers -----------------------------------------------------------
def refresh(model):
    """Return hotel-scoped SQLAlchemy table as DataFrame."""
    hotel_name = st.session_state.get("hotel_name")
    if hotel_name and hasattr(model, "hotel_name"):
        query = session.query(model).filter_by(hotel_name=hotel_name)
    else:
        query = session.query(model)
    return pd.read_sql(query.statement, session.bind)
# ---------- helper: convert weekday label -> placeholder DATE --------------
from datetime import date  # already imported at top
# map 3-letter weekday -> placeholder DATE in the dummy week
WK_TO_DATE = {
    "Mon": date(2000, 1, 3),
    "Tue": date(2000, 1, 4),
    "Wed": date(2000, 1, 5),
    "Thu": date(2000, 1, 6),
    "Fri": date(2000, 1, 7),
    "Sat": date(2000, 1, 8),
    "Sun": date(2000, 1, 9),
}
def week_cols(week_start):
    """Return 7 day headers; %#d/#m for Windows, %-d/-m for POSIX."""
    fmt = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
    return [(week_start + timedelta(d)).strftime(fmt) for d in range(7)]

def refresh_totals(session, keys):
    """
    keys → set of (position_id, date) pairs.
    Updates the 'total' source in db.Actual to equal manual + contract for each.
    """
    for pos_id, biz_date in keys:
        # Sum all relevant sources (manual + contract)
        sums = session.query(
            func.coalesce(func.sum(db.Actual.hours), 0),
            func.coalesce(func.sum(db.Actual.ot_hours), 0),
            func.coalesce(func.sum(db.Actual.reg_pay), 0),
            func.coalesce(func.sum(db.Actual.ot_pay), 0),
        ).filter(
            db.Actual.position_id == pos_id,
            db.Actual.date == biz_date,
            db.Actual.source.in_(["manual", "contract"])
        ).one()

        total_rec = (
            session.query(db.Actual)
                .filter_by(position_id=pos_id, date=biz_date, source="total")
                .one_or_none()
        )

        if not total_rec:
            total_rec = db.Actual(
                emp_id=None,
                position_id=pos_id,
                date=biz_date,
                source="total"
            )
            session.add(total_rec)

        total_rec.hours, total_rec.ot_hours, total_rec.reg_pay, total_rec.ot_pay = sums

def get_week_dates():
    from datetime import date, timedelta
    today = date.today()
    start = today - timedelta(days=today.weekday())
    return [start + timedelta(days=i) for i in range(7)]


def load_labor_standards(pos_id):
    rows = session.query(db.LaborStandard).filter_by(position_id=pos_id).all()
    return pd.DataFrame([{
        "Metric": r.metric,
        "Standard": r.standard,
        "Unit": r.unit
    } for r in rows]) if rows else pd.DataFrame(columns=["Metric", "Standard", "Unit"])

def save_labor_standards(position_id, df, hotel_name=None):
    # keep the same session pattern you use elsewhere
    session = get_scoped_session()

    # resolve hotel_name if not provided
    if not hotel_name:
        try:
            hotel_name = (st.session_state.get("user") or {}).get("hotel_name")
        except Exception:
            hotel_name = None
        if not hotel_name:
            try:
                from db import current_hotel_context
                hotel_name = current_hotel_context.get("hotel_name")
            except Exception:
                hotel_name = None

    if not hotel_name:
        # fail fast rather than inserting NULL into NOT NULL column
        raise ValueError("No hotel selected in session; cannot save scoped labor standards.")

    # delete ONLY existing standards for this position + hotel
    try:
        session.query(db.LaborStandard).filter(
            db.LaborStandard.position_id == position_id,
            getattr(db.LaborStandard, "hotel_name") == hotel_name
        ).delete(synchronize_session=False)
    except Exception:
        # if your class name is different (e.g., LaborStandards), adjust here
        session.query(db.LaborStandard).filter(
            db.LaborStandard.position_id == position_id,
            getattr(db.LaborStandard, "hotel_name") == hotel_name
        ).delete(synchronize_session=False)

    # build new rows with explicit hotel_name
    new_rows = []
    for _, r in df.iterrows():
        metric = (str(r.get("Metric") or "")).strip()
        if not metric:
            continue
        standard = r.get("Standard")
        try:
            standard = float(standard)
        except Exception:
            continue
        unit = (str(r.get("Unit") or "")).strip()

        new_rows.append(
            db.LaborStandard(
                position_id=position_id,
                metric=metric,
                standard=standard,
                unit=unit,
                hotel_name=hotel_name   # <<< important
            )
        )

    if new_rows:
        session.add_all(new_rows)

    session.commit()


# Only matters for Managers; harmless for Admins/Super Users
load_acl_once()

# ---------- Streamlit page config ---------------------------------------------

# ───── Global Top Title (Displayed Once) ─────
st.markdown("""
    <style>
    .top-bar-title {
        font-size: 22px;
        font-weight: 600;
        margin-top: -20px;
        margin-bottom: 10px;
    }

    /* Reduce spacing between radio options */
    .stRadio > div {
        gap: 0.25rem !important;
    }

    /* Style each label: prevent wrapping, reduce padding */
    .stRadio label {
        font-size: 15px !important;
        padding: 4px 10px !important;
        border-radius: 6px;
        white-space: nowrap !important;  /* Prevent wrapping */
        text-overflow: ellipsis !important;
        overflow: hidden !important;
    }

    /* Slight hover effect */
    .stRadio label:hover {
        background-color: rgba(255,255,255,0.05);
    }
    </style>
""", unsafe_allow_html=True)


# ───── Sidebar Navigation ─────

if "user" not in st.session_state:
    st.session_state.user = {"username": None, "role": None, "hotel_name": None}

# normalize role once
role = (st.session_state.user.get("role") or "").strip().lower()

menu_options = [
    "Dashboard",
    "Employees",
    "Labor ▸ Structure",
    "Labor ▸ Actual Hours",
    "Room STATs",
    "Scheduling",
    "Cost and OT Mgmt",
    "Reports",
    "Scheduled Tasks",
]

# Role-based menu access
if role == "employee":
    # Employees only see the Scheduling page
    menu_options = ["Scheduling"]
elif role == "night audit":
    # Night Audit only sees Room STATs
    menu_options = ["Room STATs"]
elif role == "asset manager":
    # Asset Managers only see the Reports page
    menu_options = ["Reports"]
else:
    # Hide Labor Structure from Managers
    if role == "manager":
        menu_options = [m for m in menu_options if m != "Labor ▸ Structure"]
    # Super User and Admin get the Admin tab
    if role in ("super user", "admin"):
        menu_options.append("Admin")

main_choice = st.sidebar.radio("Menu", menu_options, key="main_menu")

# ── Sidebar logo (bottom-centered) ─────────────────────────────────────────
import base64 as _b64, os as _os
_logo_path = _os.path.join(_os.path.dirname(__file__), "attached_assets",
                           "laborpilot_logo_nobg.png")
if _os.path.exists(_logo_path):
    with open(_logo_path, "rb") as _f:
        _logo_b64 = _b64.b64encode(_f.read()).decode()
    st.sidebar.markdown(f"""
        <style>
        .lp-sidebar-logo {{
            position : fixed;
            bottom   : 12px;
            left     : 0;
            width    : 220px;
            display  : flex;
            justify-content: center;
            padding  : 0 4px;
            box-sizing: border-box;
            z-index  : 9999;
        }}
        .lp-sidebar-logo img {{
            width    : 212px;
            max-width: 100%;
            height   : auto;
        }}
        </style>
        <div class="lp-sidebar-logo">
            <img src="data:image/png;base64,{_logo_b64}" alt="LaborPilot">
        </div>
    """, unsafe_allow_html=True)

st.markdown("""
    <style>
    /* ───── Dynamic Sidebar Gradient Background ───── */
    [data-testid="stSidebar"] {
        background: linear-gradient(135deg, #2e2e2e, #4a4a4a, #3c3c3c) !important;
        background-size: 400% 400%;
        animation: gradientShift 20s ease infinite;
        box-shadow: 2px 0 10px rgba(0,0,0,0.25);
    }

    @keyframes gradientShift {
        0%   { background-position: 0% 50%; }
        50%  { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }

    /* ───── Sidebar Width ───── */
    [data-testid="stSidebar"][aria-expanded="true"] > div:first-child {
        width: 220px;
        min-width: 220px;
    }

    /* ───── Sidebar Text Styling ───── */
    [data-testid="stSidebar"] * {
        color: #ffffff !important;
    }

    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] span {
        font-size: 13px !important;
    }

    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3 {
        color: #f0f0f0 !important;
    }



/* ✅ Active (selected) = subtle highlight + BOLD */
section[data-testid="stSidebar"] label:has(input:checked) {
    transform: translateX(6px);
    background-color: rgba(255,255,255,0.12);
    font-weight: 700 !important;   /* ✅ BOLD when selected */
}

/* ✅ Ensure hover does NOT force bold */
section[data-testid="stSidebar"] label:hover {
    font-weight: 400 !important;
}

/* ✅ "Menu" title NEVER bold, NEVER highlighted */
section[data-testid="stSidebar"] label:has(span:contains("Menu")) {
    background-color: transparent !important;
    transform: none !important;
    box-shadow: none !important;
    font-weight: 400 !important;
}
    </style>
""", unsafe_allow_html=True)

# ===== Stable page-change overlay (no tab resets) ============================
import streamlit as st, time

# 1) one-time CSS (fixed id; no uuid)
if "_overlay_css_done" not in st.session_state:
      st.markdown("""
      <style>
      /* Fixed overlay id; no dynamic keys */
      #la_overlay_fixed{
        position: fixed; inset: 0; z-index: 99999;
        display: flex; align-items: center; justify-content: center;
        background: #ffffff;
        pointer-events: none;                /* purely visual */
        opacity: 0; visibility: hidden;
      }
      /* simple fade-out keyframes (1s) */
      @keyframes la_fadeout {
        0%   { opacity: 1;   visibility: visible; }
        90%  { opacity: 1;   visibility: visible; }
        100% { opacity: 0;   visibility: hidden;  }
      }
      svg.la_spinner { width:72px; height:72px; }
      </style>
      """, unsafe_allow_html=True)
      st.session_state._overlay_css_done = True

# 2) track last page; only show overlay when main_choice changes
if "_last_page" not in st.session_state:
      st.session_state._last_page = None

page_changed = (main_choice != st.session_state._last_page)
if page_changed:
      st.session_state._last_page = main_choice

# 3) fixed container so the component tree stays stable
_overlay = st.empty()

# 4) only render overlay on actual page change, then clear after ~1s
if page_changed:
      _overlay.markdown("""
      <div id="la_overlay_fixed" aria-label="Loading"
           style="opacity:1; visibility:visible;">

        <!-- White backdrop ONLY over main content (sidebar width = 220px) -->
        <div style="
             position:fixed; top:0; left:255px; right:0; bottom:0;
             background:#ffffff; pointer-events:none;
             z-index:2147483646;                  /* below spinner */
             animation: la_fadeout 1.0s forwards;">
        </div>

        <!-- keep your spinner exactly as-is (viewport centered) -->
        <div style="
             position:fixed; z-index:2147483647;
             top:50vh; left:50vw; transform:translate(-50%,-50%);
             display:flex; align-items:center; justify-content:center;
             animation: la_fadeout 1.0s forwards;">
          <svg class="la_spinner"
               width="72" height="72"
               style="width:24px!important;height:24px!important;display:block;
                      max-width:none!important;max-height:none!important;"
               viewBox="0 0 44 44" preserveAspectRatio="xMidYMid meet"
               fill="none" xmlns="http://www.w3.org/2000/svg">
            <circle cx="22" cy="22" r="18" stroke="#E9F1FF" stroke-width="3"></circle>
            <circle cx="22" cy="22" r="18" stroke="#ff5a57" stroke-width="3"
                    stroke-linecap="round" stroke-dasharray="50 200">
              <animateTransform attributeName="transform" type="rotate"
                                from="0 22 22" to="360 22 22" dur="0.9s"
                                repeatCount="indefinite"/>
            </circle>
          </svg>
        </div>

      </div>
      """, unsafe_allow_html=True)
      time.sleep(1.05)
      _overlay.empty()
# ===== END stable overlay ====================================================



# ======================================= DASHBOARD ==========================
if main_choice == "Dashboard":
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    from datetime import date, timedelta
    import math, pandas as pd
    from sqlalchemy import func, and_
    import requests
    # ── Title: Lucide "layout-dashboard" icon + text (pure B/W) ──
    st.markdown("""
    <div class="la-title">
      <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
           viewBox="0 0 24 24" fill="none" stroke="currentColor"
           stroke-width="2" stroke-linecap="round" stroke-linejoin="round"
           aria-hidden="true">
        <rect x="3"  y="3"  width="7" height="9"></rect>
        <rect x="14" y="3"  width="7" height="5"></rect>
        <rect x="14" y="12" width="7" height="9"></rect>
        <rect x="3"  y="16" width="7" height="5"></rect>
      </svg>
      <span>Labor Analytics Dashboard</span>
    </div>
    <style>
      .la-title{
        display:flex; align-items:center; gap:10px;
        margin:0 0 10px 0; line-height:1;
      }
      .la-title svg{ color:#111; }
      .la-title span{ font-weight:700; font-size:1.6rem; }
    </style>
    """, unsafe_allow_html=True)
    st.markdown("""
    <style>

    /* ✅ OUTER STREAMLIT HOLDER */
    div[data-testid="stPlotlyChart"] {
        width: 100% !important;
        max-width: 100% !important;
        padding: 0 !important;
        overflow: visible !important;   /* ✅ allow legend breathing room */
    }

    /* ✅ BORDERED CARD */
    div[data-testid="stPlotlyChart"] > div {
        background: #ffffff !important;
        border: 1px solid #e0e0e0 !important;
        border-radius: 0px !important;      
        padding: 12px 48px 12px 12px !important;   /* ✅ gives legend extra breathing room */
        box-shadow: 0 8px 18px rgba(0,0,0,0.12) !important;
        width: 100% !important;
        max-width: 100% !important;
        overflow-y: hidden !important;   /* ✅ vertical clip only */
        overflow-x: visible !important;  /* ✅ allow legend */
        box-sizing: border-box !important;
        position: relative !important;
    }

    /* ✅ FORCE GRAPH ITSELF TO STAY INSIDE */
    div[data-testid="stPlotlyChart"] .js-plotly-plot,
    div[data-testid="stPlotlyChart"] .plot-container,
    div[data-testid="stPlotlyChart"] .svg-container,
    div[data-testid="stPlotlyChart"] svg,
    div[data-testid="stPlotlyChart"] canvas {
        max-width: 100% !important;
        width: 100% !important;
        overflow: visible !important;
    }

    /* ✅ REMOVE PLOTLY TOOLBAR */
    div[data-testid="stPlotlyChart"] .modebar {
        display: none !important;
    }

    </style>
    """, unsafe_allow_html=True)


    # STEP 2B — Load ACL for the logged-in user (manager only)
    API_URL = "http://127.0.0.1:8000"  # change if your API runs elsewhere

    def load_acl_for_current_user():
        role_norm = st.session_state.user["role"].strip().lower()
        if role_norm != "manager":
            st.session_state.acl = None
            return
        if st.session_state.get("acl_loaded"):
            return
        headers = {"Authorization": f"Bearer {st.session_state.token}"}
        # ... rest of your function ...


    # STEP 2B — Load ACL for the logged-in user (manager only)
    API_URL = "http://127.0.0.1:8000"  # change if your API runs elsewhere

    def load_acl_for_current_user():
        # Only managers need ACL; harmless for others
        role_norm = st.session_state.user["role"].strip().lower()
        if role_norm != "manager":
            st.session_state.acl = None
            return

        # Avoid reloading on every rerun
        if st.session_state.get("acl_loaded"):
            return

        headers = {"Authorization": f"Bearer {st.session_state.token}"}
        try:
            r = requests.get(f"{API_URL}/my/access", headers=headers, timeout=10)
            rows = r.json() if r.status_code == 200 else []
        except Exception:
            rows = []

        # Build allowed departments and positions-per-dept
        depts = sorted({row["department"] for row in rows})
        pos_by_dept = {}
        for row in rows:
            pos_by_dept.setdefault(row["department"], set()).add(row["position"])
        pos_by_dept = {d: sorted(list(s)) for d, s in pos_by_dept.items()}

        st.session_state.acl = {
            "departments": depts,
            "positions_by_dept": pos_by_dept,
            "can_view_hourly": any(row.get("can_view_hourly_rate") for row in rows),
        }
        st.session_state.acl_loaded = True

    # Call the loader once before building filters
    load_acl_for_current_user()


    yesterday = date.today() - timedelta(days=1)
    default_from = yesterday - timedelta(days=7)

    # ─────────── EMPLOYEE BASE (SCOPED FOR DASHBOARD) ───────────
    emp_df = refresh(db.Employee)

    # 🔒 ENFORCE MANAGER SCOPE
    emp_df = apply_manager_scope(
        emp_df.rename(columns={"role": "position"})
    ).rename(columns={"position": "role"})
    # ─────────── 1. FILTER BAR ────────────────────────────────────────────
    f_col1, f_col2, f_col3, f_col4 = st.columns(4)

    with f_col1:
        dt_from = st.date_input(
            "From",
            value=st.session_state.get("la_from", yesterday),
            key="la_from",
            format="MM/DD/YYYY",
        )

    with f_col2:
        dt_to = st.date_input(
            "To",
            value=st.session_state.get("la_to", yesterday),
            key="la_to",
            format="MM/DD/YYYY",
        )

    # ✅ FRIENDLY VALIDATION (NO CRASH)
    if dt_from > dt_to:
        st.error("⚠️ 'From' date cannot be later than 'To' date.")
        st.stop()   # stops chart rendering safely

    role_norm = st.session_state.user["role"].strip().lower()

    with f_col3:
        if role_norm == "manager" and st.session_state.acl and st.session_state.acl.get("departments"):
            # Manager: only allowed departments (keep 'All' = all allowed)
            dept_options = ["All"] + st.session_state.acl["departments"]
            sel_dept = st.selectbox(
                "Department",
                dept_options,
                key="la_dept",
            )
        else:
            # Admin / Super User: original behavior
            sel_dept = st.selectbox(
                "Department",
                ["All"] + sorted(emp_df["department"].dropna().unique()),
                key="la_dept",
            )

    with f_col4:
        if role_norm == "manager" and st.session_state.acl and st.session_state.acl.get("departments"):
            if sel_dept == "All":
                # union of all allowed positions across allowed departments
                union_positions = sorted({p for ps in st.session_state.acl["positions_by_dept"].values() for p in ps})
                pos_opts = union_positions
            else:
                pos_opts = st.session_state.acl["positions_by_dept"].get(sel_dept, [])
            sel_pos = st.selectbox(
                "Position",
                ["All"] + pos_opts,
                key="la_pos",
            )
        else:
            # Admin / Super User: original behavior
            pos_opts = (
                emp_df
                .query("department == @sel_dept")["role"]
                .dropna()
                .unique()
                if sel_dept != "All"
                else emp_df["role"].dropna().unique()
            )

            sel_pos = st.selectbox(
                "Position",
                ["All"] + sorted(pos_opts),
                key="la_pos"
            )
    # ─────────── 2. LIVE DATA HELPERS ────────────────────────────────────

    def get_actual_hours(dep, pos, start, end):
        q = (
            session.query(
                Actual.date.label("date"),
                func.sum(Actual.hours).label("hours"),
                func.sum(Actual.ot_hours).label("ot_hours"),
            )
            .join(Position, Position.id == Actual.position_id)
        )

        if dep != "All":
            q = q.join(
                Department, Department.id == Position.department_id
            ).filter(Department.name == dep)
        if pos != "All":
            q = q.filter(Position.name == pos)

        q = q.filter(Actual.date.between(start, end)).group_by(
            Actual.date
        )

        df = pd.read_sql(q.statement, session.bind)
        df["date"] = pd.to_datetime(df["date"])  # 🔧 Ensure datetime64[ns]

        all_days = pd.DataFrame({"date": pd.date_range(start, end)})
        all_days["date"] = pd.to_datetime(all_days["date"])  # 🔧 Ensure datetime64[ns]

        return all_days.merge(df, on="date", how="left").fillna(0)

    def get_standard_hours(dep, pos, start, end):
        q = (
            session.query(
                Schedule.day.label("date"),
                func.count().label("std_shift_count")  # count of scheduled shifts
            )
            .join(Employee, Employee.id == Schedule.emp_id)
            .join(Position, Position.name == Employee.role)
        )

        if dep != "All":
            q = (
                q.join(Department, Department.id == Position.department_id)
                 .filter(Department.name == dep)
            )
        if pos != "All":
            q = q.filter(Position.name == pos)

        q = q.filter(Schedule.day.between(start, end)).group_by(Schedule.day)

        df = pd.read_sql(q.statement, session.bind)
        df["date"] = pd.to_datetime(df["date"])
        df["standard_hours"] = df["std_shift_count"] * 8  # ← standard shift length
        df.drop(columns="std_shift_count", inplace=True)

        all_days = pd.DataFrame({"date": pd.date_range(start, end)})
        all_days["date"] = pd.to_datetime(all_days["date"])

        return all_days.merge(df, on="date", how="left").fillna(0)
    def get_schedule_hours(dep, pos, start, end):
        """
        Earned FTEs = one 8-hour FTE for every shift that appears
        on Schedule between `start` and `end`, filtered by Department / Position.
        """

        # ── base query: count shifts per calendar day ─────────────────────
        q = (
            session.query(
                Schedule.day.label("date"),
                func.count().label("shift_cnt")          # number of shifts that day
            )
            .join(Employee, Employee.id == Schedule.emp_id)
            .join(Position, Position.name == Employee.role)
        )

        if dep != "All":
            q = (
                q.join(Department, Department.id == Position.department_id)
                 .filter(Department.name == dep)
            )
        if pos != "All":
            q = q.filter(Position.name == pos)

        q = (
            q.filter(Schedule.day.between(start, end))
             .group_by(Schedule.day)
        )

        # rows → DataFrame
        df = pd.read_sql(q.statement, session.bind)
        df["date"] = pd.to_datetime(df["date"])

        # convert shift count → hours (8 h per shift) and FTEs
        df["sched_hours"] = df["shift_cnt"] * 8        # total hours scheduled
        df.drop(columns="shift_cnt", inplace=True)

        # ── ensure every day in the range exists ───────────────────────────
        all_days = pd.DataFrame({"date": pd.date_range(start, end)})
        all_days["date"] = pd.to_datetime(all_days["date"])

        return (
            all_days.merge(df, on="date", how="left")
                    .fillna(0)                         # days with no schedule → 0
        )
    def get_room_stats(metric, start, end):
        sub_act = (
            session.query(
                RoomActual.date,
                RoomActual.value.label("actual"),
            )
            .filter(RoomActual.kpi.ilike(metric))
            .filter(RoomActual.date.between(start, end))
            .subquery()
        )

        sub_fc = (
            session.query(
                RoomForecast.date,
                RoomForecast.value.label("forecast"),
            )
            .filter(RoomForecast.kpi.ilike(metric))
            .filter(RoomForecast.date.between(start, end))
            .subquery()
        )

        q = (
            session.query(
                sub_act.c.date.label("date"),
                sub_act.c.actual,
                sub_fc.c.forecast,
            )
            .select_from(
                sub_act.outerjoin(sub_fc, sub_act.c.date == sub_fc.c.date)
            )
            .order_by(sub_act.c.date)
        )

        df = pd.read_sql(q.statement, session.bind)
        df["date"] = pd.to_datetime(df["date"])  # 🔧 Ensure datetime64[ns]

        all_days = pd.DataFrame({"date": pd.date_range(start, end)})
        all_days["date"] = pd.to_datetime(all_days["date"])  # 🔧 Ensure datetime64[ns]

        return (
            all_days.merge(df, on="date", how="left")
            .fillna(0)
            .astype({"actual": int, "forecast": int})
        )    
# ─────────── 3. LOAD DATA ------------------------------------------------
    actual_df = get_actual_hours(sel_dept, sel_pos, dt_from, dt_to)
    sched_df  = get_schedule_hours(sel_dept, sel_pos, dt_from, dt_to)
    rooms_df  = get_room_stats("Occupied Rooms", dt_from, dt_to)
    std_df    = get_standard_hours(sel_dept, sel_pos, dt_from, dt_to)

    actual_df["date"] = pd.to_datetime(actual_df["date"])
    sched_df["date"]  = pd.to_datetime(sched_df["date"])
    rooms_df["date"]  = pd.to_datetime(rooms_df["date"])

    # --- merge actual + schedule -------------------------------------------------
    merged = actual_df.merge(sched_df, on="date", how="left").fillna(0)
    merged = merged.merge(std_df, on="date", how="left").fillna({"standard_hours": 0})

    # --- add Occupied-Rooms actuals  --------------------------------------------
    merged = merged.merge(
        rooms_df[["date", "actual"]].rename(columns={"actual": "occ_rooms"}),
        on="date", how="left"
    ).fillna({"occ_rooms": 0})

    # --- calculate standard productivity index (standard hours / occupied rooms)
    merged["standard_prod_idx"] = merged.apply(
        lambda r: r["standard_hours"] / r["occ_rooms"] if r["occ_rooms"] else 0,
        axis=1
    )

    # --- day-level KPIs ----------------------------------------------------------
    merged["actual_fte"] = (merged["hours"] + merged["ot_hours"]) / 8
    merged["sched_fte"]  = merged["sched_hours"] / 8
    merged["prod_idx"]   = merged.apply(
        lambda r: (r["actual_fte"] * 8) / r["occ_rooms"] if r["occ_rooms"] else 0,
        axis=1
    )

    # --- schedule productivity (sched hours / forecast rooms) --------------------
    merged = merged.merge(
        rooms_df[["date", "forecast"]].rename(columns={"forecast": "fc_rooms"}),
        on="date", how="left"
    ).fillna({"fc_rooms": 0})

    merged["sched_prod_idx"] = merged.apply(
        lambda r: r["sched_hours"] / r["fc_rooms"] if r["fc_rooms"] else 0,
        axis=1
    )
    # --- totals / period KPIs ----------------------------------------------------
    tot_act_fte = merged["actual_fte"].sum()
    tot_sch_fte = merged["sched_fte"].sum()
    tot_ot_pct  = (
        merged["ot_hours"].sum() / merged["hours"].sum() * 100
        if merged["hours"].sum() else 0
    )
    prod_index  = (tot_sch_fte / tot_act_fte * 100) if tot_act_fte else 0
    # ─────────── Standard Productivity (from LaborStandard table) ───────────
    std_df = refresh(db.LaborStandard)
    pos_df = refresh(db.Position)
    dept_df = refresh(db.Department)

    std_df = std_df.merge(pos_df.rename(columns={"id": "position_id", "name": "position", "department_id": "dept_id"}), on="position_id", how="left")
    std_df = std_df.merge(dept_df.rename(columns={"id": "dept_id", "name": "dept"}), on="dept_id", how="left")

    std_df = std_df[std_df["metric"] == "Occupied Rooms"]

    if sel_dept != "All":
        std_df = std_df[std_df["dept"] == sel_dept]
    if sel_pos != "All":
        std_df = std_df[std_df["position"] == sel_pos]

    raw_std = std_df["standard"].mean() if not std_df.empty else None
    standard_prod_val = (8 / raw_std) if raw_std else 0
    merged["standard_prod_idx"] = standard_prod_val  # assign same value across chart

    # ─────────── 4. KPI CARDS ------------------------------------------------
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Actual FTEs",  f"{tot_act_fte:.1f}")
    k2.metric("Earned FTEs",  f"{tot_sch_fte:.1f}",
              delta=f"{tot_sch_fte - tot_act_fte:+.1f}")
    k3.metric("Overtime %",   f"{tot_ot_pct:.1f} %")
    k4.metric("Prod Index",   f"{prod_index:.1f}")

    # ─────────── 5. CHARTS ROW ----------------------------------------------
    c_left, c_right = st.columns(2)
    is_single_day   = (dt_from == dt_to)

    # Clean label for x-axis
    merged["x_label"] = merged["date"].dt.strftime("%b %d")

    # ---------- FTE VARIANCE -------------------------------------------------
    with c_left:
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        if is_single_day:
            act_fte = merged["actual_fte"].iloc[0]
            sch_fte = merged["sched_fte"].iloc[0]
            fig.add_trace(go.Indicator(
                mode   = "number+delta",
                value  = act_fte,
                number = {
                    "valueformat": ".2f",
                    "font": {"size": 64, "color": "#1f77b4"},
                    "suffix": (
                        f"<span style='font-size:0.6em; color:#d62728'>"
                        f" / {sch_fte:.2f} Sch</span>"
                    )
                },
                delta  = {
                    "reference": sch_fte,
                    "relative" : False,
                    "position" : "bottom",
                    "increasing": {"color": "green"},
                    "decreasing": {"color": "red"}
                },
                title  = {"text": "<b>Actual FTEs vs Schedule FTEs</b>"},
                domain = {'x': [0, 1], 'y': [0, 1]}
            ))
            fig.update_layout(height=280, margin=dict(t=40, l=0, r=0, b=0))
        else:
            fig.add_trace(go.Scatter(
                x=merged["x_label"],
                y=merged["actual_fte"],
                name="Actual",
                mode="lines+markers",
                marker_color="#1f77b4",
                hovertemplate="Actual&nbsp;FTE: %{y:.2f}<extra></extra>"
            ))
            fig.add_trace(go.Scatter(
                x=merged["x_label"],
                y=merged["sched_fte"],
                name="Schedule",
                mode="lines+markers",
                marker_color="firebrick",
                line=dict(dash="dash"),
                hovertemplate="Sched&nbsp;FTE: %{y:.2f}<extra></extra>"
            ), secondary_y=True)

            fig.update_layout(
                title_text="FTE Variance",
                yaxis_title="FTEs",
                height=280,
                margin=dict(t=40, l=0, r=24, b=0),
                xaxis_type="category",
                hovermode="x unified",              # stacked hover tooltip

                legend=dict(
                    x=0.995,
                    y=1.15,
                    xanchor="right",
                    yanchor="top",
                    bgcolor="rgba(255,255,255,0.85)",
                    bordercolor="#d0d0d0",
                    borderwidth=1
                )
            )
        st.plotly_chart(fig, use_container_width=True)
        # ────────────── Totals Summary (FTEs with Variance) ──────────────
        total_actual_fte = merged["actual_fte"].sum()
        total_sched_fte  = merged["sched_fte"].sum()
        variance_fte     = total_sched_fte - total_actual_fte
        variance_arrow   = "▲" if variance_fte > 0 else ("▼" if variance_fte < 0 else "")

        st.markdown(
            f"""
            <div style='text-align: right; font-size: 16px; padding-top: 8px;'>
                <b>Total Actual FTEs:</b> {total_actual_fte:.2f} &nbsp;&nbsp;
                <b>Schedule:</b> {total_sched_fte:.2f} &nbsp;&nbsp;
                <b>Variance:</b> {variance_fte:+.2f} {variance_arrow}
            </div>
            """,
            unsafe_allow_html=True
        )

    # ────────────── Standard Productivity Index (Dashboard) ──────────────
    std_df = refresh(db.LaborStandard)
    pos_df = refresh(db.Position).rename(columns={"id": "position_id", "name": "position", "department_id": "dept_id"})
    dept_df = refresh(db.Department).rename(columns={"id": "dept_id", "name": "dept"})

    std_df = std_df.merge(pos_df, on="position_id", how="left")
    std_df = std_df.merge(dept_df, on="dept_id", how="left")

    std_df = std_df[std_df["metric"] == "Occupied Rooms"]
    if sel_dept != "All":
        std_df = std_df[std_df["dept"] == sel_dept]
    if sel_pos != "All":
        std_df = std_df[std_df["position"] == sel_pos]

    avg_standard = std_df["standard"].mean() if not std_df.empty else None

    total_output = rooms_df["actual"].sum()
    if total_output > 0 and avg_standard:
        standard_prod_val = 8 / avg_standard
    else:
        standard_prod_val = 0.0
    # ---------- PRODUCTIVITY ------------------------------------------------
    with c_right:
        if is_single_day:
            prod_val       = merged["prod_idx"].iloc[0]
            sched_prod_val = standard_prod_val  # from computed standard above

            fig = go.Figure(go.Indicator(
                mode   = "number+delta",
                value  = prod_val,
                number = {
                    "valueformat": ".2f",
                    "font"  : {"size": 64, "color": "#268bd2"},
                    "suffix": (
                        f"<span style='font-size:0.6em; color:#d62728'>"
                        f" / {sched_prod_val:.2f} Std</span>"
                    )
                },
                delta  = {
                    "reference"  : sched_prod_val,
                    "relative"   : False,
                    "position"   : "bottom",
                    "increasing" : {"color": "green"},
                    "decreasing" : {"color": "red"}
                },
                title  = {"text": "<b>Productivity (Hrs / Room)</b>"},
                domain = {'x': [0, 1], 'y': [0, 1]}
            ))
            fig.update_layout(height=280, margin=dict(t=40, l=0, r=0, b=0))

        else:
            fig = go.Figure()

            fig.add_trace(go.Scatter(
                x=merged["x_label"],
                y=merged["prod_idx"],
                mode="lines+markers",
                name="Actual",
                marker_color="#268bd2",
                hovertemplate="Actual&nbsp;Prod: %{y:.2f}<extra></extra>"
            ))

            fig.add_trace(go.Scatter(
                x=merged["x_label"],
                y=[standard_prod_val] * len(merged),
                mode="lines",
                name="Standard",
                marker_color="#d62728",
                line=dict(dash="dash"),
                hovertemplate="Standard&nbsp;Prod: %{y:.2f}<extra></extra>"
            ))

            fig.update_layout(
                title_text="Productivity (Hours / Room)",
                yaxis_title="Hrs / Room",
                height=280,
                margin=dict(t=40, l=0, r=24, b=0),
                xaxis_type="category",
                hovermode="x unified",

                legend=dict(
                    x=0.995,
                    y=1.15,
                    xanchor="right",
                    yanchor="top",
                    bgcolor="rgba(255,255,255,0.85)",
                    bordercolor="#d0d0d0",
                    borderwidth=1
                )
            )

        st.plotly_chart(fig, use_container_width=True)
        # ────────────── Totals Summary (Productivity with Variance) ──────────────
        total_hours = merged["hours"].sum() + merged["ot_hours"].sum()
        total_rooms = merged["occ_rooms"].sum()
        actual_prod_total = total_hours / total_rooms if total_rooms else 0
        prod_variance = standard_prod_val - actual_prod_total
        prod_arrow = "▲" if prod_variance > 0 else ("▼" if prod_variance < 0 else "")

        st.markdown(
            f"""
            <div style='text-align: right; font-size: 16px; padding-top: 8px;'>
                <b>Total Actual Prod:</b> {actual_prod_total:.2f} hrs/room &nbsp;&nbsp;
                <b>Standard:</b> {standard_prod_val:.2f} hrs/room &nbsp;&nbsp;
                <b>Variance:</b> {prod_variance:+.2f} {prod_arrow}
            </div>
            """,
            unsafe_allow_html=True
        )
    # ─────────── 7. DAILY ACTUAL vs FORECAST  (pick KPI) -----------------------
    kpi_options = ["Occupied Rooms", "Arrivals", "Departures", "Check-outs"]

    sel_kpi = st.selectbox("◔ KPI - DAILY ACTUAL vs FORECAST", kpi_options, key="kpi_daily")

    # fetch day-level data for the chosen KPI
    kpi_day_df = get_room_stats(sel_kpi, dt_from, dt_to).copy()
    kpi_day_df["x_label"] = kpi_day_df["date"].dt.strftime("%b %d")

    st.markdown(f"#### {sel_kpi} — Actual vs Forecast by Day")

    fig = go.Figure()

    fig.add_trace(go.Bar(
        x = kpi_day_df["x_label"],
        y = kpi_day_df["actual"],
        name = "Actual",
        marker_color = "#1f77b4",
        hovertemplate = "%{x}<br>Actual: %{y}<extra></extra>"
    ))

    fig.add_trace(go.Bar(
        x = kpi_day_df["x_label"],
        y = kpi_day_df["forecast"],
        name = "Forecast",
        marker_color = "#a0a0a0",
        hovertemplate = "%{x}<br>Forecast: %{y}<extra></extra>"
    ))

    fig.update_layout(
        barmode     = "group",
        height      = 300,
        margin      = dict(t=10, l=40, r=24, b=40),
        xaxis_title = "",
        yaxis_title = "Rooms",
        xaxis_type  = "category",
        xaxis_showgrid = False,
        yaxis       = dict(showgrid=True, gridcolor="#e0e0e0"),
        hovermode   = "x unified",

        legend=dict(
            x=0.995,
            y=1.15,
            xanchor="right",
            yanchor="top",
            bgcolor="rgba(255,255,255,0.85)",
            bordercolor="#d0d0d0",
            borderwidth=1
        )
    )

    st.plotly_chart(fig, use_container_width=True)

# ===================================================== EMPLOYEES ==============
elif main_choice == "Employees":
    st.markdown("""
    <div class="la-title">
      <!-- Lucide: users icon -->
      <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
           viewBox="0 0 24 24" fill="none" stroke="currentColor"
           stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
        <path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"></path>
        <circle cx="9" cy="7" r="4"></circle>
        <path d="M22 21v-2a4 4 0 0 0-3-3.87"></path>
        <path d="M16 3.13a4 4 0 0 1 0 7.75"></path>
      </svg>
      <span>Employees</span>
    </div>
    <style>
      .la-title{
        display:flex; align-items:center; gap:10px;
        margin:0 0 10px 0; line-height:1;
      }
      .la-title svg{ color:#111; }
      .la-title span{ font-weight:700; font-size:1.6rem; }
    </style>
    """, unsafe_allow_html=True)
    # figure out role once
    role = (st.session_state.get("user", {}).get("role") or "").strip().lower()

    # create tabs conditionally
    if role == "manager":
        tab2, tab3 = st.tabs(["➕\uFE0E Add New", "📆\uFE0E Schedule Availability"])
        has_tab1 = False
    else:
        tab1, tab2, tab3 = st.tabs(["📋\uFE0E View / Edit", "➕\uFE0E Add New", "📆\uFE0E Schedule Availability"])
        has_tab1 = True

    # ---------- 1) VIEW / EDIT -----------------------------------------------
    if has_tab1:
        with tab1:
            # ──────────────────────────────────────────────────────────────────
            # Toast helper + show any pending toast (replaces old success banners)
            def _render_toast(kind: str, text: str):
                """kind: 'success' | 'warning' | 'error' | 'info'"""
                palette = {
                    "success": "#4CAF50",
                    "warning": "#F59E0B",
                    "error"  : "#EF4444",
                    "info"   : "#3B82F6",
                }
                bg = palette.get(kind, "#3B82F6")
                st.markdown(f"""
                    <div style="
                        position: fixed;
                        bottom: 24px;
                        right: 24px;
                        background-color: {bg};
                        color: white;
                        padding: 14px 20px;
                        border-radius: 6px;
                        font-size: 15px;
                        box-shadow: 0 4px 10px rgba(0,0,0,0.25);
                        z-index: 1000;
                        animation: fadeOutToast 4s forwards;
                    ">
                        {text}
                    </div>
                    <style>
                        @keyframes fadeOutToast {{
                            0%   {{ opacity: 1; }}
                            80%  {{ opacity: 1; }}
                            100% {{ opacity: 0; display: none; }}
                        }}
                    </style>
                """, unsafe_allow_html=True)

            toast = st.session_state.pop("emp_toast", None)   # {'kind': 'success', 'msg': '...'}
            if toast:
                _render_toast(toast.get("kind", "info"), toast.get("msg", ""))

            # ▸─ Load employee data ───────────────────────────────────────────────
            emp_df = refresh(db.Employee)
            # 🔒 Enforce manager scope (Employees)
            emp_df = apply_manager_scope(
                emp_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})

            # ▸─ Filters (top row) ───────────────────────────────────────────────
            st.markdown("#### Filters")
            filter_col1, filter_col2, filter_col3 = st.columns(3)

            with filter_col1:
                depts = sorted(emp_df["department"].dropna().unique())
                sel_dept = st.selectbox("Department", ["(All)"] + depts)

            with filter_col2:
                if sel_dept != "(All)":
                    pos_opts = emp_df.loc[
                        emp_df["department"] == sel_dept, "role"
                    ].dropna().unique()
                    sel_pos = st.selectbox("Position", ["(All)"] + sorted(pos_opts))
                else:
                    sel_pos = "(All)"

            with filter_col3:
                emp_opts = ["(All)"] + sorted(emp_df["name"].dropna().unique())
                sel_emp = st.selectbox("🔍︎ Search employee", options=emp_opts, key="emp_selectbox")

            # ▸─ Apply filters ───────────────────────────────────────────────────
            filtered = emp_df.copy()
            if sel_dept != "(All)":
                filtered = filtered[filtered["department"] == sel_dept]
            if sel_pos != "(All)":
                filtered = filtered[filtered["role"] == sel_pos]
            if sel_emp != "(All)":
                filtered = filtered[filtered["name"] == sel_emp]

            # ── Deduplicate: keep highest hourly rate per name ──────────────────
            dedup = (
                filtered.sort_values("hourly_rate", ascending=False)
                        .drop_duplicates(subset="name", keep="first")
                        .reset_index(drop=True)
            )

            # ── Add "Type" column (In-House / Contract) ─────────────────────────
            if "emp_type" in dedup.columns:
                dedup["Type"] = np.where(
                    dedup["emp_type"].astype(str).str.strip().str.casefold() == "manual",
                    "Contract Labor",
                    "In-House"
                )
            else:
                dedup["Type"] = "In-House"

            # ── Split name to Last / First / ID ─────────────────────────────────
            parts = dedup["name"].str.extract(
                r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
            )
            for col in ["Last_Name", "First_Name", "ID"]:
                dedup[col] = parts[col]

            # ── Preferred column order ─────────────────────────────────────────
            display_cols = ["Last_Name", "First_Name", "ID", "Type"] + [
                c for c in dedup.columns
                if c not in ["id", "name", "Last_Name", "First_Name", "ID", "Type"]
            ] + ["id", "name"]

            # ── Build editable AgGrid ──────────────────────────────────────────
            gb_emp = GridOptionsBuilder.from_dataframe(dedup[display_cols])
            gb_emp.configure_default_column(editable=True, resizable=True)

            gb_emp.configure_column("Last_Name", header_name="Last Name", cellStyle={'textAlign': 'left'})
            gb_emp.configure_column("First_Name", header_name="First Name", cellStyle={'textAlign': 'left'})
            gb_emp.configure_column("ID", cellStyle={'textAlign': 'center'})
            gb_emp.configure_column("hourly_rate", cellStyle={'textAlign': 'center'})
            gb_emp.configure_column("department", cellStyle={'textAlign': 'left'})
            gb_emp.configure_column("role", cellStyle={'textAlign': 'left'})
            gb_emp.configure_column("Type", editable=False, cellStyle={'textAlign': 'center'})

            gb_emp.configure_column("id", hide=True)
            gb_emp.configure_column("name", hide=True)
            gb_emp.configure_column("hourly_rate", hide=True)

            gb_emp.configure_selection("multiple", use_checkbox=True)
            grid_opts = gb_emp.build()

            st.markdown("""
                <style>
                .ag-theme-streamlit .ag-root-wrapper {
                    border-radius: 12px !important;
                }
                </style>
            """, unsafe_allow_html=True)

            grid = AgGrid(
                dedup[display_cols],
                gridOptions=grid_opts,
                theme="streamlit",
                fit_columns_on_grid_load=True,
                domLayout='autoHeight'
            )


            # ───────────── Delete Logic ─────────────
            if st.button("🗑️ Delete Selected"):
                sel_raw = grid.get("selected_rows", [])
                sel_rows = (
                    sel_raw.to_dict("records")
                    if isinstance(sel_raw, pd.DataFrame) else list(sel_raw)
                )

                if not sel_rows:
                    st.session_state["emp_toast"] = {"kind": "warning", "msg": "⚠️ No employees selected for deletion."}
                    st.rerun()
                else:
                    names_to_delete = {
                        r["name"].strip()
                        for r in sel_rows
                        if isinstance(r, dict) and "name" in r and pd.notna(r["name"])
                    }
                    if names_to_delete:
                        deleted = (
                            session.query(db.Employee)
                                   .filter(db.Employee.name.in_(names_to_delete))
                                   .delete(synchronize_session=False)
                        )
                        session.commit()
                        st.session_state["emp_toast"] = {"kind": "success", "msg": f"🗑️ Deleted {deleted} record(s)."}
                        st.rerun()
                    else:
                        st.session_state["emp_toast"] = {"kind": "warning", "msg": "⚠️ Selected rows contained no valid names."}
                        st.rerun()
    # ---------- ONE-TIME PATCH: hide terminated employees globally ----------
    # Paste this near the TOP of your Employees page (AFTER st.header, BEFORE any refresh(db.Employee)).
    if not st.session_state.get("_refresh_emp_filtered", False):
        _orig_refresh = refresh  # keep a reference to the original

        def _refresh_patched(model):
            df = _orig_refresh(model)
            try:
                # If it's the Employee table, remove rows marked as 'terminated'
                if model is db.Employee and "emp_type" in df.columns:
                    mask = ~df["emp_type"].astype(str).str.strip().str.casefold().eq("terminated")
                    return df.loc[mask].reset_index(drop=True)
                return df
            except Exception:
                # If anything goes wrong, fall back to original df
                return df

        # Rebind globally so ALL pages now see Employees without 'terminated'
        globals()["refresh"] = _refresh_patched
        st.session_state["_refresh_emp_filtered"] = True


    # ---------- 2) ADD NEW ----------------------------------------------------
    with tab2:
        from sqlalchemy import func  # needed for func.lower in soft-terminate
        # ── Role & scope (for Manager restriction) ───────────────────────
        role_norm = (st.session_state.user.get("role") or "").strip().lower()
        user_scope = st.session_state.user.get("scope", [])

        st.subheader("Add Employee Manually")

        # ── persistent banner after a manual add ───────────────────────────
        if "emp_add_msg" in st.session_state:
            st.markdown(st.session_state.pop("emp_add_msg"), unsafe_allow_html=True)

        # ── Department / Position sources (dataframe style like your other page)
        dept_df = refresh(db.Department).rename(columns={"name": "dept"})

        # 🔒 Manager: restrict departments
        if role_norm == "manager" and user_scope:
            allowed_depts = {s["department"] for s in user_scope}
            dept_df = dept_df[dept_df["dept"].isin(allowed_depts)]
        pos_df = (
            refresh(db.Position)
            .merge(dept_df, left_on="department_id", right_on="id")
            [["id_x", "name", "dept"]]
            .rename(columns={"id_x": "id"})
        )

        # 🔒 Manager: restrict positions within allowed departments
        if role_norm == "manager" and user_scope:
            allowed_pairs = {
                (s["department"], s["position"]) for s in user_scope
            }
            pos_df = pos_df[
                pos_df.apply(
                    lambda r: (r["dept"], r["name"]) in allowed_pairs,
                    axis=1
                )
            ]

        depts = ["(Select)"] + sorted(dept_df["dept"].dropna().unique())

        # ⛳ Department SELECT OUTSIDE the form so Position list can refresh immediately
        st.markdown("**Department**")
        dept_sel = st.selectbox(" ", depts, key="add_emp_dept_select", label_visibility="collapsed")

        # Build position options based on the current department pick
        if dept_sel != "(Select)":
            pos_opts = sorted(pos_df.loc[pos_df["dept"] == dept_sel, "name"].dropna().unique())
        else:
            pos_opts = []

        if not depts or len(depts) == 1:
            st.warning("Create some departments and positions first in the Structure page.")
        else:
            with st.form("emp_form", clear_on_submit=True):
                col1, col2 = st.columns(2)

                # left-hand fields
                with col1:
                    first_name = st.text_input("First Name")
                    last_name  = st.text_input("Last Name")
                    emp_id     = st.text_input("ID #")
                    emp_type   = st.selectbox("Type", ["In-House", "Contract Labor"])

                # right-hand dropdowns + hourly rate
                with col2:
                    # show the already-chosen department (disabled for clarity)
                    st.text_input("Department", value=dept_sel, disabled=True)

                    # Position filtered by the selected department
                    if dept_sel == "(Select)":
                        st.info("Pick a Department above to see its positions.")
                        pos_sel = st.selectbox("Position / Role", ["(None)"])
                    else:
                        if not pos_opts:
                            st.info("No positions found in this department. Create positions in Structure → Positions.")
                            pos_sel = st.selectbox("Position / Role", ["(None)"])
                        else:
                            pos_sel = st.selectbox("Position / Role", pos_opts, key="add_emp_pos")

                    rate = st.number_input("Hourly Rate", min_value=0.0)

                # submit with validation
                submitted = st.form_submit_button("Add Employee")
                if submitted:
                    missing = []
                    if not str(first_name).strip(): missing.append("First Name")
                    if not str(last_name).strip():  missing.append("Last Name")
                    if not str(emp_id).strip():     missing.append("ID #")
                    if dept_sel == "(Select)":      missing.append("Department")
                    if pos_sel in (None, "", "(None)"): missing.append("Position / Role")

                    if missing:
                        st.error("Please complete: " + ", ".join(missing))
                    else:
                        full_name = f"{last_name.strip()}, {first_name.strip()} {emp_id.strip()}"
                        source_val = "manual" if emp_type == "Contract Labor" else "import"

                        session.merge(
                            db.Employee(
                                name        = full_name,
                                role        = pos_sel,
                                department  = dept_sel,
                                hourly_rate = rate,
                                emp_type    = source_val
                            )
                        )
                        session.commit()

                        # ✅ Toast-style popup
                        st.markdown(f"""
                            <div style="
                                position: fixed;
                                bottom: 24px;
                                right: 24px;
                                background-color: #4CAF50;
                                color: white;
                                padding: 14px 20px;
                                border-radius: 6px;
                                font-size: 15px;
                                box-shadow: 0px 4px 10px rgba(0,0,0,0.25);
                                z-index: 1000;
                                animation: fadeOut 5s forwards;
                            ">
                                ✅ Employee <b>{first_name.strip()} {last_name.strip()}</b> added successfully.
                            </div>
                            <style>
                                @keyframes fadeOut {{
                                    0%   {{ opacity: 1; }}
                                    80%  {{ opacity: 1; }}
                                    100% {{ opacity: 0; display: none; }}
                                }}
                            </style>
                        """, unsafe_allow_html=True)
        # ------------- BULK UPLOAD ------------------------------------------
        # ------------- BULK UPLOAD ------------------------------------------
        if role_norm != "manager":

            st.subheader("📥 Bulk Upload Employees")
            csv_up = st.file_uploader(
                "Upload CSV/XLSX with columns "
                "'Location Name', 'Employee', 'Job Name', 'Base Salary', 'Base Rate'",
                type=["csv", "xls", "xlsx"]
            )
            if csv_up:
                # Read file
                df_up = (pd.read_excel(csv_up)
                         if csv_up.name.lower().endswith((".xls", ".xlsx"))
                         else pd.read_csv(csv_up))

                required = {"Location Name", "Employee", "Job Name", "Base Rate"}
                if not required.issubset(df_up.columns):
                    st.error(f"Missing columns: {required - set(df_up.columns)}")
                else:
                    # normalize
                    df_up["Location Name"] = df_up["Location Name"].astype(str)
                    df_up["Employee"]      = df_up["Employee"].astype(str)
                    df_up["Job Name"]      = df_up["Job Name"].astype(str)
                    df_up["Base Rate"]     = pd.to_numeric(df_up["Base Rate"], errors="coerce").fillna(0.0)

                    df_up["loc_norm"] = df_up["Location Name"].str.strip().str.casefold()
                    df_up["job_norm"] = df_up["Job Name"].str.strip().str.casefold()

                    # ---------- Departments -------------------------------------
                    exist_depts_norm = {d.name.strip().casefold() for d in session.query(db.Department).all()}
                    new_dept_rows = (
                        df_up.loc[~df_up["loc_norm"].isin(exist_depts_norm), ["Location Name", "loc_norm"]]
                            .drop_duplicates("loc_norm")
                    )
                    for _, row in new_dept_rows.iterrows():
                        session.add(db.Department(name=row["Location Name"].strip()))
                    session.commit()

                    dept_map = {d.name.strip().casefold(): d.id for d in session.query(db.Department).all()}

                    # ---------- Positions ---------------------------------------
                    exist_pos_keys = {(p.name.strip().casefold(), p.department_id)
                                      for p in session.query(db.Position).all()}
                    new_pos_added = 0
                    for _, row in df_up.drop_duplicates(subset=["job_norm", "loc_norm"]).iterrows():
                        dept_id = dept_map.get(row["loc_norm"])
                        key = (row["job_norm"], dept_id)
                        if key not in exist_pos_keys and dept_id is not None:
                            session.add(db.Position(name=row["Job Name"].strip(), department_id=dept_id))
                            new_pos_added += 1
                    session.commit()

                    # ---------- Employees (upsert + reactivate) ------------------
                    added = updated = skipped = 0
                    for _, r in df_up.iterrows():
                        full_name = r["Employee"].strip()
                        dept_norm = r["Location Name"].strip().casefold()
                        rate_val  = float(r["Base Rate"])

                        dept_id = dept_map.get(dept_norm)
                        if dept_id is None:
                            continue

                        pos = (session.query(db.Position)
                               .filter(
                                   db.Position.name.ilike(r["Job Name"].strip()),
                                   db.Position.department_id == dept_id
                               )
                               .first())
                        if not pos:
                            continue

                        emp = session.query(db.Employee).filter_by(name=full_name).first()

                        if emp:
                            if str(emp.emp_type or "").strip().casefold() == "terminated":
                                emp.emp_type = "import"
                            emp.role        = pos.name
                            emp.department  = r["Location Name"].strip()
                            emp.hourly_rate = rate_val
                            updated += 1
                        else:
                            session.add(
                                db.Employee(
                                    name        = full_name,
                                    role        = pos.name,
                                    department  = r["Location Name"].strip(),
                                    hourly_rate = rate_val,
                                    emp_type    = "import"
                                )
                            )
                            added += 1
                    session.commit()

                    # ---------- SOFT-TERMINATE -----------------------------------
                    names_in_upload_norm = {str(n).strip().lower() for n in df_up["Employee"].dropna()}
                    terminated = (
                        session.query(db.Employee)
                               .filter(
                                   db.Employee.emp_type == "import",
                                   ~func.lower(db.Employee.name).in_(names_in_upload_norm)
                               )
                               .update({db.Employee.emp_type: "terminated"}, synchronize_session=False)
                    )
                    session.commit()

                    st.markdown(f"""
                        <div style="
                            position: fixed;
                            bottom: 24px;
                            right: 24px;
                            background-color: #4CAF50;
                            color: white;
                            padding: 14px 20px;
                            border-radius: 6px;
                            font-size: 15px;
                            box-shadow: 0px 4px 10px rgba(0,0,0,0.25);
                            z-index: 1000;
                            animation: fadeOut 6s forwards;
                        ">
                            ✅ Upload complete: <b>{added}</b> added · <b>{updated}</b> updated · <b>{terminated}</b> terminated.
                        </div>
                        <style>
                            @keyframes fadeOut {{
                                0%   {{ opacity: 1; }}
                                80%  {{ opacity: 1; }}
                                100% {{ opacity: 0; display: none; }}
                            }}
                        </style>
                    """, unsafe_allow_html=True)

        else:
            st.info("You have view-only access. Bulk upload is disabled for Managers.")

    # ---------- 3) SCHEDULE AVAILABILITY --------------------------------------
    with tab3:
        st.subheader("Weekly Schedule Availability")

        emp_df = refresh(db.Employee)

        # 🔒 Enforce manager/employee scope (Schedule Availability)
        _role_norm = (st.session_state.user.get("role") or "").strip().lower()
        if _role_norm in ("manager", "employee"):
            emp_df = apply_manager_scope(
                emp_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})

        # ── All existing availability rows (dummy week only) ───────────────
        dummy_dates = list(WK_TO_DATE.values())      # 2000-01-03 … 2000-01-09
        sched_rows = session.query(db.Schedule).filter(
            db.Schedule.day.in_(dummy_dates)         # ← filter here
        ).all()

        # ── Filter sidebar ────────────────────────────────────────────────
        col_left, col_main = st.columns([1, 3])

        with col_left:
            st.markdown("#### Filter")
            depts = sorted(emp_df["department"].dropna().unique())
            sel_dept = st.selectbox("Department", ["(All)"] + depts, key="view_dept")

            if sel_dept != "(All)":
                pos_opts = emp_df[emp_df["department"] == sel_dept]["role"].dropna().unique()
                sel_pos  = st.selectbox("Position", ["(All)"] + sorted(pos_opts))
            else:
                sel_pos = "(All)"

            emp_opts = emp_df["name"].dropna().unique()
            sel_emp = st.selectbox("Employee", ["(All)"] + sorted(emp_opts))

        # ── Apply filters ─────────────────────────────────────────────────
        filtered = emp_df.copy()
        if sel_dept != "(All)":
            filtered = filtered[filtered["department"] == sel_dept]
        if sel_pos != "(All)":
            filtered = filtered[filtered["role"] == sel_pos]
        if sel_emp != "(All)":
            filtered = filtered[filtered["name"] == sel_emp]

        filtered = filtered.drop_duplicates(subset=["name"])  # one row per employee

        # ── Build editable availability table ─────────────────────────────
        import re
        from datetime import date, datetime
        days          = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        shift_options = ["MORNING", "AFTERNOON", "EVENING", "OPEN", "OFF"]

        data = []
        pattern = re.compile(r"^\s*(?P<last>[^,]+),\s*(?P<first>[^\d]+?)\s+(?P<id>\d+)")

        for _, row in filtered.iterrows():
            emp_id   = row["id"]
            raw_name = row["name"]

            m = pattern.match(raw_name)
            if m:
                last_name  = m.group("last").strip()
                first_name = m.group("first").strip()
                id_str     = m.group("id").strip()
            else:
                last_name, first_name, id_str = raw_name, "", ""

            # Existing availability (dummy-week rows only)
            shifts = {}
            for s in sched_rows:
                if s.emp_id == emp_id:
                    lbl = s.day.strftime("%a")  # 'Mon', 'Tue', …
                    shifts[lbl] = s.shift_type

            row_data = {
                "ID":         id_str,
                "First Name": first_name,
                "Last Name":  last_name,
                "emp_id":     emp_id
            }
            for d in days:
                row_data[d] = shifts.get(d, "OPEN")
            data.append(row_data)

        schedule_edit_df = pd.DataFrame(data)

        # ── Configure AgGrid ─────────────────────────────────────────────
        gb = GridOptionsBuilder.from_dataframe(schedule_edit_df)
        gb.configure_default_column(editable=True)

        gb.configure_column("ID",         width=160, editable=False)
        gb.configure_column("First Name", width=160, editable=False)
        gb.configure_column("Last Name",  width=160, editable=False)

        for d in days:
            gb.configure_column(
                d,
                cellEditor="agSelectCellEditor",
                cellEditorParams={"values": shift_options},
                width=160
            )

        gb.configure_column("emp_id", hide=True)
        gb.configure_selection("multiple", use_checkbox=False)

        grid = AgGrid(
            schedule_edit_df,
            gridOptions=gb.build(),
            height=400,
            theme="balham"
        )

        # ── Save button ──────────────────────────────────────────────────
        if st.button("💾 Save"):
            edited = pd.DataFrame(grid["data"])

            # Delete existing dummy-week rows for these employees
            session.query(db.Schedule).filter(
                db.Schedule.emp_id.in_(edited["emp_id"].tolist()),
                db.Schedule.day.in_(dummy_dates)
            ).delete(synchronize_session=False)

            # Re-insert availability rows (always dummy-week dates)
            new_rows = []
            for _, r in edited.iterrows():
                emp = int(r["emp_id"])
                for lbl in days:
                    new_rows.append(
                        db.Schedule(
                            emp_id     = emp,
                            day        = WK_TO_DATE[lbl],          # already a date obj
                            shift_type = str(r[lbl]).strip().upper()
                        )
                    )

            session.add_all(new_rows)
            session.commit()
            st.success("✅ Weekly availability saved.")
# =============================================== LABOR ▸ STRUCTURE ============
elif main_choice == "Labor ▸ Structure":
    st.markdown("""
    <div class="la-title">
      <!-- Lucide: layers icon (monochrome) -->
      <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
           viewBox="0 0 24 24" fill="none" stroke="currentColor"
           stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
        <path d="M12 2 2 7l10 5 10-5-10-5Z"></path>
        <path d="m2 17 10 5 10-5"></path>
        <path d="m2 12 10 5 10-5"></path>
      </svg>
      <span>Labor Structure</span>
    </div>
    <style>
      .la-title{
        display:flex; align-items:center; gap:10px;
        margin:0 0 10px 0; line-height:1;
      }
      .la-title svg{ color:#111; }
      .la-title span{ font-weight:700; font-size:1.6rem; }
    </style>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📋\uFE0E View / Edit", "➕\uFE0E Add New", "🕒\uFE0E Labor Shifts", "📏\uFE0E Labor Standards", "KPIs"])

    # ---------- 5) KPIs ------------------------------------------------------
    with tab5:
        # -- DB session (do not print it)
        session = get_scoped_session()

        # -- Resolve current hotel (optional scoping)
        hotel = None
        try:
            hotel = (st.session_state.get("user") or {}).get("hotel_name")
        except Exception:
            hotel = None
        if hotel is None:
            try:
                from db import current_hotel_context
                hotel = current_hotel_context.get("hotel_name")
            except Exception:
                pass

        st.markdown("### KPI Management")

        # If the model isn't present yet, show a friendly message
        if not hasattr(db, "RoomKPI"):
            st.warning("`RoomKPI` model/table not found. Create it in `db.py` first, then refresh.")
            st.stop()

        # -------------------- Canonicalization Helpers --------------------
        import re, unicodedata
        _ref_re = re.compile(r"\[([^\]]+)\]")

        def _canon(s: str) -> str:
            s = unicodedata.normalize("NFKC", str(s or ""))
            s = s.replace("\u00A0", " ")
            s = re.sub(r"\s+", " ", s).strip()
            return s.lower()

        def _clean_expr(expr: str) -> str:
            expr = unicodedata.normalize("NFKC", str(expr or ""))
            expr = expr.replace("\u00A0", " ")
            return re.sub(r"\s+", " ", expr).strip()

        def _rewrite_to_official(expr: str, canon_to_official: dict) -> str:
            def _sub(m):
                typed = m.group(1)
                c = _canon(typed)
                off = canon_to_official.get(c, typed)
                return f"[{off}]"
            return _ref_re.sub(_sub, expr or "")

        # -------------------- Create KPI form --------------------
        with st.container():
            col_a, col_b = st.columns([3, 1])
            with col_a:
                new_name = st.text_input("KPI Name", placeholder="e.g., Studio Stayover", key="kpi_new_name")
            with col_b:
                new_sort = st.number_input("Sort Order", min_value=0, value=100, step=1, key="kpi_new_sort")

            col1, col2 = st.columns([1, 3])
            with col1:
                new_has_rule = st.checkbox("Has Rule?", key="kpi_new_has_rule")
                new_active   = st.checkbox("Active", value=True, key="kpi_new_active")
            with col2:
                rule_help = "Build the rule by inserting KPIs; brackets are added automatically. Example: [Studio Occupied Rooms] - [Studio Arrivals]"
                st.caption(rule_help)

                if "kpi_rule" not in st.session_state:
                    st.session_state.kpi_rule = ""

                rule_txt = st.text_area(
                    "Rule Expression",
                    value=st.session_state.kpi_rule,
                    placeholder="Use the KPI picker below to insert names (auto adds [brackets])",
                    disabled=not new_has_rule,
                    help=rule_help,
                    key="kpi_rule_widget"
                )
                st.session_state.kpi_rule = rule_txt

                try:
                    existing_kpis = [
                        r.name for r in session.query(db.RoomKPI)
                        .filter(db.RoomKPI.is_active == True)
                        .order_by(db.RoomKPI.name.asc()).all()
                    ]
                except Exception:
                    existing_kpis = []

                def _insert_selected_kpi():
                    sel = st.session_state.get("kpi_insert_select_simple")
                    if sel and sel != "(Select)":
                        st.session_state.kpi_rule = (st.session_state.get("kpi_rule", "") + f" [{sel}]").strip()

                if new_has_rule and existing_kpis:
                    st.selectbox(
                        "Insert KPI",
                        ["(Select)"] + existing_kpis,
                        key="kpi_insert_select_simple",
                        on_change=_insert_selected_kpi
                    )

            create_clicked = st.button("➕ Create KPI", key="kpi_create_btn")

        # -------------------- Create handler --------------------
        if create_clicked:
            name = (new_name or "").strip()
            if not name:
                st.error("KPI Name is required.")
                st.stop()

            q_dup = session.query(db.RoomKPI).filter(db.RoomKPI.name == name)
            if hasattr(db.RoomKPI, "hotel_name") and hotel:
                q_dup = q_dup.filter(db.RoomKPI.hotel_name == hotel)
            dup = q_dup.first()
            if dup:
                st.error("A KPI with this name already exists.")
                st.stop()

            rule_text = (st.session_state.get("kpi_rule") or "").strip()

            if new_has_rule and not rule_text:
                st.error("Rule Expression is required when 'Has Rule' is checked.")
                st.stop()

            if new_has_rule and f"[{name}]" in rule_text:
                st.error("A KPI rule cannot reference itself.")
                st.stop()

            # sanitize rule text
            rule_text = _clean_expr(rule_text)

            kpi = db.RoomKPI(
                name=name,
                has_rule=bool(new_has_rule),
                rule_expr=rule_text if new_has_rule else None,
                sort_order=int(new_sort or 100),
                is_active=bool(new_active),
                hotel_name=hotel
            )
            session.add(kpi)
            session.commit()

            st.session_state.kpi_rule = ""
            st.rerun()

        # -------------------- Existing KPIs table (inline edit) --------------------
        q = session.query(db.RoomKPI)
        if hasattr(db.RoomKPI, "hotel_name") and hotel:
            q = q.filter(db.RoomKPI.hotel_name == hotel)
        rows = q.order_by(db.RoomKPI.sort_order.asc(), db.RoomKPI.name.asc()).all()

        import pandas as pd
        if rows:
            df = pd.DataFrame([{
                "id": r.id,
                "Name": r.name,
                "Has Rule": bool(r.has_rule),
                "Rule Expression": r.rule_expr or "",
                "Active": bool(r.is_active),
                "Sort Order": int(r.sort_order or 100),
            } for r in rows])
        else:
            df = pd.DataFrame(columns=["id", "Name", "Has Rule", "Rule Expression", "Active", "Sort Order"])

        st.markdown("#### Existing KPIs")
        edited = st.data_editor(
            df.drop(columns=["id"]),
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            column_config={
                "Name": st.column_config.TextColumn("Name"),
                "Has Rule": st.column_config.CheckboxColumn("Has Rule"),
                "Rule Expression": st.column_config.TextColumn("Rule Expression", help="Use [KPI Name] placeholders"),
                "Active": st.column_config.CheckboxColumn("Active"),
                "Sort Order": st.column_config.NumberColumn("Sort Order", min_value=0, step=1),
            },
            key="kpi_editor"
        )

        # -------------------- Save edited KPIs --------------------
        if st.button("💾 Save KPI Changes", key="kpi_save_btn"):
            if len(edited) != len(rows):
                st.error("Unexpected row mismatch. Please refresh and try again.")
                st.stop()

            import re

            def referenced(expr):
                return set(re.findall(r"\[([^\]]+)\]", expr or ""))

            all_names = { (edited.iloc[i]["Name"] or "").strip() for i in range(len(edited)) }
            canon_to_official = { _canon(n): n for n in all_names }

            for i, r in enumerate(rows):
                nm = (edited.iloc[i]["Name"] or "").strip()
                if not nm:
                    st.error("KPI Name cannot be empty.")
                    st.stop()
                if bool(edited.iloc[i]["Has Rule"]):
                    expr = _clean_expr((edited.iloc[i]["Rule Expression"] or "").strip())
                    if not expr:
                        st.error(f"Rule required for KPI '{nm}'.")
                        st.stop()
                    if f"[{nm}]" in expr:
                        st.error(f"KPI '{nm}' cannot reference itself.")
                        st.stop()
                    missing = [x for x in referenced(expr) if _canon(x) not in canon_to_official]
                    if missing:
                        st.error(f"KPI '{nm}' references unknown KPI(s): {', '.join(missing)}.")
                        st.stop()
                    # rewrite to use canonical official names
                    expr = _rewrite_to_official(expr, canon_to_official)
                    edited.at[i, "Rule Expression"] = expr

            for i, r in enumerate(rows):
                r.name       = (edited.iloc[i]["Name"] or "").strip()
                r.has_rule   = bool(edited.iloc[i]["Has Rule"])
                expr         = (edited.iloc[i]["Rule Expression"] or "").strip()
                r.rule_expr  = expr if r.has_rule else None
                r.is_active  = bool(edited.iloc[i]["Active"])
                r.sort_order = int(edited.iloc[i]["Sort Order"] or 100)

            session.commit()
            st.success("KPI changes saved and canonicalized.")
            st.rerun()

        # -------------------- Delete KPI --------------------
        st.markdown("#### 🗑️ Delete a KPI")

        dq = session.query(db.RoomKPI)
        if hasattr(db.RoomKPI, "hotel_name") and hotel:
            dq = dq.filter(db.RoomKPI.hotel_name == hotel)
        drows = dq.order_by(db.RoomKPI.name.asc()).all()

        if drows:
            del_options = [(r.name, r.id) for r in drows]
            label_to_id = {label: _id for label, _id in del_options}

            col_a, col_b = st.columns([3, 1])
            with col_a:
                sel_label = st.selectbox(
                    "Select KPI to delete",
                    ["(Select)"] + [label for label, _ in del_options],
                    key="kpi_delete_select"
                )
            with col_b:
                confirm = st.checkbox("Confirm", key="kpi_delete_confirm")

            if st.button("🗑️ Delete", key="kpi_delete_btn"):
                if sel_label and sel_label != "(Select)" and confirm:
                    sel_id = label_to_id[sel_label]
                    session.query(db.RoomKPI).filter(db.RoomKPI.id == sel_id).delete()
                    session.commit()
                    st.success(f"KPI '{sel_label}' deleted.")
                    st.rerun()
                else:
                    st.info("Select a KPI and check **Confirm** before deleting.")
        else:
            st.info("No KPIs available to delete.")
        st.caption("Rule KPIs are hidden from Room STATs manual entry and are meant to be auto-calculated.")
    # ---------- 1) VIEW / EDIT -----------------------------------------------
    with tab1:

        #–– PERSISTENT MESSAGE ───────────────────────────────────────────────
        if "del_msg" in st.session_state:
            st.success(st.session_state.pop("del_msg"))

        st.markdown("""
            <style>
            .ag-theme-streamlit .ag-root-wrapper {
                border-radius: 12px !important;
            }
            </style>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns([1, 2])

        with col1:
            st.subheader("Departments")
            dept_df = refresh(db.Department)
            gb_dept = GridOptionsBuilder.from_dataframe(dept_df)
            gb_dept.configure_default_column(editable=True, resizable=True)
            gb_dept.configure_selection("multiple", use_checkbox=True)

            dept_grid = AgGrid(
                dept_df,
                gridOptions=gb_dept.build(),
                theme="streamlit",
                fit_columns_on_grid_load=True,
                height=350
            )
            sel_depts = dept_grid.get("selected_rows", [])

        with col2:
            st.subheader("Positions")
            pos_df = refresh(db.Position).merge(
                dept_df, left_on="department_id", right_on="id",
                suffixes=("", "_dept"))
            gb_pos = GridOptionsBuilder.from_dataframe(pos_df)
            gb_pos.configure_default_column(editable=True, resizable=True)
            gb_pos.configure_selection("multiple", use_checkbox=True)

            pos_grid = AgGrid(
                pos_df,
                gridOptions=gb_pos.build(),
                theme="streamlit",
                fit_columns_on_grid_load=True,
                height=350
            )
            sel_pos = pos_grid.get("selected_rows", [])

        col_a, col_b = st.columns(2)
        if col_a.button("💾 Save"):
            # save department edits
            for row in dept_grid["data"]:
                session.query(db.Department).filter_by(id=row["id"]).update({"name": row["name"]})
            # save position edits
            for row in pos_grid["data"]:
                session.query(db.Position).filter_by(id=row["id"]).update({"name": row["name"]})
            session.commit()
            st.success("Changes saved.")
            st.rerun()

        if col_b.button("🗑️ Delete Selected"):
            # ---- helper to ensure we always iterate over a list ------------
            def _rows(sel):
                if sel is None:
                    return []
                if isinstance(sel, list):
                    return sel
                if isinstance(sel, pd.DataFrame):
                    return sel.to_dict("records")
                return []

            pos_rows  = _rows(sel_pos)
            dept_rows = _rows(sel_depts)

            # ---- delete positions first -----------------------------------
            for r in pos_rows:
                rec_id = r["id"] if isinstance(r, dict) and "id" in r else r
                try:
                    rec_id = int(rec_id)
                    session.query(db.Position).filter_by(id=rec_id).delete()
                except Exception:
                    pass  # skip invalid ids

            # ---- then delete departments (skip if referenced) -------------
            for r in dept_rows:
                rec_id  = r["id"] if isinstance(r, dict) and "id" in r else r
                r_name  = r.get("name", str(r)) if isinstance(r, dict) else str(r)
                try:
                    rec_id = int(rec_id)
                    session.query(db.Department).filter_by(id=rec_id).delete()
                except Exception:
                    st.warning(f"Dept '{r_name}' not deleted (still referenced).")

            session.commit()
            #–– SAVE BANNER ─────────────────────────────────────────────────
            st.session_state["del_msg"] = (
                f"Deleted {len(pos_rows)} positions and {len(dept_rows)} departments."
            )
            st.rerun()

    # ---------- 2) ADD NEW ----------------------------------------------------
    with tab2:
        st.subheader("Add Department")
        with st.form("dept_form", clear_on_submit=True):
            new_dept = st.text_input("Department name")
            if st.form_submit_button("Add Department") and new_dept:
                session.merge(db.Department(name=new_dept))
                session.commit()
                st.success("Department added!")
                st.rerun()

        st.subheader("Add Position")
        depts = refresh(db.Department)
        if depts.empty:
            st.warning("Create a department first.")
        else:
            with st.form("pos_form", clear_on_submit=True):
                pos_name = st.text_input("Position name")
                dept_sel = st.selectbox("Department", depts["name"])
                if st.form_submit_button("Add Position") and pos_name:
                    dept_id = int(depts.set_index("name").loc[dept_sel, "id"])
                    session.merge(db.Position(name=pos_name, department_id=dept_id))
                    session.commit()
                    st.success("Position added!")
                    st.rerun()

        # Bulk upload (structure)  – unchanged
        st.subheader("📥 Bulk Upload (Location Name → Department, Job Name → Position)")
        csv_file = st.file_uploader(
            "Upload CSV/XLSX with columns 'Location Name' and 'Job Name'",
            type=["csv", "xls", "xlsx"],
            key="struct_upl"
        )
        if csv_file:
            df_up = (pd.read_excel(csv_file)
                     if csv_file.name.endswith((".xls", ".xlsx"))
                     else pd.read_csv(csv_file))

            if not {"Location Name", "Job Name"}.issubset(df_up.columns):
                st.error("Required columns not found.")
            else:
                df_up["loc_norm"] = df_up["Location Name"].astype(str).str.strip().str.casefold()
                df_up["job_norm"] = df_up["Job Name"].astype(str).str.strip().str.casefold()

                exist_depts_norm = {
                    d.name.strip().casefold()
                    for d in session.query(db.Department).all()
                }
                new_dept_rows = (
                    df_up.loc[~df_up["loc_norm"].isin(exist_depts_norm),
                              ["Location Name", "loc_norm"]]
                    .drop_duplicates("loc_norm")
                )
                for _, row in new_dept_rows.iterrows():
                    session.add(db.Department(name=row["Location Name"].strip()))
                session.commit()

                dept_map = {
                    d.name.strip().casefold(): d.id
                    for d in session.query(db.Department).all()
                }

                exist_pos_keys = {
                    (p.name.strip().casefold(), p.department_id)
                    for p in session.query(db.Position).all()
                }
                added = 0
                for _, row in (
                    df_up[["Job Name", "job_norm", "loc_norm"]]
                    .dropna()
                    .drop_duplicates()
                ).iterrows():
                    dept_id = dept_map.get(row["loc_norm"])
                    key = (row["job_norm"], dept_id)
                    if key not in exist_pos_keys:
                        session.add(
                            db.Position(
                                name=row["Job Name"].strip(),
                                department_id=dept_id
                            )
                        )
                        added += 1
                session.commit()
                st.success(
                    f"Imported {len(new_dept_rows)} new departments and {added} new positions."
                )

    # ---------- 3) LABOR SHIFTS -------------------------------------------
    with tab3:
        st.subheader("Define Shift Times by Position")

        # 1. Fetch departments / positions
        depts = sorted({d.name for d in session.query(db.Department).all()})
        dept_sel = st.selectbox("Department", ["(All)"] + depts)

        if dept_sel != "(All)":
            dept_positions = (
                session.query(db.Position)
                       .join(db.Department, db.Position.department_id == db.Department.id)
                       .filter(db.Department.name == dept_sel)
                       .order_by(db.Position.name)
                       .all()
            )
            pos_names = [p.name for p in dept_positions]
            sel_pos = st.selectbox("Position", ["(All)"] + pos_names)
        else:
            sel_pos = "(All)"

        # 👉  Save button — at the top
        save_clicked = st.button("💾 Save All Shift Times", key="save_shifts")

        # 2. Filter positions
        pos_query = session.query(db.Position)
        if dept_sel != "(All)":
            dept_id = session.query(db.Department).filter_by(name=dept_sel).first().id
            pos_query = pos_query.filter_by(department_id=dept_id)
        if sel_pos != "(All)":
            pos_query = pos_query.filter_by(name=sel_pos)

        positions = pos_query.all()
        shift_periods = ["Morning", "Afternoon", "Evening"]

        # 3. UI to edit shifts
        for pos in positions:
            st.markdown(f"#### {pos.name}")

            for period in shift_periods:
                existing = (
                    session.query(db.ShiftTime)
                    .filter_by(position_id=pos.id, period=period)
                    .order_by(db.ShiftTime.id)
                    .all()
                )

                with st.expander(f"🕒 {period} shifts for {pos.name}", expanded=bool(existing)):
                    existing_cnt = len(existing)
                    rows = st.number_input(
                        f"# of {period} shifts",
                        min_value=0, max_value=30,
                        value=existing_cnt,
                        key=f"{pos.id}_{period}_count"
                    )

                    for i in range(rows):
                        c1, c2 = st.columns(2)
                        def_time = lambda obj, field: getattr(obj, field) if obj else None
                        start = c1.time_input(
                            "Start",
                            value=def_time(existing[i] if i < existing_cnt else None, "start"),
                            step=1800,
                            key=f"{pos.id}_{period}_start_{i}"
                        )
                        end = c2.time_input(
                            "End",
                            value=def_time(existing[i] if i < existing_cnt else None, "end"),
                            step=1800,
                            key=f"{pos.id}_{period}_end_{i}"
                        )

        # 4. Persist when Save pressed
        if save_clicked:
            # clear only the shifts for displayed positions
            session.query(db.ShiftTime).filter(
                db.ShiftTime.position_id.in_([p.id for p in positions])
            ).delete(synchronize_session=False)

            for pos in positions:
                for period in shift_periods:
                    rows = st.session_state.get(f"{pos.id}_{period}_count", 0)
                    for i in range(rows):
                        start = st.session_state.get(f"{pos.id}_{period}_start_{i}")
                        end = st.session_state.get(f"{pos.id}_{period}_end_{i}")
                        if start and end:
                            session.add(
                                db.ShiftTime(
                                    position_id=pos.id,
                                    period=period,
                                    start=start,
                                    end=end
                                )
                            )
            session.commit()
            st.success("✅ Shifts saved successfully.")

    # ---------- 4) LABOR STANDARDS ------------------------------------------
    with tab4:
        # -- DB session (do not print it)
        session = get_scoped_session()

        # -- Resolve current hotel (optional scoping)
        hotel = None
        try:
            hotel = (st.session_state.get("user") or {}).get("hotel_name")
        except Exception:
            hotel = None
        if hotel is None:
            try:
                from db import current_hotel_context
                hotel = current_hotel_context.get("hotel_name")
            except Exception:
                pass

        # -- Load KPIs strictly from db.RoomKPI (no fallback)
        available_metrics = []
        try:
            ModelKPI = getattr(db, "RoomKPI")  # will raise if model not defined
            q = session.query(ModelKPI)
            if hasattr(ModelKPI, "hotel_name") and hotel:
                q = q.filter(getattr(ModelKPI, "hotel_name") == hotel)
            if hasattr(ModelKPI, "is_active"):
                q = q.filter(getattr(ModelKPI, "is_active") == True)

            rows = q.all()

            # build (order, name) list safely
            items = []
            for r in rows:
                name = getattr(r, "name", None) or getattr(r, "kpi", None)
                if name:
                    order = getattr(r, "sort_order", 9999)
                    items.append((order, str(name)))

            items.sort(key=lambda t: (t[0], t[1].lower()))
            available_metrics = [n for _, n in items]
        except Exception:
            # If RoomKPI model/table doesn't exist yet, just show the message below
            available_metrics = []

        if not available_metrics:
            st.info("No KPIs are available. Go to **Labor ▸ Structure → KPIs** to add KPI names first.")
            st.stop()

        emp_df = refresh(db.Employee)
        depts = emp_df["department"].dropna().unique().tolist()

        col_dept, col_pos = st.columns(2)
        with col_dept:
            sel_dept = st.selectbox("Department", ["(Select)"] + sorted(depts))

        with col_pos:
            if sel_dept != "(Select)":
                pos_opts = (
                    emp_df.loc[emp_df["department"] == sel_dept, "role"]
                    .dropna()
                    .unique()
                    .tolist()
                )
                sel_pos = st.selectbox("Position", ["(Select)"] + sorted(pos_opts))
            else:
                sel_pos = "(Select)"

        if sel_pos == "(Select)":
            st.info("Select a department **and** position to view labor standards.")
            st.stop()

        pos_rec = (
            session.query(db.Position)
            .filter(func.lower(db.Position.name) == sel_pos.lower())
            .first()
        )
        if not pos_rec:
            st.error("Position not found in database.")
            st.stop()
        pos_id = pos_rec.id

        std_df = load_labor_standards(pos_id)
        if std_df.empty:
            std_df = pd.DataFrame([{"Metric": "", "Standard": None}])
        else:
            std_df = std_df[["Metric", "Standard"]].copy()
        std_df["Standard"] = pd.to_numeric(std_df["Standard"], errors="coerce")

        st.markdown("#### Labor Standards for This Position")

        edited_df = st.data_editor(
            std_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "Metric": st.column_config.SelectboxColumn(
                    "Metric", options=available_metrics
                ),
                "Standard": st.column_config.NumberColumn(
                    "Standard per FTE", format="%.2f"
                ),
            },
            key=f"labor_std_editor_{pos_id}"  # ensures reset when switching position
        )

        if st.button("💾 Save Standard"):
            edited_df = edited_df[
                edited_df["Metric"].notna() & edited_df["Standard"].notna()
            ]
            edited_df["Standard"] = edited_df["Standard"].astype(float)
            edited_df["Unit"] = edited_df["Metric"].astype(str) + " per FTE"
            save_labor_standards(pos_id, edited_df, hotel_name=hotel)
            st.success("Labor standards saved.")

# =========================================== LABOR ▸ ACTUAL HOURS =============
elif main_choice == "Labor ▸ Actual Hours":
      import os, re, pandas as pd
      from datetime import date, timedelta
      from dateutil.relativedelta import relativedelta, MO
      from sqlalchemy import or_, func
      from st_aggrid import AgGrid, GridOptionsBuilder

      # ────────────────────────── PAGE HEADER ───────────────────────────────
      st.markdown("""
      <div class="la-title">
            <!-- Lucide-style: calendar with clock -->
            <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
                  viewBox="0 0 24 24" fill="none" stroke="currentColor"
                  stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
                  <!-- calendar body -->
                  <rect x="3" y="4" width="18" height="17" rx="2"></rect>
                  <!-- binding rings -->
                  <line x1="8"  y1="2" x2="8"  y2="6"></line>
                  <line x1="16" y1="2" x2="16" y2="6"></line>
                  <!-- header separator -->
                  <line x1="3" y1="10" x2="21" y2="10"></line>
                  <!-- clock in lower-right -->
                  <circle cx="16" cy="16" r="5"></circle>
                  <path d="M16 13v3l2 1"></path>
            </svg>
            <span>Actual Hours</span>
      </div>
      <style>
            .la-title{
                  display:flex; align-items:center; gap:10px;
                  margin:0 0 10px 0; line-height:1;
            }
            .la-title svg{ color:#111; }      /* pure monochrome via currentColor */
            .la-title span{ font-weight:700; font-size:1.6rem; }
      </style>
      """, unsafe_allow_html=True)

      latest_date = (
            session.query(func.max(db.Actual.date))
                   .filter(or_(db.Actual.hours    != 0,
                               db.Actual.ot_hours != 0,
                               db.Actual.reg_pay  != 0,
                               db.Actual.ot_pay   != 0))
                   .scalar()
      )
      if latest_date:
            st.info(f"📅 Latest actual hours with activity: **{latest_date:%B %d, %Y}**")
      else:
            st.info("📅 No non‑zero actual‑hours data found yet.")

      # ── Role (same rule as other pages) ────────────────────────────────
      role_norm = (st.session_state.user.get("role") or "").strip().lower()
      # ───────────────────────── TAB BAR ────────────────────────────────────
      tab_import, tab_emp = st.tabs(["📥\uFE0E Import / Enter", "🕒\uFE0E Actual Hours"])

      # =====================================================================
      # TAB A — IMPORT / ENTER (Payroll CSV → db.Actual, Contract editor)
      # =====================================================================
      with tab_import:

            if role_norm != "manager":

                  st.subheader("📥 Import Payroll CSV (auto-map to Manual)")
                  pay_file = st.file_uploader(
                        "Drop payroll CSV here",
                        type=["csv"],
                        key="payroll_csv"
                  )

                  def norm_code(code: str) -> str:
                        return code.strip().casefold().replace(" ", "").replace(".", "")

                  if pay_file is not None:

                        pay_df = pd.read_csv(pay_file)

                        # ── column aliases --------------------------------------------------
                        COL_POS   = "Job"
                        COL_DATE  = "Business Date"
                        COL_HOURS = "Hours"
                        COL_AMT   = "Pay Amount"
                        COL_CODE  = "Pay Category"
                        COL_ID    = "Number"

                        pay_df[COL_DATE] = pd.to_datetime(pay_df[COL_DATE]).dt.date

                        pos_lookup = {p.name.strip(): p.id for p in session.query(db.Position).all()}

                        RAW = {
                              "Reg":      ("Reg Hours", "Reg Pay"),
                              "OT 1.5":   ("OT Hours",  "OT Pay"),
                              "OT1.5":    ("OT Hours",  "OT Pay"),
                        }
                        CODE_TO_METRIC = {norm_code(k): v for k, v in RAW.items()}

                        added_rows = skipped_pos = skipped_code = 0
                        examples_pos, examples_code = [], []

                        with session.no_autoflush:
                              overwrite_keys = {
                                    (pos_lookup.get(str(r[COL_POS]).strip()), r[COL_DATE])
                                    for _, r in pay_df.iterrows()
                                    if pos_lookup.get(str(r[COL_POS]).strip())
                              }
                              for pos_id, biz_date in overwrite_keys:
                                    session.query(db.Actual).filter_by(
                                          position_id=pos_id,
                                          date=biz_date,
                                          source="manual"
                                    ).delete()

                        for _, r in pay_df.iterrows():
                              pos_key = str(r[COL_POS]).strip()
                              pos_id  = pos_lookup.get(pos_key)
                              if not pos_id:
                                    skipped_pos += 1
                                    if len(examples_pos) < 5:
                                          examples_pos.append(pos_key)
                                    continue

                              paycode_norm = norm_code(str(r[COL_CODE]))
                              pair = CODE_TO_METRIC.get(paycode_norm)
                              if not pair:
                                    skipped_code += 1
                                    if len(examples_code) < 5:
                                          examples_code.append(r[COL_CODE])
                                    continue

                              metric_hrs, metric_pay = pair
                              biz_date   = r[COL_DATE]
                              hrs_val    = r[COL_HOURS]
                              pay_val    = r[COL_AMT]
                              emp_id_val = str(r[COL_ID]).strip().removesuffix(".0")

                              rec = db.Actual(
                                    emp_id      = emp_id_val,
                                    position_id = pos_id,
                                    date        = biz_date,
                                    hours       = hrs_val if metric_hrs == "Reg Hours" else 0,
                                    ot_hours    = hrs_val if metric_hrs == "OT Hours" else 0,
                                    reg_pay     = pay_val if metric_pay == "Reg Pay"   else 0,
                                    ot_pay      = pay_val if metric_pay == "OT Pay"    else 0,
                                    source      = "manual"
                              )
                              session.add(rec)
                              added_rows += 1

                        session.commit()

                        st.success(
                              f"✅ Imported {added_rows} rows · {skipped_pos} unknown positions · {skipped_code} unknown pay-codes"
                        )
                        if examples_pos:
                              st.warning(f"Unmatched position examples: {examples_pos}")
                        if examples_code:
                              st.warning(f"Unmatched pay-code examples: {examples_code}")

                        # ── 3) de-duplicate cached uploads so totals don’t double ——
                        key_cols = [COL_ID, COL_DATE, COL_POS, COL_CODE, COL_HOURS, COL_AMT]

                        if "payroll_cache" in st.session_state and not st.session_state["payroll_cache"].empty:
                              combined = pd.concat(
                                    [st.session_state["payroll_cache"], pay_df],
                                    ignore_index=True
                              )
                              combined.sort_values(key_cols, inplace=True)
                              combined.drop_duplicates(subset=key_cols, keep="last", inplace=True)
                              st.session_state["payroll_cache"] = combined.reset_index(drop=True)
                        else:
                              st.session_state["payroll_cache"] = pay_df.copy()

            else:
                  st.info("You have view-only access. Payroll imports are disabled for Managers.")

            # ────────── POSITION FILTER + WEEK NAVIGATOR + CONTRACT EDITOR ──────────
            st.subheader("Filter")

            dept_df = refresh(db.Department).rename(columns={"name": "dept"})
            pos_df  = (
                  refresh(db.Position)
                  .merge(dept_df, left_on="department_id", right_on="id")
                  [["id_x", "name", "dept"]]
                  .rename(columns={"id_x": "id"})
            )

            # 🔒 Manager scope (Actual Hours – Import tab)
            role_norm  = (st.session_state.user.get("role") or "").strip().lower()
            user_scope = st.session_state.user.get("scope", [])

            if role_norm == "manager" and user_scope:
                  allowed_depts  = {s["department"] for s in user_scope}
                  allowed_pairs = {(s["department"], s["position"]) for s in user_scope}

                  dept_df = dept_df[dept_df["dept"].isin(allowed_depts)]

                  pos_df = pos_df[
                        pos_df.apply(
                              lambda r: (r["dept"], r["name"]) in allowed_pairs,
                              axis=1
                        )
                  ]

            dept_opts = ["(All)"] + sorted(dept_df["dept"].dropna().unique())
            f1, f2 = st.columns(2)
            with f1:
                  sel_dept = st.selectbox("Department", dept_opts)
            if sel_dept != "(All)":
                  pos_opts = ["(All)"] + sorted(
                        pos_df.loc[pos_df["dept"] == sel_dept, "name"].dropna().unique()
                  )
            else:
                  pos_opts = ["(All)"]
            with f2:
                  sel_pos = st.selectbox("Position", pos_opts)

            if sel_pos != "(All)":
                  sel_pos_id = int(
                        pos_df.loc[(pos_df["dept"] == sel_dept) &
                                   (pos_df["name"] == sel_pos), "id"].values[0]
                  )
                  sel_path = f"{sel_dept}/{sel_pos}"
            else:
                  sel_pos_id, sel_path = None, None

            if "week_start" not in st.session_state:
                  st.session_state.week_start = date.today() + relativedelta(weekday=MO(-1))

            cprev, crange, cnext = st.columns([1, 3, 1])
            if cprev.button("◀", key="week_prev"):
                  st.session_state.week_start -= timedelta(days=7)
            if cnext.button("▶", key="week_next"):
                  st.session_state.week_start += timedelta(days=7)

            week_start = st.session_state.week_start
            week_end   = week_start + timedelta(days=6)
            days       = [week_start + timedelta(d) for d in range(7)]
            fmt = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
            crange.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

            if sel_pos_id is None:
                  st.info("Select a Department and Position to view or edit hours.")
            else:
                  st.markdown(f"#### {sel_path}")

                  act_rows = (session.query(db.Actual)
                                      .filter(db.Actual.position_id == sel_pos_id,
                                              db.Actual.date.between(week_start, week_end))
                                      .all())
                  df = pd.DataFrame([r.__dict__ for r in act_rows])

                  idx = pd.MultiIndex.from_product(
                        [["Manual", "Contract", "Total"],
                         ["Reg Hours", "OT Hours", "Reg Pay", "OT Pay"]],
                        names=["Block", "Metric"]
                  )
                  tbl = pd.DataFrame(index=idx,
                                     columns=[d.strftime(fmt) for d in days]).fillna(0.0)

                  def _add(block, metric, d, val):
                        tbl.loc[(block, metric), d.strftime(fmt)] += float(val or 0)

                  if not df.empty:
                        for _, row in df.iterrows():
                              block = row["source"].capitalize()
                              _add(block, "Reg Hours", row["date"], row["hours"])
                              _add(block, "OT Hours",  row["date"], row["ot_hours"])
                              _add(block, "Reg Pay",   row["date"], row["reg_pay"])
                              _add(block, "OT Pay",    row["date"], row["ot_pay"])

                  for d in days:
                        col = d.strftime(fmt)
                        for m in ["Reg Hours", "OT Hours", "Reg Pay", "OT Pay"]:
                              tbl.loc[("Total", m), col] = (
                                    tbl.loc[("Manual", m), col] +
                                    tbl.loc[("Contract", m), col]
                              )

                  full_df = tbl.reset_index()
                  st.markdown("""
                        <style>
                        .ag-theme-streamlit .ag-root-wrapper {
                              border-radius: 12px !important;
                        }
                        </style>
                  """, unsafe_allow_html=True)

                  gb = GridOptionsBuilder.from_dataframe(full_df)
                  gb.configure_default_column(resizable=True, flex=1, minWidth=100)
                  gb.configure_default_column(editable=True, type=["numericColumn"], precision=2)
                  gb.configure_column("Block", editable=False, pinned="left")
                  gb.configure_column("Metric", editable=False, pinned="left")
                  gb.configure_grid_options(suppressHorizontalScroll=False)
                  gb.configure_grid_options(forceFitColumns=True)

                  grid = AgGrid(
                        full_df,
                        gridOptions=gb.build(),
                        theme="streamlit",
                        fit_columns_on_grid_load=True,
                        allow_unsafe_jscode=True,
                        domLayout='autoHeight',
                        key=f"grid_{sel_pos_id}_{week_start}"
                  )
                  if st.button("💾 Save CL Entries"):
                        edited_df   = pd.DataFrame(grid["data"])
                        contract_df = edited_df[edited_df["Block"] == "Contract"].set_index("Metric")

                        affected_keys = set()
                        for d in days:
                              col = d.strftime(fmt)
                              session.query(db.Actual).filter(
                                    db.Actual.position_id == sel_pos_id,
                                    db.Actual.date        == d,
                                    db.Actual.source      == "contract"
                              ).delete(synchronize_session=False)

                              for metric in ["Reg Hours", "OT Hours", "Reg Pay", "OT Pay"]:
                                    val = float(contract_df.at[metric, col])
                                    if val == 0 or pd.isna(val):
                                          continue
                                    session.add(db.Actual(
                                          emp_id      = None,
                                          position_id = sel_pos_id,
                                          date        = d,
                                          hours       = val if metric == "Reg Hours" else 0,
                                          ot_hours    = val if metric == "OT Hours" else 0,
                                          reg_pay     = val if metric == "Reg Pay"  else 0,
                                          ot_pay      = val if metric == "OT Pay"   else 0,
                                          source      = "contract"
                                    ))
                              affected_keys.add((sel_pos_id, d))

                        refresh_totals(session, affected_keys)
                        session.commit()
                        st.success("✅ Contract data saved.")

      with tab_emp:

            st.subheader("Employee Actual Hours")

            # ── 1) Filters ---------------------------------------------------
            dept_df = refresh(db.Department).rename(columns={"name": "dept"})
            pos_df  = (
                  refresh(db.Position)
                  .merge(dept_df, left_on="department_id", right_on="id")
                  [["id_x", "name", "dept"]]
                  .rename(columns={"id_x": "id"})
            )

            # 🔒 Manager scope (Actual Hours – Employee tab)
            role_norm  = (st.session_state.user.get("role") or "").strip().lower()
            user_scope = st.session_state.user.get("scope", [])

            if role_norm == "manager" and user_scope:
                  allowed_depts  = {s["department"] for s in user_scope}
                  allowed_pairs = {(s["department"], s["position"]) for s in user_scope}

                  dept_df = dept_df[dept_df["dept"].isin(allowed_depts)]

                  pos_df = pos_df[
                        pos_df.apply(
                              lambda r: (r["dept"], r["name"]) in allowed_pairs,
                              axis=1
                        )
                  ]

            dept_opts = ["(All)"] + sorted(dept_df["dept"].dropna().unique())
            col1, col2 = st.columns(2)
            with col1:
                  sel_dept = st.selectbox("Department", dept_opts, key="emp_dept")
            if sel_dept != "(All)":
                  pos_opts = ["(All)"] + sorted(
                        pos_df.loc[pos_df["dept"] == sel_dept, "name"].dropna().unique()
                  )
            else:
                  pos_opts = ["(All)"]
            with col2:
                  sel_pos = st.selectbox("Position", pos_opts, key="emp_pos")

            # ── 2) Week navigator -------------------------------------------
            if "emp_week_start" not in st.session_state:
                  st.session_state.emp_week_start = date.today() + relativedelta(weekday=MO(-1))

            col_prev, col_range, col_next = st.columns([1, 3, 1])
            if col_prev.button("◀", key="emp_prev"):
                  st.session_state.emp_week_start -= timedelta(days=7)
            if col_next.button("▶", key="emp_next"):
                  st.session_state.emp_week_start += timedelta(days=7)

            week_start = st.session_state.emp_week_start
            week_end   = week_start + timedelta(days=6)
            days       = [week_start + timedelta(d) for d in range(7)]
            fmt        = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
            fmt_cols   = [d.strftime(fmt) for d in days]
            col_range.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

            # ── 3) Pull Actual rows from DB (week-range + filters) ----------
            q = (
                  session.query(
                        db.Actual.emp_id.label("Number"),
                        db.Actual.date.label("Business Date"),
                        (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
                        db.Position.name.label("Position"),
                        db.Department.name.label("Department")
                  )
                  .join(db.Position,   db.Actual.position_id == db.Position.id)
                  .join(db.Department, db.Position.department_id == db.Department.id)
                  .filter(db.Actual.date.between(week_start, week_end))
                  .filter(or_(db.Actual.hours != 0, db.Actual.ot_hours != 0))
            )

            if sel_dept != "(All)":
                  q = q.filter(db.Department.name == sel_dept)
            if sel_pos != "(All)":
                  q = q.filter(db.Position.name == sel_pos)

            raw = pd.DataFrame(q.all(),
                               columns=["Number", "Business Date", "Hours", "Position", "Department"])

            # Allow blank emp_id (None) and still show hours
            raw["Number"] = raw["Number"].astype(str).str.strip().fillna("")

            raw["Business Date"] = pd.to_datetime(raw["Business Date"]).dt.date

            if raw.empty:
                  st.warning("No rows match the current filters / week.")
                  st.stop()

            # ── 4) Parse employee names -------------------------------------
            emp_df = refresh(db.Employee).copy()
            # 🔒 Manager scope (Employee Actual Hours – employee lookup)
            emp_df = apply_manager_scope(
                  emp_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})
            parts = emp_df["name"].astype(str).str.extract(
                  r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
            )
            emp_df["ID"]         = parts["ID"].fillna("").astype(str).str.strip().str.zfill(5)
            emp_df["First Name"] = parts["First_Name"].str.strip()
            emp_df["Last Name"]  = parts["Last_Name"].str.strip()

            raw = pd.DataFrame(q.all())

            # Drop records with no employee ID (emp_id is None)
            raw = raw[raw["Number"].notna()]

            # Normalize 'Number' to match the format of emp_df["ID"]
            raw["Number"] = (
                  pd.to_numeric(raw["Number"], errors="coerce")
                    .fillna(0)
                    .astype(int)
                    .astype(str)
                    .str.zfill(5)
            )

            # ── 5) Pivot table by day ---------------------------------------
            pivot = (
                  raw.pivot_table(index="Number",
                                  columns="Business Date",
                                  values="Hours",
                                  aggfunc="sum",
                                  fill_value=0)
                     .reset_index()
            )

            # ── Fix ID parsing and cleanup ----------------------------------
            pivot = pivot[pd.to_numeric(pivot["Number"], errors="coerce").notna()]
            pivot["ID"] = (
                  pivot["Number"]
                        .astype(str)
                        .str.strip()
                        .str.replace(r"\.0$", "", regex=True)
            )
            pivot.drop(columns="Number", inplace=True)

            # create temp keys with stripped leading zeros for matching only
            pivot["match_ID"]   = pivot["ID"].astype(str).str.lstrip("0")
            emp_df["match_ID"]  = emp_df["ID"].astype(str).str.lstrip("0")
            # ── FIX: remove duplicate employees with same ID (keep active or latest) ──
            if "emp_type" in emp_df.columns:
                  # Prefer active employees over terminated
                  emp_df = (
                        emp_df.sort_values(["match_ID", "emp_type"], ascending=[True, True])
                               .drop_duplicates(subset=["match_ID"], keep="first")
                  )
            else:
                  # Fallback: keep only the last occurrence
                  emp_df = emp_df.drop_duplicates(subset=["match_ID"], keep="last")

            pivot = pivot.merge(emp_df[["match_ID", "First Name", "Last Name"]],
                                on="match_ID", how="left")

            pivot.drop(columns="match_ID", inplace=True)
            rename_map = {d: label for d, label in zip(days, fmt_cols)}
            pivot.rename(columns=rename_map, inplace=True)

            # ensure all day columns exist
            for label in fmt_cols:
                  if label not in pivot.columns:
                        pivot[label] = 0.0

            final_cols = ["ID", "First Name", "Last Name"] + fmt_cols
            pivot = pivot[final_cols]

            missing_emp_rows = raw[raw["Number"].isna()]
            st.warning(f"Dropped {len(missing_emp_rows)} rows with missing emp_id.")

            # ── 6) Show table -----------------------------------------------
            gb = GridOptionsBuilder.from_dataframe(pivot)
            gb.configure_default_column(resizable=True, type=["numericColumn"], precision=2)
            for col in ["ID", "First Name", "Last Name"]:
                  gb.configure_column(col, pinned="left", editable=False)

            st.markdown("""
                  <style>
                  .ag-theme-streamlit .ag-root-wrapper {
                        border-radius: 12px !important;
                  }
                  </style>
            """, unsafe_allow_html=True)

            gb.configure_grid_options(domLayout='autoHeight', suppressHorizontalScroll=False)
            gb.configure_grid_options(forceFitColumns=True)

            AgGrid(
                  pivot,
                  gridOptions=gb.build(),
                  theme="streamlit",
                  fit_columns_on_grid_load=True,
                  allow_unsafe_jscode=True,
                  domLayout='autoHeight',
                  key=f"emp_table_{week_start}"
            )
# =============================================== ROOM STATs ================
elif main_choice == "Room STATs":
    import os
    import datetime
    from datetime import date, timedelta
    from dateutil.relativedelta import relativedelta, MO
    import pandas as pd  # ensure available for df build

    st.markdown("""
    <div class="la-title">
        <!-- Lucide-style: building (windows + door) -->
        <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
             viewBox="0 0 24 24" fill="none" stroke="currentColor"
             stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
            <!-- main building -->
            <rect x="7" y="3" width="10" height="18" rx="2"></rect>
            <!-- ground line -->
            <path d="M3 21h18"></path>
            <!-- windows -->
            <path d="M10 7h.01M14 7h.01M10 10h.01M14 10h.01M10 13h.01M14 13h.01"></path>
            <!-- door -->
            <path d="M11 21v-3a1 1 0 0 1 2 0v3"></path>
        </svg>
        <span>Room STATs</span>
    </div>
    <style>
        .la-title{
            display:flex; align-items:center; gap:10px;
            margin:0 0 10px 0; line-height:1;
        }
        .la-title svg{ color:#111; }
        .la-title span{ font-weight:700; font-size:1.6rem; }
    </style>
    """, unsafe_allow_html=True)

    # ─────────────────────────── 1. Week context ───────────────────────
    if "rs_week_start" not in st.session_state:
        st.session_state.rs_week_start = date.today() + relativedelta(weekday=MO(-1))

    week_start = st.session_state.rs_week_start
    week_end   = week_start + timedelta(days=6)
    week_dates = [week_start + timedelta(days=i) for i in range(7)]
    fmt_day    = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
    day_cols   = [d.strftime(fmt_day) for d in week_dates]

    # resolve current hotel
    hotel = None
    try:
        hotel = (st.session_state.get("user") or {}).get("hotel_name")
    except Exception:
        hotel = None
    if hotel is None:
        try:
            from db import current_hotel_context
            hotel = current_hotel_context.get("hotel_name")
        except Exception:
            pass

    # ─────────────────────────── 2. Tab-style Selectors ─────────────────
    subtab_rs = st.tabs(["📈\uFE0E Actuals", "📅\uFE0E Forecast", "📊\uFE0E OTB + Pickup"])
    tab_labels = ["📈\uFE0E Actuals", "📅\uFE0E Forecast", "📊\uFE0E OTB + Pickup"]
    tab_models = [db.RoomActual, db.RoomForecast, db.RoomOTBPickup]

    # ─────────────────────────── 3. Tab Content ────────────────────────
    for tab, label, Model in zip(subtab_rs, tab_labels, tab_models):
        with tab:
            # ── Week navigator ─────────────────────
            nav_prev, nav_range, nav_next = st.columns([1, 3, 1])
            if nav_prev.button("◀", key=f"rs_week_prev_{label}"):
                st.session_state.rs_week_start -= timedelta(days=7)
                st.rerun()
            if nav_next.button("▶", key=f"rs_week_next_{label}"):
                st.session_state.rs_week_start += timedelta(days=7)
                st.rerun()
            nav_range.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

            # ───────────── Fetch data ─────────────
            session = get_scoped_session()

            entry_kpis = []
            rule_kpis  = {}
            if hasattr(db, "RoomKPI"):
                kpi_q = session.query(db.RoomKPI).filter(db.RoomKPI.is_active == True)
                if hasattr(db.RoomKPI, "hotel_name") and hotel:
                    kpi_q = kpi_q.filter(db.RoomKPI.hotel_name == hotel)
                kpi_rows = kpi_q.order_by(db.RoomKPI.sort_order.asc(), db.RoomKPI.name.asc()).all()
                for r in kpi_rows:
                    nm = str(r.name).strip()
                    if not nm:
                        continue
                    if bool(getattr(r, "has_rule", False)):
                        rule_kpis[nm] = (r.rule_expr or "")
                    else:
                        entry_kpis.append(nm)
            else:
                entry_kpis = []
                rule_kpis  = {}

            if not entry_kpis and not rule_kpis:
                st.info("No KPIs are available. Go to **Labor ▸ Structure → KPIs** to add KPI names first.")
                st.stop()

            kpis = entry_kpis

            # Pull existing rows for the week
            if hotel and hasattr(Model, "hotel_name"):
                rows = session.query(Model).filter(
                    Model.date.between(week_start, week_end),
                    getattr(Model, "hotel_name") == hotel
                ).all()
            else:
                rows = session.query(Model).filter(
                    Model.date.between(week_start, week_end)
                ).all()

            data_dict = {k: {"KPI": k} for k in kpis}
            for r in rows:
                col_lbl = r.date.strftime(fmt_day)
                if r.kpi in data_dict:
                    data_dict[r.kpi][col_lbl] = r.value
            for k in kpis:
                for col in day_cols:
                    data_dict[k].setdefault(col, 0)

            df_edit = pd.DataFrame(data_dict.values()).loc[:, ["KPI"] + day_cols]

            # ───────────── Save Button ─────────────
            save_col = st.columns([10, 1])
            with save_col[1]:
                save_click = st.button("💾", key=f"save_btn_{Model.__name__}")

            # ───────────── Styling ─────────────
            st.markdown("""
                <style>
                section[data-testid="stDataEditor"] thead tr {
                    background-color: #2D2D2D !important;
                }
                section[data-testid="stDataEditor"] thead th {
                    color: white !important;
                    font-weight: bold !important;
                    text-align: center !important;
                    border-right: 1px solid #ddd !important;
                }
                section[data-testid="stDataEditor"] div[data-testid^="cell-"][data-testid$="-0"] {
                    font-weight: bold !important;
                    color: #1F4E79 !important;
                    background-color: #E3F2FD !important;
                }
                section[data-testid="stDataEditor"] div[data-testid^="cell-"] {
                    border-right: 1px solid #eee !important;
                    white-space: nowrap;
                    overflow: hidden;
                    text-overflow: ellipsis;
                }
                </style>
            """, unsafe_allow_html=True)

            # ───────────── Auto height for data_editor ─────────────
            _ROW_PX     = 36
            _HEADER_PX  = 42
            _FOOTER_PX  = 8
            _nrows = max(1, len(df_edit))
            grid_height = _HEADER_PX + _ROW_PX * _nrows + _FOOTER_PX
            _MAX_PX = 900
            grid_height = min(grid_height, _MAX_PX)

            # ───────────── Editable Table ─────────────
            edited_df = st.data_editor(
                df_edit,
                column_config={
                    "KPI": st.column_config.TextColumn("KPI", disabled=True),
                },
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                key=f"room_stats_editor_{Model.__name__}",
                height=grid_height,  # auto-fit height
            )

            # ───────────── Save Logic ─────────────
            if save_click:
                if hasattr(Model, "hotel_name") and not hotel:
                    st.error("No hotel selected in session; cannot save scoped stats.")
                else:
                    edited = edited_df.set_index("KPI")

                    base_filter = [Model.date.between(week_start, week_end)]
                    if hasattr(Model, "hotel_name") and hotel:
                        base_filter.append(getattr(Model, "hotel_name") == hotel)

                    session.query(Model).filter(*base_filter).delete(synchronize_session=False)

                    to_add = []
                    for kpi in kpis:
                        for d, col in zip(week_dates, day_cols):
                            try:
                                val = int(edited.at[kpi, col])
                            except Exception:
                                val = 0
                            kwargs = dict(kpi=kpi, date=d, value=val)
                            if hasattr(Model, "hotel_name") and hotel:
                                kwargs["hotel_name"] = hotel
                            to_add.append(Model(**kwargs))
                    if to_add:
                        session.add_all(to_add)
                    session.commit()

                    # ───────────── DEPENDENCY-AWARE RULE EVAL (replacement) ─────────────
                    if rule_kpis:
                        # base values present for this week/model/hotel (manual entries we just saved)
                        rows_week = session.query(Model).filter(*base_filter).all()
                        base_values = {(str(r.kpi), r.date): float(r.value or 0.0) for r in rows_week}

                        import re, ast, operator as _op
                        _TOKEN_RE = re.compile(r"\[([^\]]+)\]")
                        _OPS = {ast.Add:_op.add, ast.Sub:_op.sub, ast.Mult:_op.mul,
                                ast.Div:_op.truediv, ast.Pow:_op.pow, ast.USub:_op.neg}

                        def _safe_eval(expr: str) -> float:
                            node = ast.parse(expr, mode="eval").body
                            def _ev(n):
                                if isinstance(n, ast.Num):      return float(n.n)
                                if isinstance(n, ast.UnaryOp):  return _OPS[type(n.op)](_ev(n.operand))
                                if isinstance(n, ast.BinOp):    return _OPS[type(n.op)](_ev(n.left), _ev(n.right))
                                return 0.0
                            return float(_ev(node) or 0.0)

                        memo = {}
                        visiting = set()

                        def resolve(kpi_name: str, day: date) -> float:
                            key = (kpi_name, day)
                            if key in memo:
                                return memo[key]
                            if key in visiting:
                                return 0.0
                            # base (manual) available?
                            if key in base_values:
                                memo[key] = float(base_values[key] or 0.0)
                                return memo[key]
                            # rule-defined?
                            expr = rule_kpis.get(kpi_name)
                            if not expr:
                                memo[key] = 0.0
                                return 0.0
                            visiting.add(key)
                            def _subst(m):
                                dep = m.group(1)
                                return str(resolve(dep, day))
                            substituted = _TOKEN_RE.sub(_subst, expr or "")
                            try:
                                val = _safe_eval(substituted)
                            except Exception:
                                val = 0.0
                            visiting.remove(key)
                            memo[key] = float(val or 0.0)
                            return memo[key]

                        rule_rows = []
                        for kpi_name, expr in rule_kpis.items():
                            for d in week_dates:
                                v = resolve(kpi_name, d)
                                kwargs = dict(kpi=kpi_name, date=d, value=int(round(v)))
                                if hasattr(Model, "hotel_name") and hotel:
                                    kwargs["hotel_name"] = hotel
                                rule_rows.append(Model(**kwargs))
                        if rule_rows:
                            session.add_all(rule_rows)
                            session.commit()

                    st.markdown(f"""
                        <div style="
                            position: fixed;
                            bottom: 24px;
                            right: 24px;
                            background-color: #4CAF50;
                            color: white;
                            padding: 14px 20px;
                            border-radius: 6px;
                            font-size: 15px;
                            box-shadow: 0px 4px 10px rgba(0,0,0,0.25);
                            z-index: 1000;
                            animation: fadeOut 5s forwards;
                        ">
                            ✅ Room stats for <b>{label}</b> saved successfully.
                        </div>
                        <style>
                            @keyframes fadeOut {{
                                0%   {{ opacity: 1; }}
                                80%  {{ opacity: 1; }}
                                100% {{ opacity: 0; display: none; }}
                            }}
                        </style>
                    """, unsafe_allow_html=True)

            # ───────────── DEBUG: Rule KPI numeric evaluation (this tab's data) ─────────────
            with st.expander("🧩 Debug: Rule KPI Calculations (this tab's data)"):
                if not rule_kpis:
                    st.info("No rule-based KPIs defined (has_rule=True).")
                else:
                    # base values for this tab/week/hotel
                    dq = session.query(Model).filter(Model.date.between(week_start, week_end))
                    if hasattr(Model, "hotel_name") and hotel:
                        dq = dq.filter(getattr(Model, "hotel_name") == hotel)
                    dbg_rows = dq.all()
                    base_values = {(str(r.kpi), r.date): float(r.value or 0.0) for r in dbg_rows}

                    import re, ast, operator as _op
                    _TOKEN_RE = re.compile(r"\[([^\]]+)\]")
                    _OPS = {ast.Add:_op.add, ast.Sub:_op.sub, ast.Mult:_op.mul,
                            ast.Div:_op.truediv, ast.Pow:_op.pow, ast.USub:_op.neg}

                    def _safe_eval(expr: str) -> float:
                        node = ast.parse(expr, mode="eval").body
                        def _ev(n):
                            if isinstance(n, ast.Num):      return float(n.n)
                            if isinstance(n, ast.UnaryOp):  return _OPS[type(n.op)](_ev(n.operand))
                            if isinstance(n, ast.BinOp):    return _OPS[type(n.op)](_ev(n.left), _ev(n.right))
                            return 0.0
                        return float(_ev(node) or 0.0)

                    memo = {}
                    visiting = set()

                    def resolve(kpi_name: str, day: date) -> float:
                        key = (kpi_name, day)
                        if key in memo: return memo[key]
                        if key in visiting: return 0.0
                        if key in base_values:
                            memo[key] = float(base_values[key] or 0.0)
                            return memo[key]
                        expr = rule_kpis.get(kpi_name)
                        if not expr:
                            memo[key] = 0.0
                            return 0.0
                        visiting.add(key)
                        # substitute recursively with numeric values
                        def _subst_num(m):
                            dep = m.group(1)
                            return str(resolve(dep, day))
                        num_expr = _TOKEN_RE.sub(_subst_num, expr or "")
                        try:
                            val = _safe_eval(num_expr)
                        except Exception:
                            val = 0.0
                        visiting.remove(key)
                        memo[key] = float(val or 0.0)
                        return memo[key]

                    def explain(expr: str, day: date):
                        """Return (annotated_expr, numeric_expr, value)."""
                        if not expr:
                            return ("", "0", 0.0)

                        # annotated: [KPI] -> (KPI=number)
                        def _subst_ann(m):
                            dep = m.group(1)
                            return f"({dep}={resolve(dep, day)})"

                        # numeric: [KPI] -> number
                        def _subst_num(m):
                            dep = m.group(1)
                            return str(resolve(dep, day))

                        annotated = _TOKEN_RE.sub(_subst_ann, expr or "")
                        numeric   = _TOKEN_RE.sub(_subst_num, expr or "")
                        try:
                            value = _safe_eval(numeric)
                        except Exception:
                            value = 0.0
                        return (annotated, numeric, float(value or 0.0))

                    # Build tables
                    summary = {"KPI": []}
                    for col in day_cols:
                        summary[col] = []

                    details_rows = []  # KPI, Date, Rule, Annotated, Numeric, Value

                    for kpi_name, expr in rule_kpis.items():
                        summary["KPI"].append(kpi_name)
                        for d, col in zip(week_dates, day_cols):
                            ann, num, val = explain(expr, d)
                            summary[col].append(round(val, 2))
                            details_rows.append({
                                "KPI": kpi_name,
                                "Date": d.strftime(fmt_day),
                                "Rule": expr or "",
                                "Annotated": ann,     # e.g., (Stayovers=15) - (Average Guests per Room=2.5)
                                "Numeric": num,       # e.g., 15 - 2.5
                                "Value": round(val, 4)
                            })

                    st.markdown("**Summary (result by day)**")
                    st.dataframe(pd.DataFrame(summary), use_container_width=True, hide_index=True)

                    st.markdown("**Details (token → number substitution)**")
                    st.dataframe(pd.DataFrame(details_rows), use_container_width=True, hide_index=True)
# ---------- SCHEDULING ----------------------------------------------------
elif main_choice == "Scheduling":
    import os, json, pandas as pd
    from datetime import date, timedelta, datetime, time
    from dateutil.relativedelta import relativedelta, MO
    from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
    from sqlalchemy import func
    from collections import defaultdict

    st.markdown("""
    <div class="la-title">
      <!-- Lucide: calendar -->
      <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
           viewBox="0 0 24 24" fill="none" stroke="currentColor"
           stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
        <rect x="3" y="4" width="18" height="16" rx="2"></rect>
        <path d="M16 2v4M8 2v4M3 10h18"></path>
      </svg>
      <span>Weekly Scheduling</span>
    </div>
    <style>
      .la-title{
        display:flex; align-items:center; gap:10px;
        margin:0 0 10px 0; line-height:1;
      }
      .la-title svg{ color:#111; }
      .la-title span{ font-weight:700; font-size:1.6rem; }
    </style>
    """, unsafe_allow_html=True)

    # ---------- role guard ----------------------------------------------------
    _sch_role = (st.session_state.user.get("role") or "").strip().lower()
    _is_employee = _sch_role == "employee"

    # ---------- 1) week reference ------------------------------------------
    if "sch_week_start" not in st.session_state:
        st.session_state.sch_week_start = date.today() + relativedelta(weekday=MO(-1))

    week_start = st.session_state.sch_week_start
    week_end   = week_start + timedelta(days=6)
    week_dates = [week_start + timedelta(i) for i in range(7)]
    week_dates_str = [d.strftime("%Y-%m-%d") for d in week_dates]
    fmt_day    = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
    day_cols   = [d.strftime(fmt_day) for d in week_dates]

    # ---------- 2) Filters --------------------------------------------------
    emp_df = refresh(db.Employee)

    emp_df = emp_df.loc[
        emp_df["emp_type"].fillna("").str.strip().str.casefold() != "terminated"
    ].copy()

    emp_df = apply_manager_scope(
        emp_df.rename(columns={"role": "position"})
    ).rename(columns={"position": "role"})
    col_dept, col_pos = st.columns(2)
    with col_dept:
        sel_dept = st.selectbox("Department*", ["(Select)"] + sorted(emp_df["department"].dropna().unique()))
    with col_pos:
        if sel_dept != "(Select)":
            pos_opts = emp_df.loc[emp_df["department"] == sel_dept, "role"].dropna().unique()
            sel_pos  = st.selectbox("Position*", ["(Select)"] + sorted(pos_opts))
            if sel_pos != "(Select)":
                st.session_state["selected_pos"] = sel_pos
            st.session_state["selected_pos"] = sel_pos
            st.session_state["selected_pos"] = sel_pos
        else:
            sel_pos = "(Select)"

    if sel_dept == "(Select)" or sel_pos == "(Select)":
        st.info("Select **Department** and **Position** to view schedule grid.")
        st.stop()

    # ---------- 3) pull employees ------------------------------------------
    emp_sub = emp_df[(emp_df["department"] == sel_dept) & (emp_df["role"] == sel_pos)]
    emp_sub = emp_sub.drop_duplicates(subset=["id"]).reset_index(drop=True)
    if emp_sub.empty:
        st.warning("No employees match that Department / Position.")
        st.stop()

    ids    = emp_sub["name"].str.extract(r"(\d+)$")[0].fillna("")
    firsts = emp_sub["name"].str.extract(r",\s*([^\d]+)")[0].str.strip()
    lasts  = emp_sub["name"].str.extract(r"^\s*([^,]+)")[0].str.strip()

    sched_df = pd.DataFrame({
        "ID": ids,
        "First Name": firsts,
        "Last Name": lasts,
        "emp_id": emp_sub["id"],
    })
    for dc in day_cols:
        sched_df[dc] = ""
    sched_df["Total"] = ""
    sched_df = sched_df.drop_duplicates(subset=["ID", "First Name", "Last Name"]).reset_index(drop=True)

    # ---------- 4) load saved rows & OFF highlight -------------------------
    saved_rows = session.query(db.Schedule).filter(
        db.Schedule.emp_id.in_(emp_sub["id"]),
        db.Schedule.day.in_(week_dates_str)
    ).all()

    avail_rows = session.query(db.Schedule).filter(
        db.Schedule.emp_id.in_(emp_sub["id"]),
        db.Schedule.day.in_(WK_TO_DATE.values())
    ).all()

    off_map = {}
    for r in avail_rows:
        if r.shift_type.upper() == "OFF":
            lbl = r.day.strftime("%a")
            off_map.setdefault(r.emp_id, set()).add(lbl)

    for r in saved_rows:
        col = r.day.strftime(fmt_day)
        idx = sched_df.index[sched_df["emp_id"] == r.emp_id][0]
        sched_df.at[idx, col] = r.shift_type

    cell_styles = {}
    for idx, row in sched_df.iterrows():
        for d, col in zip(week_dates, day_cols):
            if row["emp_id"] in off_map and d.strftime("%a") in off_map[row["emp_id"]]:
                sched_df.at[idx, col] = "OFF"
                cell_styles[f"{row['ID']}|||{col}"] = {"backgroundColor": "#f8d7da"}

    # ---------- helpers -----------------------------------------------------
    def do_rerun():
        st.rerun() if hasattr(st, "rerun") else st.experimental_rerun()

    if "undo_backup" not in st.session_state:
        st.session_state["undo_backup"] = {}

    def backup_diffs(old_map, _new_map):
        import copy
        st.session_state["undo_backup"] = copy.deepcopy(old_map)

    def undo_last_change():
        undo_map = st.session_state.get("undo_backup", {})
        if not undo_map:
            st.warning("Nothing to undo.")
            return

        emp_ids  = {k[0] for k in undo_map}
        day_vals = {k[1] if isinstance(k[1], date) else datetime.strptime(k[1], "%Y-%m-%d").date()
                    for k in undo_map}

        session.query(db.Schedule).filter(
            db.Schedule.emp_id.in_(emp_ids),
            db.Schedule.day.in_(day_vals)
        ).delete(synchronize_session=False)

        session.add_all([
            db.Schedule(
                emp_id=k[0],
                day=k[1] if isinstance(k[1], date) else datetime.strptime(k[1], "%Y-%m-%d").date(),
                shift_type=v
            )
            for k, v in undo_map.items()
        ])
        session.commit()

        st.session_state["undo_backup"] = {}
        st.success("Undo complete.")
        do_rerun()

    # ---------- 5) shift list + quick add -----------------------------------
    st.markdown("#### Shifts available for this position")
    ShiftTbl = db.ShiftTime
    pos_rec  = session.query(db.Position).filter(func.lower(db.Position.name)==sel_pos.lower()).first()
    shift_rows = session.query(ShiftTbl).filter(ShiftTbl.position_id==pos_rec.id).all()
    fmt = lambda t: t.strftime("%H:%M") if t else ""
    shift_strings = sorted({f"{fmt(r.start)}-{fmt(r.end)}" for r in shift_rows if r.start and r.end})
    buckets = {"Morning": [], "Afternoon": [], "Evening": []}
    for r in shift_rows:
        buckets[r.period].append(f"{fmt(r.start)}-{fmt(r.end)}")

    def _to12(hhmm: str) -> str:
        try:
            h, m = map(int, hhmm.split(":"))
            t = datetime(2000, 1, 1, h, m)
            s = t.strftime("%I:%M %p")
            return s.lstrip("0")
        except:
            return hhmm

    def _to12_range(rng: str) -> str:
        try:
            a, b = rng.split("-")
            return f"{_to12(a)} - {_to12(b)}"
        except:
            return rng

    for p in ["Morning", "Afternoon", "Evening"]:
        shown = ", ".join(sorted(_to12_range(s) for s in buckets[p])) if buckets[p] else "*No shifts created*"
        st.markdown(f"**{p}:** {shown}")

    shift_opts = shift_strings[:]

    if not _is_employee:
        with st.expander("➕ Add new shift", expanded=False):
            period = st.selectbox("Time of day", ["Morning", "Afternoon", "Evening"])
            c1, c2 = st.columns(2)
            with c1: t_start = st.time_input("Start", value=time(0,0), step=1800)
            with c2: t_end   = st.time_input("End",   value=time(12,0), step=1800)
            if st.button("💾 Save Shift"):
                if t_end <= t_start:
                    st.error("End time must be after start.")
                else:
                    new_shift = f"{t_start.strftime('%H:%M')}-{t_end.strftime('%H:%M')}"
                    if new_shift in shift_opts:
                        st.warning("That shift already exists.")
                    else:
                        session.add(ShiftTbl(position_id=pos_rec.id, period=period, start=t_start, end=t_end))
                        session.commit(); st.success("Shift saved."); do_rerun()

    # ---------- 6) editable grid -------------------------------------------
    df_view = sched_df.copy()

    # --- dynamic Total hours (30-min break per shift) ----------------------
    total_js = JsCode(f"""
        function(p) {{
            const cols = {json.dumps(day_cols)};
            let t = 0;

            function hoursWithBreak(s) {{
                if (!s) return 0;
                const S = s.toString().trim();
                if (!S || S.toUpperCase() === "OFF") return 0;
                const a = S.split("-");
                if (a.length !== 2) return 0;

                const [h0,m0] = a[0].split(":").map(Number);
                const [h1,m1] = a[1].split(":").map(Number);
                if ([h0,m0,h1,m1].some(v => isNaN(v))) return 0;

                let diffMin = (h1*60 + m1) - (h0*60 + m0);
                if (diffMin <= 0) diffMin += 1440;   // cross-midnight
                let hrs = diffMin / 60;
                hrs = Math.max(0, hrs - 0.5);        // subtract 30-min break
                return hrs;
            }}

            cols.forEach(c => t += hoursWithBreak(p.data[c] || ""));
            return t.toFixed(2);
        }}
    """)

    icon_renderer = JsCode("""
        function(p) {
            const v = Number(p.value || 0).toFixed(2);
            const icon = v > 40 ? " ⚠️" : "";
            return v + icon;
        }
    """)

    style_total = JsCode("""
        function(p) {
            const v = Number(p.value || 0);
            if (v > 40) {
                return {fontWeight: 'bold', color: '#d9534f'};
            }
            return {fontWeight: 'bold'};
        }
    """)

    # display 12-hour in cells (values remain 24h)
    fmt12_js = JsCode("""
    function(p){
      const v = (p.value || "").toString().trim();
      if (!v || v.toUpperCase() === "OFF") return v;
      const parts = v.split("-");
      if (parts.length !== 2) return v;

      function fmt(hhmm){
        const seg = hhmm.split(":");
        if (seg.length !== 2) return hhmm;
        let h = parseInt(seg[0], 10);
        const m = parseInt(seg[1], 10);
        if (isNaN(h) || isNaN(m)) return hhmm;
        const suff = h >= 12 ? " PM" : " AM";
        h = h % 12; if (h === 0) h = 12;
        const mm = (m < 10 ? "0" : "") + m;
        return `${h}:${mm}${suff}`;
      }

      return `${fmt(parts[0])} - ${fmt(parts[1])}`;
    }
    """)

    gb = GridOptionsBuilder.from_dataframe(df_view)
    _grid_editable = not _is_employee
    gb.configure_default_column(editable=_grid_editable)
    gb.configure_default_column(editable=_grid_editable, singleClickEdit=_grid_editable)

    for col in ["ID", "First Name", "Last Name", "emp_id"]:
        gb.configure_column(col, editable=False, hide=True if col=="emp_id" else False)
    gb.configure_column("ID", cellStyle={"fontWeight": "bold"})
    gb.configure_column("First Name", cellStyle={"fontWeight": "bold"})
    gb.configure_column("Last Name", cellStyle={"fontWeight": "bold"})

    gb.configure_grid_options(enableRangeSelection=True, enableFillHandle=True,
                              undoRedoCellEditing=True, undoRedoCellEditingLimit=100,
                              clipboardDelimiter=",", domLayout="normal")
    gb.configure_grid_options(getContextMenuItems=JsCode("""
        function(p){return ['copy','copyWithHeaders','paste','undo','redo'];}
    """))

    js_style = JsCode(f"""
        function(p){{
            const m = {json.dumps(cell_styles)};
            const k = p.data['ID'] + '|||' + p.colDef.field;
            return m[k] || null;
        }}
    """)

    for col in day_cols:
        if _is_employee:
            gb.configure_column(
                col,
                editable=False,
                cellStyle=js_style,
                valueFormatter=fmt12_js,
                flex=1
            )
        else:
            gb.configure_column(
                col,
                cellEditor="agSelectCellEditor",
                cellEditorParams={"values": shift_opts + ["OFF", ""]},
                cellStyle=js_style,
                valueFormatter=fmt12_js,
                flex=1
            )

    gb.configure_column(
        "Total",
        editable=False,
        valueGetter=total_js,
        cellRenderer=icon_renderer,
        cellStyle=JsCode("""
            function(p) {
                const v = Number(p.value || 0);
                let style = {
                    backgroundColor: '#FEF1E6',
                    fontWeight: 'bold',
                    textAlign: 'center'
                };
                if (v > 40) {
                    style.color = '#d9534f';
                }
                return style;
            }
        """),
        flex=1
    )

    # ---------- navigator & buttons ----------------------------------------
    nav_prev2, nav_range2, nav_next2 = st.columns([1,3,1])
    if nav_prev2.button("◀", key="sch_prev2"):
        st.session_state.sch_week_start -= timedelta(days=7); do_rerun()
    if nav_next2.button("▶", key="sch_next2"):
        st.session_state.sch_week_start += timedelta(days=7); do_rerun()
    nav_range2.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

    if _is_employee:
        undo_clicked = save_clicked = copy_clicked = False
    else:
        hdr_l, hdr_u, hdr_s, hdr_c = st.columns([6,1,1,2])
        with hdr_u: undo_clicked  = st.button("🔄 Undo")
        with hdr_s: save_clicked  = st.button("💾 Save")
        with hdr_c: copy_clicked  = st.button("📋 Copy Forward")

    # ---------- grid CSS ----------------------------------------------------
    st.markdown("""
        <style>
        .ag-theme-alpine .ag-cell,
        .ag-theme-alpine .ag-header-cell-label{
            font-size:13px!important;padding:4px 8px!important;
        }
        .ag-theme-alpine .ag-header-cell-label{
            font-weight:bold;justify-content:center!important;
        }
        .ag-theme-alpine .ag-cell,
        .ag-theme-alpine .ag-header-cell{
            box-shadow:inset -1px 0 #b3b3b3,inset 1px 0 #b3b3b3;
        }
        .ag-theme-alpine .ag-row,
        .ag-theme-alpine .ag-header-row{
            border-bottom:1px solid #dcdcdc!important;
        }
        .ag-theme-alpine .ag-root-wrapper{
            border:1px solid #a0a0a0;border-radius:4px;
        }
        .ag-theme-alpine .total-normal{font-weight:bold;}
        .ag-theme-alpine .total-over  {font-weight:bold;color:#d9534f;}
        </style>
    """, unsafe_allow_html=True)

    st.markdown("""
        <style>
        .ag-theme-streamlit .ag-cell,
        .ag-theme-streamlit .ag-header-cell-label {
            font-size:13px!important; padding:4px 8px!important;
        }
        .ag-theme-streamlit .ag-header-cell-label {
            font-weight:bold; justify-content:center!important;
        }
        .ag-theme-streamlit .ag-cell,
        .ag-theme-streamlit .ag-header-cell {
            box-shadow:inset -1px 0 #b3b3b3,inset 1px 0 #b3b3b3;
        }
        .ag-theme-streamlit .ag-row,
        .ag-theme-streamlit .ag-header-row {
            border-bottom:1px solid #dcdcdc!important;
        }
        .ag-theme-streamlit .ag-root-wrapper {
            border:1px solid #a0a0a0; border-radius:6px;
        }
        </style>
    """, unsafe_allow_html=True)


    grid_response = AgGrid(
        df_view,
        gridOptions=gb.build(),
        theme="streamlit",
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=True,
        domLayout="autoHeight",
        data_return_mode="AS_INPUT",
        update_mode="MODEL_CHANGED"
    )
    edited_df = grid_response["data"]
    try:
        selected_pos = selected_pos
    except NameError:
        selected_pos = None

    pos_id = (
        session.query(Position.id)
        .filter_by(name=selected_pos)
        .scalar()
        if selected_pos else None
    )

    # ---------- PLANNING PERIOD SUMMARY ------------------------------------
    def parse_shift_hours(s):
        try:
            if not s or str(s).strip().upper() == "OFF":
                return 0.0
            a, b = s.split("-")
            t0 = datetime.strptime(a.strip(), "%H:%M")
            t1 = datetime.strptime(b.strip(), "%H:%M")
            diff = (t1 - t0).seconds / 3600.0
            if diff <= 0:
                diff += 24.0                 # cross-midnight
            hrs = max(0.0, diff - 0.5)       # subtract 30-min break
            return round(hrs, 2)
        except:
            return 0.0

    summary = defaultdict(lambda: {"Shifts":0, "Total":0.0})
    for col in day_cols:
        for _, r in edited_df.iterrows():
            s = str(r[col]).strip()
            if s and s.upper() != "OFF":
                summary[col]["Shifts"] += 1
                summary[col]["Total"]  += parse_shift_hours(s)

    st.markdown("#### 🗓️ Schedule Summary")
    summary_df = pd.DataFrame({
        "Metric": ["Shifts", "Total Hours"],
        **{c: [int(summary[c]["Shifts"]), float(summary[c]["Total"])] for c in day_cols}
    }).set_index("Metric")

    summary_fmt = summary_df.copy()
    for c in summary_fmt.columns:
        summary_fmt[c] = summary_fmt[c].astype(int)

    st.dataframe(
        summary_fmt.style.format("{:.0f}").set_properties(**{"text-align": "right"}),
        height=120,
        use_container_width=True
    )

    visible_positions = [selected_pos] if selected_pos and selected_pos != "(Select)" else []
    st.session_state["planning_summary_all"] = st.session_state.get("planning_summary_all", {})
    for pos in visible_positions:
        st.session_state["planning_summary_all"][pos] = summary_df

    selected_pos = st.session_state.get("selected_pos")

    # ---------- PROJECTED / STANDARD HOURS (Totals Only) ----------------------
    if not selected_pos or selected_pos == "(Select)":
        st.info("Select a position above to see projected hours.")
        st.stop()

    std_rows = (
        session.query(LaborStandard)
        .join(Position, LaborStandard.position_id == Position.id)
        .filter(Position.name == selected_pos)
        .all()
    )
    if not std_rows:
        st.info("No labor standards defined for this position.")
        st.stop()

    hrs_per_fte = 8
    week_start  = st.session_state.get("sch_week_start", date.today())
    week_dates  = [week_start + timedelta(days=i) for i in range(7)]
    fmt_day     = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
    col_labels  = [d.strftime(fmt_day) for d in week_dates]

    # --- normalize function so names match regardless of case/spacing -------
    def _norm(s: str) -> str:
        return (s or "").strip().lower()

    # --- OPTIONAL: rules map used in your debug table (metric -> expr) ------
    # Expecting something like: {"Cars Parked": "[Studio Occupied Rooms] * 0.15", ...}
    kpi_rules = st.session_state.get("room_kpi_rules", {})
    kpi_rules_norm = {_norm(k): v for k, v in kpi_rules.items()}

    # --- load all base KPI values once (date -> {kpi: value}) ---------------
    def _load_kpi_map(model):
        rows = (
            session.query(model)
            .filter(model.date.in_(week_dates))
            .all()
        )
        m = {}
        for r in rows:
            d = r.date
            if d not in m:
                m[d] = {}
            m[d][_norm(r.kpi)] = float(r.value or 0)
        return m

    fc_all  = _load_kpi_map(RoomForecast)
    otb_all = _load_kpi_map(db.RoomOTBPickup)

    # --- evaluator for expressions with bracketed KPI names -----------------
    import re
    _bracket_re = re.compile(r"\[([^\]]+)\]")

    def _eval_expr(expr: str, values_for_day: dict) -> float:
        def _sub(match):
            key = _norm(match.group(1))
            return str(values_for_day.get(key, 0))
        safe = _bracket_re.sub(_sub, (expr or "0"))
        if re.fullmatch(r"[0-9.+\-*/() ]+", safe):
            try:
                return float(eval(safe, {"__builtins__": {}}, {}))
            except Exception:
                return 0.0
        return 0.0

    # --- resolve a metric value: direct value OR rule-based computed value ---
    def _resolve_metric(metric_name: str, day_values: dict) -> float:
        key = _norm(metric_name)
        if key in day_values:
            return day_values[key]
        expr = kpi_rules_norm.get(key)
        if expr:
            return _eval_expr(expr, day_values)
        return 0.0

    # ---------------------------- compute Projected --------------------------
    grand_tot = {lbl: {"Shifts": 0.0, "Hours": 0.0} for lbl in col_labels}

    for std in std_rows:
        metric      = (std.metric or "").strip()
        std_per_fte = float(std.standard or 0) or 0.0

        for d, lbl in zip(week_dates, col_labels):
            metric_val = _resolve_metric(metric, fc_all.get(d, {}))
            if metric_val and std_per_fte:
                shifts = round(metric_val / std_per_fte, 2)
                hours  = round(shifts * hrs_per_fte, 2)
            else:
                shifts = 0.0
                hours  = 0.0
            grand_tot[lbl]["Shifts"] += shifts
            grand_tot[lbl]["Hours"]  += hours

    df_tot = pd.DataFrame({
        "Metric": ["Total Shifts", "Total Hours"],
        **{
            lbl: [
                f'{grand_tot[lbl]["Shifts"]:.2f}',
                f'{grand_tot[lbl]["Hours"]:.2f}'
            ] for lbl in col_labels
        }
    }).set_index("Metric")

    df_tot_fmt = df_tot.copy()
    for c in df_tot_fmt.columns:
        df_tot_fmt[c] = df_tot_fmt[c].astype(float).round(0).astype(int)

    st.markdown("#### 📈 Projected / Standard Hours")
    st.dataframe(
        df_tot_fmt.style.format("{:.0f}").set_properties(**{"text-align": "right"}),
        height=120,
        use_container_width=True
    )

    st.session_state["projected_hours_all"] = st.session_state.get("projected_hours_all", {})
    for pos in visible_positions:
        st.session_state["projected_hours_all"][pos] = df_tot

    # ---------- OTB + Pickup Projected Hours (Totals Only) -------------------
    otb_tot = {lbl: {"Shifts": 0.0, "Hours": 0.0} for lbl in col_labels}

    for std in std_rows:
        metric      = (std.metric or "").strip()
        std_per_fte = float(std.standard or 0) or 0.0

        for d, lbl in zip(week_dates, col_labels):
            metric_val = _resolve_metric(metric, otb_all.get(d, {}))
            if metric_val and std_per_fte:
                shifts = round(metric_val / std_per_fte, 2)
                hours  = round(shifts * hrs_per_fte, 2)
            else:
                shifts = 0.0
                hours  = 0.0
            otb_tot[lbl]["Shifts"] += shifts
            otb_tot[lbl]["Hours"]  += hours

    df_otb = pd.DataFrame({
        "Metric": ["Total Shifts", "Total Hours"],
        **{
            lbl: [
                f'{otb_tot[lbl]["Shifts"]:.2f}',
                f'{otb_tot[lbl]["Hours"]:.2f}'
            ] for lbl in col_labels
        }
    }).set_index("Metric")

    #st.markdown("#### 📊 OTB + Pickup Projected Hours")
    #st.dataframe(df_otb, height=120, use_container_width=True)

    st.session_state["otb_hours_all"] = st.session_state.get("otb_hours_all", {})
    for pos in visible_positions:
        st.session_state["otb_hours_all"][pos] = df_otb
    # ---------- SAVE --------------------------------------------------------
    if save_clicked:
        old_rows = session.query(db.Schedule).filter(
            db.Schedule.emp_id.in_(sched_df["emp_id"]),
            db.Schedule.day.in_(week_dates_str)
        ).all()
        old_map = {(r.emp_id, r.day): r.shift_type for r in old_rows}

        new_map = {}
        for _, r in edited_df.iterrows():
            emp = int(r["emp_id"])
            for d, col in zip(week_dates, day_cols):
                val = str(r[col]).strip()
                if val:
                    new_map[(emp, d.strftime("%Y-%m-%d"))] = val

        backup_diffs(old_map, new_map)

        session.query(db.Schedule).filter(
            db.Schedule.emp_id.in_(sched_df["emp_id"]),
            db.Schedule.day.in_(week_dates_str)
        ).delete(synchronize_session=False)

        from datetime import datetime as _dt
        new_rows = [
            db.Schedule(
                emp_id=k[0],
                day=_dt.strptime(k[1], "%Y-%m-%d").date(),
                shift_type=v
            )
            for k, v in new_map.items()
        ]

        if new_rows:
            session.add_all(new_rows)
            session.commit()

        st.markdown("""
            <div style="
                position: fixed;
                bottom: 24px;
                right: 24px;
                background-color: #4CAF50;
                color: white;
                padding: 14px 20px;
                border-radius: 6px;
                font-size: 15px;
                box-shadow: 0px 4px 10px rgba(0,0,0,0.25);
                z-index: 1000;
                animation: fadeOut 5s forwards;
            ">
                ✅ Schedule saved.
            </div>
            <style>
                @keyframes fadeOut {
                    0%   { opacity: 1; }
                    80%  { opacity: 1; }
                    100% { opacity: 0; display: none; }
                }
            </style>
        """, unsafe_allow_html=True)

    # ---------- UNDO --------------------------------------------------------
    if undo_clicked:
        undo_last_change()

    # ---------- COPY FORWARD (always overwrite target week) -----------------
    if copy_clicked:
        src_rows = session.query(db.Schedule).filter(
            db.Schedule.emp_id.in_(sched_df["emp_id"]),
            db.Schedule.day.in_(week_dates_str)
        ).all()

        if not src_rows:
            st.info("There is nothing saved for the current week yet.")
        else:
            tgt_dates = [d + timedelta(days=7) for d in week_dates]

            session.query(db.Schedule).filter(
                db.Schedule.emp_id.in_(sched_df["emp_id"]),
                db.Schedule.day.in_(tgt_dates)
            ).delete(synchronize_session=False)

            hotel_name = st.session_state.get("user", {}).get("hotel_name")
            new_rows = [
                db.Schedule(
                    emp_id=r.emp_id,
                    day=(r.day if isinstance(r.day, date) else datetime.strptime(r.day, "%Y-%m-%d").date()) + timedelta(days=7),
                    shift_type=r.shift_type
                )
                for r in src_rows
            ]
            session.add_all(new_rows)
            session.commit()
            st.success("Copied current week to next week and overwrote any existing data.")
# ------------------ PAGE: COST AND OT MGMT ------------------
elif main_choice == "Cost and OT Mgmt":
      # ─────────────── Page Header ───────────────
      st.markdown("""
      <div class="la-title">
        <!-- Lucide: briefcase (fits Cost & OT Mgmt theme) -->
        <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
             viewBox="0 0 24 24" fill="none" stroke="currentColor"
             stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
          <rect x="2" y="7" width="20" height="14" rx="2" ry="2"></rect>
          <path d="M16 7V5a2 2 0 0 0-2-2h-4a2 2 0 0 0-2 2v2"></path>
        </svg>
        <span>Cost and OT Management</span>
      </div>
      <style>
        .la-title{
          display:flex; align-items:center; gap:10px;
          margin:0 0 10px 0; line-height:1;
        }
        .la-title svg{ color:#111; }
        .la-title span{ font-weight:700; font-size:1.6rem; }
      </style>
      """, unsafe_allow_html=True)

      # ─────────────── Tabs ───────────────
      tab1, tab2 = st.tabs(["🚨\uFE0E OT Risk", "📊\uFE0E Cost Mgmt"])

      with tab1:
            st.header("Overtime Risk Overview")
            # ─────────────── Filters + Week navigator ───────────────
            dept_df = refresh(db.Department).rename(columns={"name": "dept"})

            pos_df  = (
                  refresh(db.Position)
                  .merge(dept_df, left_on="department_id", right_on="id")
                  [["id_x", "name", "dept"]]
                  .rename(columns={"id_x": "id"})
            )

            pos_df = apply_manager_scope(
                  pos_df.rename(columns={
                        "dept": "department",
                        "name": "position"
                  })
            ).rename(columns={
                  "department": "dept",
                  "position": "name"
            })

            allowed_depts = sorted(
                  pos_df["dept"].dropna().unique().tolist()
            )

            allowed_positions = sorted(
                  pos_df["name"].dropna().unique().tolist()
            )

            dept_opts = ["(All)"] + sorted(pos_df["dept"].dropna().unique())
            col1, col2 = st.columns(2)
            with col1:
                  sel_dept = st.selectbox("Department", dept_opts, key="otrisk_dept")
            if sel_dept != "(All)":
                  pos_opts = ["(All)"] + sorted(
                        pos_df.loc[pos_df["dept"] == sel_dept, "name"].dropna().unique()
                  )
            else:
                  pos_opts = ["(All)"]
            with col2:
                  sel_pos = st.selectbox("Position", pos_opts, key="otrisk_pos")

            if "otrisk_week_start" not in st.session_state:
                  st.session_state.otrisk_week_start = date.today() + relativedelta(weekday=MO(-1))

            col_prev, col_range, col_next = st.columns([1, 3, 1])
            if col_prev.button("◀", key="otrisk_prev"):
                  st.session_state.otrisk_week_start -= timedelta(days=7)
            if col_next.button("▶", key="otrisk_next"):
                  st.session_state.otrisk_week_start += timedelta(days=7)

            week_start = st.session_state.otrisk_week_start
            week_end   = week_start + timedelta(days=6)
            col_range.markdown(f"### {week_start:%d %b %Y} – {week_end:%d %b %Y}")

            # ─────────────── Query actual hours for the week ───────────────
            q = (
                  session.query(
                        db.Actual.emp_id.label("Number"),
                        db.Actual.date.label("Business Date"),
                        (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
                        db.Position.name.label("Position"),
                        db.Department.name.label("Department")
                  )
                  .join(db.Position,   db.Actual.position_id == db.Position.id)
                  .join(db.Department, db.Position.department_id == db.Department.id)
                  .filter(db.Actual.date.between(week_start, week_end))
                  .filter(or_(db.Actual.hours != 0, db.Actual.ot_hours != 0))
            )

            if sel_dept == "(All)":
                  q = q.filter(db.Department.name.in_(allowed_depts))
            else:
                  q = q.filter(db.Department.name == sel_dept)

            if sel_pos == "(All)":
                  q = q.filter(db.Position.name.in_(allowed_positions))
            else:
                  q = q.filter(db.Position.name == sel_pos)

            raw = pd.DataFrame(q.all(),
                               columns=["Number", "Business Date", "Hours", "Position", "Department"])
            # === REMOVE ONLY VISUAL DUPLICATES BEFORE MERGE, KEEP BUSINESS DATE ===
            raw = raw.drop_duplicates(subset=["Number", "Business Date", "Hours", "Position", "Department"])

            if raw.empty:
                  st.warning("No actual hours data found for selected filters.")
                  st.stop()

            raw["Number"] = (
                  pd.to_numeric(raw["Number"], errors="coerce")
                    .fillna(0)
                    .astype(int)
                    .astype(str)
                    .str.zfill(5)
            )

            # Merge with employee names
            emp_df = refresh(db.Employee).copy()
            parts = emp_df["name"].astype(str).str.extract(
                  r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
            )
            emp_df["ID"]         = parts["ID"].fillna("").astype(str).str.strip().str.zfill(5)
            emp_df["First Name"] = parts["First_Name"].str.strip()
            emp_df["Last Name"]  = parts["Last_Name"].str.strip()
            emp_df["match_ID"]   = emp_df["ID"].astype(str).str.lstrip("0")
            raw["match_ID"]      = raw["Number"].astype(str).str.lstrip("0")

            # ── FIX: HARD DEDUPE — keep only ONE employee row per ID ──
            emp_df = emp_df.drop_duplicates(subset=["match_ID"], keep="last")

            merged = raw.merge(emp_df[["match_ID", "First Name", "Last Name"]],
                               on="match_ID", how="left")

            # ─────────────── Aggregate actuals ───────────────
            agg = merged.groupby(["Number", "First Name", "Last Name"]).agg(
                  total_hours=("Hours", "sum"),
                  days_worked=("Business Date", pd.Series.nunique)
            ).reset_index()

            # ─────────────── Pull Scheduled Days ───────────────
            sched_rows = (
                  session.query(db.Employee.name, db.Schedule.day, db.Schedule.shift_type)
                  .join(db.Employee, db.Employee.id == db.Schedule.emp_id)
                  .filter(db.Schedule.day.between(week_start, week_end))
                  .all()
            )
            sched_df = pd.DataFrame(sched_rows, columns=["name", "day", "shift_type"])

            if not sched_df.empty:
                  sched_df["shift_type"] = sched_df["shift_type"].fillna("").astype(str).str.upper().str.strip()
                  sched_df = sched_df[sched_df["shift_type"] != "OFF"]

                  sched_df["Number"] = sched_df["name"].str.extract(r"(\d+)$")[0].fillna("").str.zfill(5)
                  sched_df["day"] = pd.to_datetime(sched_df["day"])
                  merged["Business Date"] = pd.to_datetime(merged["Business Date"])

                  last_worked = (
                        merged.groupby("Number")["Business Date"]
                        .max()
                        .reset_index()
                        .rename(columns={"Business Date": "last_worked"})
                  )

                  sched_df = sched_df.merge(last_worked, on="Number", how="left")
                  sched_df["after_work"] = sched_df["day"] > sched_df["last_worked"]

                  sched_counts = sched_df.groupby("Number")["day"].nunique().reset_index(name="Days Scheduled")

                  sched_future = sched_df[sched_df["after_work"]].copy()
                  days_remaining = sched_future.groupby("Number")["day"].nunique().reset_index(name="Days Remaining")

                  def parse_shift_to_hours(shift_str):
                        try:
                              start, end = shift_str.split("-")
                              start_dt = pd.to_datetime(start, format="%H:%M")
                              end_dt = pd.to_datetime(end, format="%H:%M")
                              hours = (end_dt - start_dt).total_seconds() / 3600
                              if hours < 0:
                                    hours += 24
                              return max(0, hours - 0.5)
                        except:
                              return 0

                  sched_future["shift_hours"] = sched_future["shift_type"].apply(parse_shift_to_hours)
                  future_hours = sched_future.groupby("Number")["shift_hours"].sum().reset_index()
                  future_hours.rename(columns={"shift_hours": "Future Scheduled Hrs"}, inplace=True)

                  agg = agg.merge(sched_counts, how="left", on="Number")
                  agg = agg.merge(days_remaining, how="left", on="Number")
                  agg = agg.merge(future_hours, how="left", on="Number")

            else:
                  agg["Days Scheduled"] = 0
                  agg["Days Remaining"] = 0
                  agg["Future Scheduled Hrs"] = 0

            agg["Days Scheduled"] = agg["Days Scheduled"].fillna(0).astype(int)
            agg["Days Remaining"] = agg["Days Remaining"].fillna(0).astype(int)
            agg["Future Scheduled Hrs"] = agg["Future Scheduled Hrs"].fillna(0)

            agg.rename(columns={"days_worked": "Days Worked"}, inplace=True)
            agg["OT Risk"] = agg["total_hours"].apply(lambda h: "No Risk" if h <= 40 else "At Risk")
            agg["OT Risk %"] = agg["total_hours"].apply(
                  lambda h: "0%" if pd.isna(h) or h <= 40 else f"{round(((h - 40)/40)*100)}%"
            )
            agg["Projected OT"] = agg["total_hours"].apply(lambda h: max(round(h - 40, 2), 0))
            agg["Future Scheduled Hrs"] = pd.to_numeric(agg["Future Scheduled Hrs"], errors="coerce").fillna(0)
            agg["Total Hrs Worked + Schedule"] = (agg["total_hours"] + agg["Future Scheduled Hrs"]).round(2)

            def classify_ot_risk(row):
                  if row["Total Hrs Worked + Schedule"] <= 40:
                        return "No Risk"
                  elif row["Days Remaining"] > 0:
                        return "At Risk"
                  else:
                        return "OT"

            def estimate_risk_percent(row):
                  if row["Total Hrs Worked + Schedule"] <= 40:
                        return "0%"
                  if row["Days Remaining"] == 0:
                        return "100%"
                  elif row["Days Remaining"] == 1:
                        return "80%"
                  elif row["Days Remaining"] == 2:
                        return "60%"
                  elif row["Days Remaining"] == 3:
                        return "40%"
                  else:
                        return "20%"

            agg["OT Risk"] = agg.apply(classify_ot_risk, axis=1)
            agg["OT Risk %"] = agg.apply(estimate_risk_percent, axis=1)
            agg["Projected OT"] = agg["Total Hrs Worked + Schedule"].apply(lambda h: max(round(h - 40, 2), 0))
            # ── VISUAL-ONLY DEDUPE: keep one row per employee ID ──
            agg["vis_key"] = agg["Number"]
            agg = agg.drop_duplicates(subset=["vis_key"], keep="last")
            agg = agg.drop(columns=["vis_key"])

            # ─────────────── Merge Employee Rate + OT Cost ───────────────
            emp_df = refresh(db.Employee).copy()

            emp_df = apply_manager_scope(
                  emp_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})
            emp_df["ID"] = (
                  emp_df["name"]
                  .astype(str)
                  .str.extract(r"(\d+)$")[0]
                  .fillna("")
                  .str.zfill(5)
            )

            # Ensure consistent string types before merge
            emp_df["ID"] = emp_df["ID"].astype(str)

            if "hourly_rate" in emp_df.columns:
                  emp_df["rate"] = emp_df["hourly_rate"].fillna(0)
            else:
                  emp_df["rate"] = 0.00

            # Coerce Number to string too
            agg["Number"] = agg["Number"].astype(str)

            # Merge safely
            agg = agg.merge(emp_df[["ID", "rate"]], left_on="Number", right_on="ID", how="left")
            agg["rate"] = agg["rate"].fillna(0)
            agg["OT Cost"] = (agg["Projected OT"] * agg["rate"] * 1.5).round(2)
            # ─────────────── Display Table ───────────────
            show_cols = [
                  "Number", "First Name", "Last Name",
                  "OT Risk", "OT Risk %", "Projected OT", "OT Cost",
                  "Days Worked", "Days Remaining",
                  "Total Hrs Worked + Schedule"
            ]

            # hide OT Cost for Managers only
            role = (st.session_state.get("user", {}).get("role") or "").strip().lower()
            if role == "manager" and "OT Cost" in show_cols:
                  show_cols.remove("OT Cost")
            # --- VISUAL FILTER: SHOW ONLY UNIQUE NUMBERS ---
            agg = agg.drop_duplicates(subset=["Number"], keep="first")

            # ✅ sanity: make sure all show_cols exist in agg
            missing = [c for c in show_cols if c not in agg.columns]
            if missing:
                  st.error(f"Missing columns in data: {missing}")
                  st.caption(f"Available columns: {list(agg.columns)}")
            else:
                  from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

                  risk_icon_renderer = JsCode("""
                        function(params) {
                              if (params.value === "OT") {
                                    return "⛔ OT";
                              } else if (params.value === "At Risk") {
                                    return "⚠️ At Risk";
                              } else {
                                    return "✅ No Risk";
                              }
                        }
                  """)

                  risk_pct_renderer = JsCode("""
                        function(params) {
                              let pct = parseFloat(params.value.replace('%', ''));
                              if (pct >= 100) {
                                    return "🔴 " + params.value;
                              } else if (pct >= 80) {
                                    return "🟤 " + params.value;
                              } else if (pct >= 40) {
                                    return "🟠 " + params.value;
                              } else if (pct > 0) {
                                    return "🟡 " + params.value;
                              } else {
                                    return "🟢 " + params.value;
                              }
                        }
                  """)

                  currency_renderer = JsCode("""
                        function(params) {
                              if (!params.value || isNaN(params.value)) return "";
                              return "$" + Number(params.value).toFixed(2);
                        }
                  """)

                  # build grid on the validated columns
                  gb = GridOptionsBuilder.from_dataframe(agg[show_cols])
                  gb.configure_column("OT Risk", cellRenderer=risk_icon_renderer)
                  gb.configure_column("OT Risk %", cellRenderer=risk_pct_renderer)
                  if "OT Cost" in show_cols:
                        gb.configure_column("OT Cost", cellRenderer=currency_renderer, type=["numericColumn"])
                  gb.configure_default_column(editable=False, filter=True, resizable=True)

                  AgGrid(
                        agg[show_cols],
                        gridOptions=gb.build(),
                        fit_columns_on_grid_load=True,
                        height=420,
                        allow_unsafe_jscode=True,
                        enable_enterprise_modules=False
                  )
            st.session_state["ot_risk_final"] = agg.copy()
            st.session_state["ot_risk_filters"] = {
                  "week_start": week_start,
                  "week_end": week_end,
                  "department": sel_dept,
                  "position": sel_pos
            }

# ─────────────── Cost Mgmt Table ───────────────
      with tab2:
            st.header("Cost Management: FTE Variance")

            from collections import defaultdict

            dept_df = refresh(db.Department).rename(columns={"name": "dept"})

            pos_df  = (
                  refresh(db.Position)
                  .merge(dept_df, left_on="department_id", right_on="id")
                  [["id_x", "name", "dept"]]
                  .rename(columns={"id_x": "id"})
            )

            pos_df = apply_manager_scope(
                  pos_df.rename(columns={
                        "dept": "department",
                        "name": "position"
                  })
            ).rename(columns={
                  "department": "dept",
                  "position": "name"
            })

            dept_list = sorted(pos_df["dept"].dropna().unique())
            sel_dept = st.selectbox("Select Department", dept_list)

            if "cost_week_start" not in st.session_state:
                  st.session_state.cost_week_start = date.today() + relativedelta(weekday=MO(-1))

            prev_col, mid_col, next_col = st.columns([1,3,1])
            if prev_col.button("⬅", key="cost_prev"):
                  st.session_state.cost_week_start -= timedelta(days=7)
            if next_col.button("➡", key="cost_next"):
                  st.session_state.cost_week_start += timedelta(days=7)

            week_start = st.session_state.cost_week_start
            week_end   = week_start + timedelta(days=6)
            week_dates = [week_start + timedelta(days=i) for i in range(7)]
            fmt_day    = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
            day_cols   = [d.strftime(fmt_day) for d in week_dates]
            mid_col.markdown(f"### Week of {week_start:%b %d, %Y}")

            # ─────────────── Actual Hours ───────────────
            actual_q = (
                  session.query(
                        db.Position.name.label("Position"),
                        db.Actual.date.label("Date"),
                        func.sum(db.Actual.hours).label("Actual Hours")
                  )
                  .join(db.Position, db.Actual.position_id == db.Position.id)
                  .join(db.Department, db.Position.department_id == db.Department.id)
                  .filter(db.Department.name == sel_dept)
                  .filter(db.Actual.date.between(week_start, week_end))
                  .group_by(db.Position.name, db.Actual.date)
            )
            actual_df = pd.DataFrame(actual_q.all())
            if not actual_df.empty:
                  actual_df["FTE"] = actual_df["Actual Hours"] / 8
                  actual_df["Date Label"] = pd.to_datetime(actual_df["Date"]).dt.strftime(fmt_day)
            else:
                  actual_df = pd.DataFrame(columns=["Position", "Date", "Actual Hours", "FTE", "Date Label"])

            # ─────────────── Scheduled FTEs ───────────────
            emp_scope_df = refresh(db.Employee)

            emp_scope_df = apply_manager_scope(
                  emp_scope_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})

            sched_q = (
                  session.query(
                        db.Employee.role.label("Position"),
                        db.Schedule.day,
                        db.Schedule.shift_type
                  )
                  .join(db.Employee, db.Employee.id == db.Schedule.emp_id)
                  .filter(db.Employee.id.in_(emp_scope_df["id"].tolist()))
                  .filter(db.Employee.department == sel_dept)
                  .filter(db.Schedule.day.between(week_start, week_end))
            )
            sched_df_raw = pd.DataFrame(sched_q.all())

            def parse_hours(shift):
                  if not shift or shift.upper() == "OFF": return 0
                  try:
                        a,b = shift.split("-")
                        t0 = datetime.strptime(a.strip(), "%H:%M")
                        t1 = datetime.strptime(b.strip(), "%H:%M")
                        diff = (t1 - t0).seconds / 3600
                        return diff if diff > 0 else diff + 24
                  except:
                        return 0

            if not sched_df_raw.empty:
                  sched_df_raw["Hours"] = sched_df_raw["shift_type"].apply(parse_hours)
                  sched_df_raw["Date Label"] = pd.to_datetime(sched_df_raw["day"]).dt.strftime(fmt_day)
                  sched_group = sched_df_raw.groupby(["Position", "Date Label"])["Hours"].sum().reset_index()
                  sched_group["FTE"] = sched_group["Hours"] / 8
            else:
                  sched_group = pd.DataFrame(columns=["Position", "Date Label", "Hours", "FTE"])

            # ─────────────── OTB FTEs ───────────────
            std_q = (
                  session.query(
                        db.Position.name.label("Position"),
                        db.LaborStandard.metric,
                        db.LaborStandard.standard
                  )
                  .join(db.Position, db.Position.id == db.LaborStandard.position_id)
                  .join(db.Department, db.Position.department_id == db.Department.id)
                  .filter(db.Department.name == sel_dept)
            )
            std_df = pd.DataFrame(std_q.all())

            otb_q = (
                  session.query(
                        db.RoomOTBPickup.date.label("date"),
                        db.RoomOTBPickup.kpi.label("kpi"),
                        db.RoomOTBPickup.value.label("value")
                  )
                  .filter(db.RoomOTBPickup.date.between(week_start, week_end))
            )
            otb_df_raw = pd.DataFrame(otb_q.all())
            if otb_df_raw.empty:
                  otb_df_raw = pd.DataFrame(columns=["date", "kpi", "value"])

            otb_rows = []
            for _, row in std_df.iterrows():
                  pos, metric, std_val = row["Position"], row["metric"], row["standard"]
                  subset = otb_df_raw[otb_df_raw["kpi"].str.lower() == metric.lower()]
                  for _, r in subset.iterrows():
                        date_lbl = pd.to_datetime(r["date"]).strftime(fmt_day)
                        fte = (r["value"] / std_val) if std_val else 0
                        otb_rows.append({
                              "Position": pos,
                              "Date Label": date_lbl,
                              "OTB FTE": fte
                        })
            otb_df = pd.DataFrame(otb_rows)
            if otb_df.empty:
                  otb_df = pd.DataFrame(columns=["Position", "Date Label", "OTB FTE"])

            # ─────────────── No Data Message ───────────────
            if actual_df.empty and sched_group.empty and otb_df.empty:
                  st.warning("No Actual, Scheduled, or OTB data available for the selected week and department.")

            # ─────────────── Final Variance Table ───────────────
            pos_list = sorted(set(actual_df.get("Position", pd.Series()))
                              .union(sched_group.get("Position", pd.Series()))
                              .union(otb_df.get("Position", pd.Series())))

            data = []
            for pos in pos_list:
                  row = {"Position": pos}
                  for d in week_dates:
                        lbl = d.strftime(fmt_day)
                        otb_fte = otb_df[(otb_df["Position"] == pos) & (otb_df["Date Label"] == lbl)]["OTB FTE"].sum()
                        actual_fte = actual_df[(actual_df["Position"] == pos) & (actual_df["Date Label"] == lbl)]["FTE"].sum()
                        sched_fte  = sched_group[(sched_group["Position"] == pos) & (sched_group["Date Label"] == lbl)]["FTE"].sum()

                        if actual_fte > 0:
                              fte_var = actual_fte - otb_fte
                              color = "#ffe6cc"
                              font_color = "red"
                              arrow = "🔺"
                        else:
                              fte_var = sched_fte - otb_fte
                              color = "#e6f0ff"
                              font_color = "black"
                              arrow = "🔽" if fte_var < 0 else ""

                        cell = f"<div style='color:{font_color};background:{color};padding:4px;border-radius:4px;text-align:center'>{arrow} {fte_var:.2f}</div>"
                        row[lbl] = cell
                  data.append(row)

            var_df = pd.DataFrame(data)

            st.markdown("""
            <style>
                  .fancy-table td {
                        text-align: center;
                        vertical-align: middle;
                  }
            </style>
            """, unsafe_allow_html=True)

            st.write(var_df.to_html(escape=False, index=False, classes="fancy-table"), unsafe_allow_html=True)

            # ─────────────── Debug Outputs ───────────────
            with st.expander("🔍 Actual FTE DataFrame", expanded=False):
                  st.dataframe(actual_df, use_container_width=True)

            with st.expander("🔍 Scheduled FTE DataFrame", expanded=False):
                  st.dataframe(sched_group, use_container_width=True)

            with st.expander("🔍 OTB FTE DataFrame", expanded=False):
                  st.dataframe(otb_df, use_container_width=True)


# ─────────────── Reports Page ───────────────
elif main_choice == "Reports":
      from datetime import date, timedelta
      from dateutil.relativedelta import relativedelta, MO
      import pandas as pd
      import io
      from sqlalchemy import or_

      st.markdown("""
      <div class="la-title">
        <!-- Lucide: file-text (Reports) -->
        <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
             viewBox="0 0 24 24" fill="none" stroke="currentColor"
             stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
          <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path>
          <path d="M14 2v6h6"></path>
          <path d="M16 13H8"></path>
          <path d="M16 17H8"></path>
          <path d="M10 9H8"></path>
        </svg>
        <span>Reports</span>
      </div>
      <style>
        .la-title{
          display:flex; align-items:center; gap:10px;
          margin:0 0 10px 0; line-height:1;
        }
        .la-title svg{ color:#111; }      /* pure B/W via currentColor */
        .la-title span{ font-weight:700; font-size:1.6rem; }
      </style>
      """, unsafe_allow_html=True)

      # ─────────────── Report Type Selection (hide OT Risk for Managers) ───────────────
      role = (st.session_state.get("user", {}).get("role") or "").strip().lower()

      _all_reports = [
            "Department Schedule",
            "Labor Variance",
            "Forecast Variance",
            "OT Risk",
            "Productivity Index",
            "Labor Standards",
            "Schedule Variance",
            "Actual Hours",
            "Cost Mgmt FTE Variance",
      ]
      _visible_reports = [r for r in _all_reports if not (role == "manager" and r == "OT Risk")]

      report_type = st.selectbox("🧾\uFE0E Select Report Type", _visible_reports, key="report_type_sel")
      # ─────────────── Export Options ───────────────
      st.markdown("### 📤 Export Options")
      export_format = st.radio("Choose Export Format", ["Excel", "PDF", "CSV"], horizontal=True)

      # ─────────────── OT RISK REPORT ───────────────
      if report_type == "OT Risk":
            st.markdown("### OT Risk Report Filters")

            # independent keys for OT Risk date range (ACTUALS window)
            if "ot_risk_start" not in st.session_state:
                  st.session_state.ot_risk_start = date.today()
            if "ot_risk_end" not in st.session_state:
                  st.session_state.ot_risk_end = date.today()

            col1, col2 = st.columns(2)
            with col1:
                  sel_start = st.date_input(
                        "From Date",
                        value=st.session_state.ot_risk_start,
                        key="ot_risk_start"
                  )
            with col2:
                  sel_end = st.date_input(
                        "To Date",
                        value=st.session_state.ot_risk_end,
                        key="ot_risk_end"
                  )

            # keep same var names your pipeline uses for ACTUALS
            week_start = sel_start
            week_end   = sel_end

            # compute the full schedule week (Mon → Sun) that contains sel_start
            # e.g., if sel_start is Tue 2025-10-28, week is 2025-10-27 .. 2025-11-02
            sched_week_start = sel_start + relativedelta(weekday=MO(-1))
            sched_week_end   = sched_week_start + timedelta(days=6)

            if week_end < week_start:
                  st.error("End date cannot be earlier than start date.")
            else:
                  st.markdown(f"📅 **Selected Range (Actuals):** {week_start} to {week_end}")
                  st.markdown(f"🗓️ **Schedule Week:** {sched_week_start} to {sched_week_end}")

            # ---- Department / Position filters (same style as yours) ----
            dept_df = refresh(db.Department).rename(columns={"name": "Department"})
            dept_list = sorted(dept_df["Department"].dropna().unique())
            ot_risk_dept = st.selectbox("Select Department (OT Risk)", ["(All)"] + dept_list, key="ot_risk_dept")

            if ot_risk_dept == "(All)":
                  ot_risk_dept_id = None
            else:
                  try:
                        ot_risk_dept_id = dept_df[dept_df["Department"] == ot_risk_dept]["id"].values[0]
                  except IndexError:
                        ot_risk_dept_id = None

            pos_df = refresh(db.Position)  # uses columns: id, name, department_id
            if ot_risk_dept_id is None:
                  filtered_positions = pos_df
            else:
                  filtered_positions = pos_df[pos_df["department_id"] == ot_risk_dept_id]

            pos_list = sorted(filtered_positions["name"].dropna().unique())
            ot_risk_pos = st.selectbox("Select Position (OT Risk)", ["(All)"] + pos_list, key="ot_risk_pos")

            # From here, use:
            #   - week_start, week_end (ACTUALS inclusive range)
            #   - sched_week_start, sched_week_end (Mon→Sun SCHEDULE week)
            #   - ot_risk_dept_id (None = all)
            #   - ot_risk_pos ("(All)" = all)
            # ─────────────── Generate Button ───────────────
            if st.button("📊 Generate OT Risk Report"):

                  # === QUERY ACTUALS (respect selected date range) ===
                  q = (
                        session.query(
                              db.Actual.emp_id.label("Number"),
                              db.Actual.date.label("Business Date"),
                              (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
                              db.Position.name.label("Position"),
                              db.Department.name.label("Department")
                        )
                        .join(db.Position, db.Actual.position_id == db.Position.id)
                        .join(db.Department, db.Position.department_id == db.Department.id)
                        .filter(db.Actual.date.between(week_start, week_end))
                        .filter(or_(db.Actual.hours != 0, db.Actual.ot_hours != 0))
                  )
                  if ot_risk_dept != "(All)":
                        q = q.filter(db.Department.name == ot_risk_dept)
                  if ot_risk_pos != "(All)":
                        q = q.filter(db.Position.name == ot_risk_pos)

                  raw = pd.DataFrame(q.all(), columns=["Number", "Business Date", "Hours", "Position", "Department"])

                  if raw.empty:
                        st.warning("⚠️ No OT Risk data found. Cannot generate report.")
                  else:
                        # normalize employee IDs and names
                        raw["Number"] = pd.to_numeric(raw["Number"], errors="coerce").fillna(0).astype(int).astype(str).str.zfill(5)

                        emp_df = refresh(db.Employee).copy()
                        parts = emp_df["name"].astype(str).str.extract(r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)")
                        emp_df["ID"] = parts["ID"].fillna("").astype(str).str.strip().str.zfill(5)
                        emp_df["First Name"] = parts["First_Name"].str.strip()
                        emp_df["Last Name"] = parts["Last_Name"].str.strip()
                        emp_df["match_ID"] = emp_df["ID"].astype(str).str.lstrip("0")
                        raw["match_ID"] = raw["Number"].astype(str).str.lstrip("0")

                        merged = raw.merge(emp_df[["match_ID", "First Name", "Last Name"]], on="match_ID", how="left")

                        agg = merged.groupby(["Number", "First Name", "Last Name"]).agg(
                              total_hours=("Hours", "sum"),
                              days_worked=("Business Date", pd.Series.nunique)
                        ).reset_index()

                        # === QUERY SCHEDULE (ALWAYS the full Mon→Sun of selected week) ===
                        sched_rows = (
                              session.query(db.Employee.name, db.Schedule.day, db.Schedule.shift_type)
                              .join(db.Employee, db.Employee.id == db.Schedule.emp_id)
                              .filter(db.Schedule.day.between(sched_week_start, sched_week_end))
                              .all()
                        )
                        sched_df = pd.DataFrame(sched_rows, columns=["name", "day", "shift_type"])

                        if not sched_df.empty:
                              # normalize schedule rows
                              sched_df["shift_type"] = sched_df["shift_type"].fillna("").astype(str).str.upper().str.strip()
                              sched_df = sched_df[sched_df["shift_type"] != "OFF"]
                              sched_df["Number"] = sched_df["name"].str.extract(r"(\d+)$")[0].fillna("").str.zfill(5)
                              sched_df["day"] = pd.to_datetime(sched_df["day"])

                              # keep only employees that appeared in ACTUALS (dept/pos filters already applied)
                              valid_numbers = set(agg["Number"].astype(str).unique())
                              sched_df = sched_df[sched_df["Number"].isin(valid_numbers)]

                              merged["Business Date"] = pd.to_datetime(merged["Business Date"])
                              last_worked = (
                                    merged.groupby("Number")["Business Date"]
                                    .max().reset_index().rename(columns={"Business Date": "last_worked"})
                              )

                              sched_df = sched_df.merge(last_worked, on="Number", how="left")
                              # if an employee has no worked day in the selected actuals window,
                              # treat all scheduled days in the Mon→Sun week as "future"
                              sched_df["last_worked"] = sched_df["last_worked"].fillna(
                                    pd.to_datetime(week_start) - pd.Timedelta(days=1)
                              )
                              sched_df["after_work"] = sched_df["day"] > sched_df["last_worked"]

                              # days scheduled in the Mon→Sun week
                              sched_counts = sched_df.groupby("Number")["day"].nunique().reset_index(name="Days Scheduled")
                              # remaining days in that week after last worked day
                              sched_future = sched_df[sched_df["after_work"]].copy()
                              days_remaining = sched_future.groupby("Number")["day"].nunique().reset_index(name="Days Remaining")

                              def parse_shift_to_hours(shift_str):
                                    try:
                                          start, end = shift_str.split("-")
                                          start_dt = pd.to_datetime(start, format="%H:%M")
                                          end_dt = pd.to_datetime(end, format="%H:%M")
                                          hours = (end_dt - start_dt).total_seconds() / 3600
                                          if hours < 0:
                                                hours += 24
                                          # subtract 0.5h meal break if you want to keep your original logic
                                          return max(0, hours - 0.5)
                                    except:
                                          return 0

                              sched_future["shift_hours"] = sched_future["shift_type"].apply(parse_shift_to_hours)
                              future_hours = sched_future.groupby("Number")["shift_hours"].sum().reset_index()
                              future_hours.rename(columns={"shift_hours": "Future Scheduled Hrs"}, inplace=True)

                              agg = agg.merge(sched_counts, how="left", on="Number")
                              agg = agg.merge(days_remaining, how="left", on="Number")
                              agg = agg.merge(future_hours, how="left", on="Number")
                        else:
                              agg["Days Scheduled"] = 0
                              agg["Days Remaining"] = 0
                              agg["Future Scheduled Hrs"] = 0

                        # finalize fields
                        agg["Days Scheduled"] = agg["Days Scheduled"].fillna(0).astype(int)
                        agg["Days Remaining"] = agg["Days Remaining"].fillna(0).astype(int)
                        agg["Future Scheduled Hrs"] = agg["Future Scheduled Hrs"].fillna(0)
                        agg.rename(columns={"days_worked": "Days Worked"}, inplace=True)
                        agg["Total Hrs Worked + Schedule"] = (agg["total_hours"] + agg["Future Scheduled Hrs"]).round(2)

                        def classify_ot_risk(row):
                              if row["Total Hrs Worked + Schedule"] <= 40:
                                    return "No Risk"
                              elif row["Days Remaining"] > 0:
                                    return "At Risk"
                              else:
                                    return "OT"

                        def estimate_risk_percent(row):
                              if row["Total Hrs Worked + Schedule"] <= 40:
                                    return "0%"
                              if row["Days Remaining"] == 0:
                                    return "100%"
                              elif row["Days Remaining"] == 1:
                                    return "80%"
                              elif row["Days Remaining"] == 2:
                                    return "60%"
                              elif row["Days Remaining"] == 3:
                                    return "40%"
                              else:
                                    return "20%"

                        agg["OT Risk"] = agg.apply(classify_ot_risk, axis=1)
                        agg["OT Risk %"] = agg.apply(estimate_risk_percent, axis=1)
                        agg["Projected OT"] = agg["Total Hrs Worked + Schedule"].apply(lambda h: max(round(h - 40, 2), 0))

                        emp_df["ID"] = emp_df["ID"].astype(str)
                        agg["Number"] = agg["Number"].astype(str)
                        emp_df["rate"] = emp_df.get("hourly_rate", 0).fillna(0)
                        agg = agg.merge(emp_df[["ID", "rate"]], left_on="Number", right_on="ID", how="left")
                        agg["rate"] = agg["rate"].fillna(0)
                        agg["OT Cost"] = (agg["Projected OT"] * agg["rate"] * 1.5).round(2)

                        export_df = agg[[
                              "Number", "First Name", "Last Name", "Days Worked",
                              "Days Scheduled", "Days Remaining", "Total Hrs Worked + Schedule",
                              "OT Risk", "OT Risk %", "Projected OT", "OT Cost"
                        ]].rename(columns={"Total Hrs Worked + Schedule": "Total"})

                        # Export

                        # ===== DROP DUPLICATES BY NUMBER BEFORE EXPORT =====
                        export_df = export_df.drop_duplicates(subset=["Number"]).reset_index(drop=True)
                        # ===================================================
                        # Export
                        if export_df.empty:
                              st.warning("Report generated, but no OT Risk data was found for the selected filters.")
                        else:
                              if export_format == "Excel":
                                    import io
                                    output = io.BytesIO()
                                    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                                          # shift table down to row 7 (blank row included after metadata)
                                          export_df.to_excel(
                                                writer,
                                                sheet_name="OT Risk Report",
                                                startrow=7,
                                                startcol=1,
                                                index=False,
                                                header=False
                                          )
                                          workbook  = writer.book
                                          worksheet = writer.sheets["OT Risk Report"]

                                          # ===== Solid background =====
                                          solid_bg_fmt = workbook.add_format({'bg_color': '#FFFFFF', 'pattern': 1})
                                          worksheet.set_column(0, 51, None, solid_bg_fmt)
                                          for r in range(0, 2000):
                                                worksheet.set_row(r, None, solid_bg_fmt)

                                          worksheet.set_column(0, 0, 2)   # gutter

                                          # ===== User info =====
                                          user_obj = st.session_state.get("user") or {}
                                          username = user_obj.get("username") or user_obj.get("email") or "User"
                                          hotel_name = user_obj.get("hotel_name") or "Hotel"

                                          from datetime import datetime
                                          created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                                          # ===== BORDERLESS formats for top-right =====
                                          top_right_hotel_fmt = workbook.add_format({
                                                'bold': True, 'align': 'right', 'border': 0, 'bg_color': '#FFFFFF'
                                          })
                                          top_right_created_fmt = workbook.add_format({
                                                'align': 'right', 'border': 0, 'bg_color': '#FFFFFF'
                                          })

                                          # ===== Borderless metadata =====
                                          borderless_label_fmt = workbook.add_format({
                                                'bold': True,
                                                'border': 0,
                                                'bg_color': '#FFFFFF'
                                          })
                                          borderless_value_fmt = workbook.add_format({
                                                'border': 0,
                                                'bg_color': '#FFFFFF'
                                          })

                                          # ===== Title =====
                                          worksheet.write("B1", "OT Risk Report", workbook.add_format({'bold': True, 'font_size': 14}))

                                          # ===== Top-right hotel & creation (NO BORDER NOW) =====
                                          rightmost_col = 1 + len(export_df.columns) - 1

                                          worksheet.write(0, rightmost_col, hotel_name, top_right_hotel_fmt)
                                          worksheet.write(1, rightmost_col,
                                                f"Created by {username} on {created_str}",
                                                top_right_created_fmt)

                                          # ===== Left metadata + DATE FORMAT FIX =====

                                          worksheet.write("B3", "Department:", borderless_label_fmt)
                                          worksheet.write("C3", ot_risk_dept or "(All)", borderless_value_fmt)

                                          worksheet.write("B4", "Position:", borderless_label_fmt)
                                          worksheet.write("C4", ot_risk_pos or "(All)", borderless_value_fmt)

                                          # ---- DATE FORMAT UPDATED (MM-DD-YYYY) ----
                                          worksheet.write("B5", "Week (Schedule):", borderless_label_fmt)
                                          worksheet.write("C5",
                                                f"{sched_week_start:%m-%d-%Y} to {sched_week_end:%m-%d-%Y}",
                                                borderless_value_fmt)

                                          worksheet.write("B6", "Actuals Range:", borderless_label_fmt)
                                          worksheet.write("C6",
                                                f"{week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}",
                                                borderless_value_fmt)

                                          # ===== Blank row =====
                                          # row 6 is intentionally blank

                                          # ===== Table formats (Labor Variance style) =====
                                          border_fmt = {'border': 1, 'border_color': '#A6A6A6'}

                                          header_fmt = workbook.add_format({
                                                **border_fmt,
                                                'bold': True,
                                                'bg_color': '#16365C',
                                                'font_color': '#FFFFFF',
                                                'align': 'center',
                                                'pattern': 1
                                          })

                                          default_fmt = workbook.add_format({
                                                **border_fmt,
                                                'align': 'center'
                                          })

                                          first3_fmt = workbook.add_format({
                                                **border_fmt,
                                                'bold': True,
                                                'align': 'center'
                                          })

                                          last2_fmt = workbook.add_format({
                                                **border_fmt,
                                                'bg_color': '#FEF6F0',
                                                'align': 'center',
                                                'pattern': 1
                                          })

                                          green_fmt = workbook.add_format({
                                                **border_fmt,
                                                'font_color': '#008000',
                                                'align': 'center'
                                          })

                                          red_fmt = workbook.add_format({
                                                **border_fmt,
                                                'font_color': '#FF0000',
                                                'align': 'center'
                                          })

                                          dollar_red = workbook.add_format({
                                                **border_fmt,
                                                'num_format': '$#,##0.00',
                                                'font_color': '#FF0000',
                                                'align': 'center'
                                          })

                                          dollar_norm = workbook.add_format({
                                                **border_fmt,
                                                'num_format': '$#,##0.00',
                                                'align': 'center'
                                          })

                                          # ===== Header Row =====
                                          header_row = 7
                                          header_col = 1
                                          for col_num, col_name in enumerate(export_df.columns):
                                                worksheet.write(header_row, header_col + col_num, col_name, header_fmt)

                                          # ===== Data Rows =====
                                          data_start_row = 8
                                          data_start_col = 1

                                          for row_idx, row in export_df.iterrows():
                                                for col_idx, col_name in enumerate(export_df.columns):
                                                      val = row[col_name]

                                                      if col_idx <= 2:
                                                            fmt = first3_fmt
                                                      elif col_idx >= 9:
                                                            fmt = last2_fmt
                                                      elif col_name == "OT Risk" and val == "OT":
                                                            fmt = red_fmt
                                                      elif col_name == "OT Risk %" and val == "100%":
                                                            fmt = red_fmt
                                                      elif col_name == "OT Risk %" and val == "0%":
                                                            fmt = green_fmt
                                                      elif col_name == "OT Cost":
                                                            try:
                                                                  numeric = float(val)
                                                                  fmt = dollar_red if numeric > 0 else dollar_norm
                                                            except:
                                                                  fmt = default_fmt
                                                      else:
                                                            fmt = default_fmt

                                                      worksheet.write(
                                                            data_start_row + row_idx,
                                                            data_start_col + col_idx,
                                                            val,
                                                            fmt
                                                      )

                                          # ===== Summary Block =====
                                          summary_row = 3
                                          summary_col = 1 + (len(export_df.columns) - 2)

                                          summary_header_fmt = workbook.add_format({
                                                'bold': True,
                                                'bg_color': '#16365C',
                                                'font_color': '#FFFFFF',
                                                'border': 1,
                                                'align': 'center',
                                                'pattern': 1,
                                                'border_color': '#A6A6A6'
                                          })

                                          worksheet.merge_range(
                                                summary_row,
                                                summary_col,
                                                summary_row,
                                                summary_col + 1,
                                                "Total",
                                                summary_header_fmt
                                          )

                                          worksheet.write(summary_row + 1, summary_col, "Projected OT",
                                                workbook.add_format({'border': 1, 'border_color': '#A6A6A6'}))
                                          worksheet.write(summary_row + 1, summary_col + 1,
                                                export_df["Projected OT"].sum(),
                                                dollar_norm)

                                          worksheet.write(summary_row + 2, summary_col, "OT Cost",
                                                workbook.add_format({'border': 1, 'border_color': '#A6A6A6'}))
                                          worksheet.write(summary_row + 2, summary_col + 1,
                                                export_df["OT Cost"].sum(),
                                                dollar_norm)

                                          # ===== Footer =====
                                          worksheet.write(
                                                data_start_row + len(export_df) + 2,
                                                1,
                                                "Confidential | 2025 Labor Pilot",
                                                workbook.add_format({'align': 'left'})
                                          )

                                          # ===== Auto column widths =====
                                          for i, col in enumerate(export_df.columns):
                                                max_width = max(len(str(col)), export_df[col].astype(str).str.len().max())
                                                worksheet.set_column(1 + i, 1 + i, max(max_width + 2, 12))

                                    st.download_button(
                                          "⬇️ Download Excel",
                                          data=output.getvalue(),
                                          file_name=f"OT_Risk_Report_{sched_week_start:%Y-%m-%d}_to_{sched_week_end:%Y-%m-%d}.xlsx",
                                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )

                              elif export_format == "PDF":
                                    from reportlab.platypus import (
                                          SimpleDocTemplate, Table, TableStyle,
                                          Paragraph, Spacer
                                    )
                                    from reportlab.lib.pagesizes import landscape, letter
                                    from reportlab.lib.styles import getSampleStyleSheet
                                    from reportlab.lib import colors
                                    from datetime import datetime
                                    import io

                                    buffer = io.BytesIO()
                                    doc = SimpleDocTemplate(
                                          buffer,
                                          pagesize=landscape(letter),
                                          leftMargin=28,
                                          rightMargin=28,
                                          topMargin=24,
                                          bottomMargin=24
                                    )

                                    styles = getSampleStyleSheet()
                                    elements = []

                                    # ---------- TITLE ----------
                                    elements.append(Paragraph("<b>OT Risk Report</b>", styles["Heading2"]))
                                    elements.append(Spacer(1, 6))

                                    # ---------- LEFT METADATA ----------
                                    left_meta = [
                                          Paragraph(f"<b>Department:</b> {ot_risk_dept or '(All)'}", styles["Normal"]),
                                          Paragraph(f"<b>Position:</b> {ot_risk_pos or '(All)'}", styles["Normal"]),
                                          Paragraph(f"<b>Week (Schedule):</b> {sched_week_start:%m-%d-%Y} to {sched_week_end:%m-%d-%Y}", styles["Normal"]),
                                          Paragraph(f"<b>Actuals Range:</b> {week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}", styles["Normal"]),
                                    ]

                                    # ---------- RIGHT SUMMARY BOX (MATCH EXCEL) ----------
                                    summary_data = [
                                          ["Total", ""],   # merge next cell
                                          ["Projected OT", f"{export_df['Projected OT'].sum():.2f}"],
                                          ["OT Cost", f"${export_df['OT Cost'].sum():,.2f}"]
                                    ]

                                    summary_table = Table(summary_data, colWidths=[90, 70])
                                    summary_table.setStyle(TableStyle([
                                          ("SPAN", (0, 0), (1, 0)),  # merge Total + blank cell
                                          ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                                          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                          ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                          ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                          ("BACKGROUND", (0, 1), (-1, -1), colors.white),  # remove orange background
                                          ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                                          ("FONTSIZE", (0, 0), (-1, -1), 9),
                                          ("TOPPADDING", (0, 0), (-1, -1), 3),
                                          ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                                    ]))

                                    # ---------- 2 COLUMN LAYOUT ----------
                                    layout_data = [
                                          [left_meta, summary_table]
                                    ]

                                    layout_table = Table(layout_data, colWidths=[None, 160])
                                    layout_table.setStyle(TableStyle([
                                          ("VALIGN", (0, 0), (-1, -1), "TOP"),
                                          ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                                          ("LEFTPADDING", (0, 0), (-1, -1), 0),
                                          ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                                          ("TOPPADDING", (0, 0), (-1, -1), 0),
                                          ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                                    ]))

                                    elements.append(layout_table)
                                    elements.append(Spacer(1, 16))

                                    # ---------- TABLE DATA ----------
                                    pdf_data = [export_df.columns.tolist()] + export_df.values.tolist()
                                    table = Table(pdf_data)

                                    table_style = TableStyle([
                                          ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                                          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                          ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                          ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                          ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                                          ("FONTSIZE", (0, 0), (-1, -1), 9),
                                          ("TOPPADDING", (0, 0), (-1, -1), 4),
                                          ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                                    ])

                                    # last 2 columns shading (#FEF6F0)
                                    last2_start = len(export_df.columns) - 2
                                    last2_end = len(export_df.columns) - 1

                                    table_style.add(
                                          "BACKGROUND",
                                          (last2_start, 1),
                                          (last2_end, len(pdf_data) - 1),
                                          colors.HexColor("#FEF6F0")
                                    )

                                    table.setStyle(table_style)
                                    elements.append(table)

                                    # ---------- HEADER / FOOTER ----------
                                    user_obj = st.session_state.get("user") or {}
                                    username = user_obj.get("username") or user_obj.get("email") or "User"
                                    hotel_name = user_obj.get("hotel_name") or "Hotel"
                                    created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                                    def _first_page(canvas, doc):
                                          canvas.saveState()
                                          canvas.setFont("Helvetica-Bold", 10)
                                          canvas.drawRightString(
                                                doc.pagesize[0] - doc.rightMargin,
                                                doc.pagesize[1] - 40,     # moved down 20px
                                                hotel_name
                                          )
                                          canvas.setFont("Helvetica", 8)
                                          canvas.drawRightString(
                                                doc.pagesize[0] - doc.rightMargin,
                                                doc.pagesize[1] - 54,     # moved down 20px
                                                f"Created by {username} on {created_str}"
                                          )
                                          canvas.setFont("Helvetica", 8)
                                          canvas.drawString(
                                                doc.leftMargin,
                                                doc.bottomMargin - 12,
                                                "Confidential | 2025 Labor Pilot"
                                          )
                                          canvas.drawRightString(
                                                doc.pagesize[0] - doc.rightMargin,
                                                doc.bottomMargin - 12,
                                                f"Page | {canvas.getPageNumber()}"
                                          )
                                          canvas.restoreState()
                                    def _later_pages(canvas, doc):
                                          canvas.saveState()
                                          canvas.setFont("Helvetica", 8)
                                          canvas.drawString(doc.leftMargin, doc.bottomMargin - 12,
                                                            "Confidential | 2025 Labor Pilot")
                                          canvas.drawRightString(doc.pagesize[0] - doc.rightMargin,
                                                                 doc.bottomMargin - 12,
                                                                 f"Page | {canvas.getPageNumber()}")
                                          canvas.restoreState()

                                    doc.build(elements, onFirstPage=_first_page, onLaterPages=_later_pages)

                                    buffer.seek(0)
                                    st.download_button(
                                          "⬇️ Download PDF",
                                          buffer.getvalue(),
                                          file_name=f"OT_Risk_Report_{sched_week_start:%Y-%m-%d}_to_{sched_week_end:%Y-%m-%d}.pdf",
                                          mime="application/pdf"
                                    )

      # ─────────────── ACTUAL HOURS REPORT ───────────────
      if report_type == "Actual Hours":
            st.markdown("### Actual Hours Report Filters")

            # ---------------------------------------------------------
            # WEEK SELECTION (MATCHES LABOR VARIANCE)
            # ---------------------------------------------------------
            if "actual_hours_week_date" not in st.session_state:
                  st.session_state.actual_hours_week_date = date.today()

            sel_day = st.date_input(
                  "Select any date in the week",
                  value=st.session_state.actual_hours_week_date,
                  key="actual_hours_week_date"
            )

            week_start = sel_day - timedelta(days=sel_day.weekday())
            week_end   = week_start + timedelta(days=6)

            st.markdown(
                  f"📅 **Selected Week:** {week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}"
            )

            # ---------------------------------------------------------
            # DEPARTMENT / POSITION (SCOPED – ACTUAL HOURS)
            # ---------------------------------------------------------
            base_df = (
                  refresh(db.Position)
                  .merge(
                        refresh(db.Department),
                        left_on="department_id",
                        right_on="id",
                        suffixes=("", "_dept")
                  )
                  [["name", "name_dept"]]
                  .rename(columns={
                        "name": "position",
                        "name_dept": "department"
                  })
            )

            # 🔒 APPLY MANAGER SCOPE FIRST
            base_df = apply_manager_scope(base_df)

            dept_opts = ["(All)"] + sorted(
                  base_df["department"].dropna().unique().tolist()
            )

            act_dept = st.selectbox(
                  "Select Department (Actual Hours)",
                  dept_opts,
                  key="actual_hours_dept"
            )

            if act_dept != "(All)":
                  pos_opts = ["(All)"] + sorted(
                        base_df.loc[
                              base_df["department"] == act_dept,
                              "position"
                        ]
                        .dropna()
                        .unique()
                        .tolist()
                  )
            else:
                  pos_opts = ["(All)"] + sorted(
                        base_df["position"]
                        .dropna()
                        .unique()
                        .tolist()
                  )

            act_pos = st.selectbox(
                  "Select Position (Actual Hours)",
                  pos_opts,
                  key="actual_hours_pos"
            )

            # ─────────────── Generate Button ───────────────
            if st.button("📊 Generate Actual Hours Report"):

                  # ---------------------------------------------------------
                  # QUERY ACTUALS (ALWAYS SCOPED)
                  # ---------------------------------------------------------
                  q = (
                        session.query(
                              db.Actual.emp_id.label("Number"),
                              db.Actual.date.label("Business Date"),
                              (db.Actual.hours + db.Actual.ot_hours).label("Hours"),
                              db.Position.name.label("Position"),
                              db.Department.name.label("Department")
                        )
                        .join(db.Position,   db.Actual.position_id == db.Position.id)
                        .join(db.Department, db.Position.department_id == db.Department.id)
                        .filter(db.Actual.date.between(week_start, week_end))
                        .filter(
                              db.Department.name.in_(
                                    base_df["department"].unique().tolist()
                              )
                        )
                        .filter(
                              db.Position.name.in_(
                                    base_df["position"].unique().tolist()
                              )
                        )
                  )

                  if act_dept != "(All)":
                        q = q.filter(db.Department.name == act_dept)

                  if act_pos != "(All)":
                        q = q.filter(db.Position.name == act_pos)

                  raw = pd.DataFrame(
                        q.all(),
                        columns=["Number", "Business Date", "Hours", "Position", "Department"]
                  )

                  if raw.empty:
                        st.warning("⚠️ No Actual Hours data found for the selected week.")
                        st.stop()

                  # ---------------------------------------------------------
                  # NORMALIZE IDS
                  # ---------------------------------------------------------
                  raw["Number"] = (
                        pd.to_numeric(raw["Number"], errors="coerce")
                          .fillna(0)
                          .astype(int)
                          .astype(str)
                          .str.zfill(5)
                  )

                  # ✅ FIX: Convert datetime → DATE BEFORE pivot
                  raw["Business Date"] = pd.to_datetime(raw["Business Date"]).dt.date

                  # ---------------------------------------------------------
                  # PIVOT TO MON–SUN
                  # ---------------------------------------------------------
                  pivot = (
                        raw.pivot_table(
                              index="Number",
                              columns="Business Date",
                              values="Hours",
                              aggfunc="sum",
                              fill_value=0
                        )
                        .reset_index()
                  )

                  pivot["ID"] = pivot["Number"]
                  pivot.drop(columns="Number", inplace=True)

                  # ---------------------------------------------------------
                  # ATTACH EMPLOYEE NAMES
                  # ---------------------------------------------------------
                  emp_df = refresh(db.Employee).copy()
                  parts = emp_df["name"].astype(str).str.extract(
                        r"^\s*(?P<Last_Name>[^,]+),\s*(?P<First_Name>[^\d]+?)\s+(?P<ID>\d+)"
                  )

                  emp_df["ID"]         = parts["ID"].fillna("").astype(str).str.strip().str.zfill(5)
                  emp_df["First Name"] = parts["First_Name"].str.strip()
                  emp_df["Last Name"]  = parts["Last_Name"].str.strip()

                  pivot["match_ID"]  = pivot["ID"].str.lstrip("0")
                  emp_df["match_ID"] = emp_df["ID"].str.lstrip("0")

                  emp_df = emp_df.drop_duplicates(subset=["match_ID"], keep="last")

                  pivot = pivot.merge(
                        emp_df[["match_ID", "First Name", "Last Name"]],
                        on="match_ID",
                        how="left"
                  )

                  pivot.drop(columns="match_ID", inplace=True)

                  # ---------------------------------------------------------
                  # ✅ FIXED DAY LIST (NO .date() CALL)
                  # ---------------------------------------------------------
                  days = [week_start + timedelta(d) for d in range(7)]
                  fmt  = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
                  fmt_cols = [(week_start + timedelta(d)).strftime(fmt) for d in range(7)]

                  rename_map = {d: label for d, label in zip(days, fmt_cols)}
                  pivot.rename(columns=rename_map, inplace=True)

                  for label in fmt_cols:
                        if label not in pivot.columns:
                              pivot[label] = 0.0

                  export_df = pivot[["ID", "First Name", "Last Name"] + fmt_cols]

                  # =========================================================
                  # =================== EXCEL EXPORT =========================
                  # =========================================================
                  if export_format == "Excel":
                        import io
                        output = io.BytesIO()

                        # -----------------------------------------------------
                        # ✅ ROUND ALL NUMBERS TO 2 DECIMALS
                        # -----------------------------------------------------
                        numeric_cols = export_df.columns[3:]
                        export_df[numeric_cols] = export_df[numeric_cols].round(2)

                        # -----------------------------------------------------
                        # ✅ ADD TOTAL COLUMN (ROW SUM)
                        # -----------------------------------------------------
                        export_df["Total"] = export_df[numeric_cols].sum(axis=1).round(2)

                        # -----------------------------------------------------
                        # ✅ ADD TOTAL ROW (ONLY FIRST COL = "TOTAL")
                        # -----------------------------------------------------
                        total_row = {}
                        for col in export_df.columns:
                              if col == export_df.columns[0]:
                                    total_row[col] = "TOTAL"
                              elif col in ["First Name", "Last Name"]:
                                    total_row[col] = ""
                              else:
                                    total_row[col] = round(export_df[col].sum(), 2)

                        export_df = pd.concat(
                              [export_df, pd.DataFrame([total_row])],
                              ignore_index=True
                        )

                        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                              export_df.to_excel(
                                    writer,
                                    sheet_name="Actual Hours",
                                    startrow=7,
                                    startcol=1,
                                    index=False,
                                    header=False
                              )

                              workbook  = writer.book
                              worksheet = writer.sheets["Actual Hours"]

                              # ===== Solid background =====
                              solid_bg_fmt = workbook.add_format({'bg_color': '#FFFFFF', 'pattern': 1})
                              worksheet.set_column(0, 60, None, solid_bg_fmt)
                              for r in range(0, 3000):
                                    worksheet.set_row(r, None, solid_bg_fmt)

                              worksheet.set_column(0, 0, 2)

                              user_obj = st.session_state.get("user") or {}
                              username = user_obj.get("username") or user_obj.get("email") or "User"
                              hotel_name = user_obj.get("hotel_name") or "Hotel"

                              from datetime import datetime
                              created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                              # ===== TOP-RIGHT (MATCH OT RISK) =====
                              top_right_hotel_fmt = workbook.add_format({
                                    'bold': True, 'align': 'right', 'border': 0, 'bg_color': '#FFFFFF'
                              })
                              top_right_created_fmt = workbook.add_format({
                                    'align': 'right', 'border': 0, 'bg_color': '#FFFFFF'
                              })

                              rightmost_col = 1 + len(export_df.columns) - 1
                              worksheet.write(0, rightmost_col, hotel_name, top_right_hotel_fmt)
                              worksheet.write(
                                    1, rightmost_col,
                                    f"Created by {username} on {created_str}",
                                    top_right_created_fmt
                              )

                              # ========== METADATA (LEFT ALIGNED, NO BORDERS – MATCH DEPT SCHEDULE) ==========

                              # ---- Title (merged like other report) ----
                              worksheet.merge_range("B1:E1", "Actual Hours Report", workbook.add_format({
                                    'bold': True,
                                    'font_size': 14,
                                    'align': 'left',
                                    'border': 0,
                                    'bg_color': '#FFFFFF'
                              }))

                              # ---- Label + Value formats ----
                              meta_label_fmt = workbook.add_format({
                                    'bold': True,
                                    'align': 'left',
                                    'border': 0,
                                    'bg_color': '#FFFFFF'
                              })

                              meta_value_fmt = workbook.add_format({
                                    'align': 'left',
                                    'border': 0,
                                    'bg_color': '#FFFFFF'
                              })

                              # ---- Department ----
                              worksheet.write("B3", "Department:", meta_label_fmt)
                              worksheet.write("C3", act_dept or "(All)", meta_value_fmt)

                              # ---- Position ----
                              worksheet.write("B4", "Position:", meta_label_fmt)
                              worksheet.write("C4", act_pos or "(All)", meta_value_fmt)

                              # ---- Week ----
                              worksheet.write("B5", "Week:", meta_label_fmt)
                              worksheet.write(
                                    "C5",
                                    f"{week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}",
                                    meta_value_fmt
                              )

                              # ===== Header / Data Formatting =====
                              border_fmt = {'border': 1, 'border_color': '#A6A6A6'}

                              header_fmt = workbook.add_format({
                                    **border_fmt,
                                    'bold': True,
                                    'bg_color': '#16365C',
                                    'font_color': '#FFFFFF',
                                    'align': 'center'
                              })

                              default_fmt = workbook.add_format({**border_fmt, 'align': 'center'})
                              first3_fmt  = workbook.add_format({**border_fmt, 'bold': True, 'align': 'center'})

                              # ✅ Total Column = #FEF6F0
                              total_col_fmt = workbook.add_format({
                                    **border_fmt,
                                    'bg_color': '#FEF6F0',
                                    'align': 'center'
                              })

                              # ✅ Total Row = #E6EEF8
                              total_row_fmt = workbook.add_format({
                                    **border_fmt,
                                    'bold': True,
                                    'bg_color': '#E6EEF8',
                                    'align': 'center'
                              })

                              # ===== Header Row =====
                              header_row = 7
                              header_col = 1
                              for col_num, col_name in enumerate(export_df.columns):
                                    worksheet.write(header_row, header_col + col_num, col_name, header_fmt)

                              # ===== Data Rows =====
                              last_row_index = len(export_df) - 1
                              total_col_index = export_df.columns.get_loc("Total")

                              for row_idx, row in export_df.iterrows():
                                    for col_idx, col_name in enumerate(export_df.columns):
                                          val = row[col_name]

                                          is_total_row = (row_idx == last_row_index)
                                          is_total_col = (col_idx == total_col_index)

                                          if is_total_row:
                                                fmt = total_row_fmt
                                          elif is_total_col:
                                                fmt = total_col_fmt
                                          elif col_idx <= 2:
                                                fmt = first3_fmt
                                          else:
                                                fmt = default_fmt

                                          worksheet.write(8 + row_idx, 1 + col_idx, val, fmt)

                              # ===== Auto Column Widths =====
                              for i, col in enumerate(export_df.columns):
                                    max_width = max(len(str(col)), export_df[col].astype(str).str.len().max())
                                    worksheet.set_column(1 + i, 1 + i, max(max_width + 2, 12))

                              # ✅ Confidential Footer ONE ROW BELOW TABLE
                              footer_row = 8 + len(export_df) + 1
                              worksheet.write(
                                    footer_row,
                                    1,
                                    "Confidential | 2025 Labor Pilot",
                                    workbook.add_format({'align': 'left'})
                              )

                        st.download_button(
                              "⬇️ Download Excel",
                              data=output.getvalue(),
                              file_name=f"Actual_Hours_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                  # =========================================================
                  # ===================== PDF EXPORT =========================
                  # =========================================================
                  elif export_format == "PDF":
                        import io
                        from datetime import datetime
                        from reportlab.lib.pagesizes import landscape, letter
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors

                        # -----------------------------------------------------
                        # ✅ FORCE Total column + Total row (same as Excel)
                        # -----------------------------------------------------
                        numeric_cols = export_df.columns[3:]
                        export_df[numeric_cols] = export_df[numeric_cols].round(2)

                        if "Total" not in export_df.columns:
                              export_df["Total"] = export_df[numeric_cols].sum(axis=1).round(2)

                        total_row = {}
                        for col in export_df.columns:
                              if col == export_df.columns[0]:
                                    total_row[col] = "TOTAL"
                              elif col in ["First Name", "Last Name"]:
                                    total_row[col] = ""
                              else:
                                    total_row[col] = round(export_df[col].sum(), 2)

                        export_df = pd.concat(
                              [export_df, pd.DataFrame([total_row])],
                              ignore_index=True
                        )

                        # -----------------------------------------------------
                        # ✅ Build PDF table data
                        # -----------------------------------------------------
                        pdf_data = [export_df.columns.tolist()] + export_df.values.tolist()

                        total_row_index = len(pdf_data) - 1
                        total_col_index = len(pdf_data[0]) - 1

                        # -----------------------------------------------------
                        # ✅ PDF Document (LANDSCAPE + STANDARD MARGINS)
                        # -----------------------------------------------------
                        buffer = io.BytesIO()
                        doc = SimpleDocTemplate(
                              buffer,
                              pagesize=landscape(letter),
                              leftMargin=28, rightMargin=28, topMargin=24, bottomMargin=24
                        )

                        styles = getSampleStyleSheet()
                        elements = []

                        # -----------------------------------------------------
                        # ✅ HEADER META TEXT (MATCHES YOUR STANDARD)
                        # -----------------------------------------------------
                        elements.append(Paragraph("Actual Hours Report", styles["Heading2"]))
                        elements.append(Spacer(1, 6))

                        elements.append(
                              Paragraph(
                                    f"<b>Department:</b> {st.session_state.get('actual_hours_dept', 'All')}",
                                    styles["Normal"]
                              )
                        )
                        elements.append(
                              Paragraph(
                                    f"<b>Position:</b> {st.session_state.get('actual_hours_pos', 'All')}",
                                    styles["Normal"]
                              )
                        )
                        elements.append(
                              Paragraph(
                                    f"<b>Week:</b> {week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}",
                                    styles["Normal"]
                              )
                        )

                        elements.append(Spacer(1, 10))

                        # -----------------------------------------------------
                        # ✅ ACTUAL HOURS TABLE
                        # -----------------------------------------------------
                        table = Table(
                              pdf_data,
                              repeatRows=1,
                              hAlign="CENTER",
                        )

                        style = TableStyle([
                              ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                              ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("FONTSIZE", (0, 0), (-1, -1), 9),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                              ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ])

                        # ✅ First 3 columns bold
                        n_rows = len(pdf_data) - 1
                        for r in range(1, n_rows + 1):
                              for c_idx in range(0, 3):
                                    style.add("FONTNAME", (c_idx, r), (c_idx, r), "Helvetica-Bold")

                        # ✅ Total COLUMN shading (#FEF6F0)
                        style.add(
                              "BACKGROUND",
                              (total_col_index, 1),
                              (total_col_index, total_row_index - 1),
                              colors.HexColor("#FEF6F0")
                        )

                        # ✅ Total ROW shading (#E6EEF8 + bold)
                        style.add(
                              "BACKGROUND",
                              (0, total_row_index),
                              (total_col_index, total_row_index),
                              colors.HexColor("#E6EEF8")
                        )
                        for c_idx in range(0, len(export_df.columns)):
                              style.add(
                                    "FONTNAME",
                                    (c_idx, total_row_index),
                                    (c_idx, total_row_index),
                                    "Helvetica-Bold"
                              )

                        table.setStyle(style)
                        elements.append(table)

                        # -----------------------------------------------------
                        # ✅ HEADER / FOOTER (MATCHES YOUR OTHER REPORTS)
                        # -----------------------------------------------------
                        user_obj = st.session_state.get("user") or {}
                        username = (
                              user_obj.get("username")
                              or user_obj.get("email")
                              or st.session_state.get("username")
                              or "User"
                        )
                        hotel_name = (
                              user_obj.get("hotel_name")
                              or st.session_state.get("hotel_name")
                              or "Hotel"
                        )
                        created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                        def _first_page(canvas, doc):
                              canvas.saveState()

                              # ----- top-right: hotel + created by -----
                              x_right = doc.pagesize[0] - doc.rightMargin
                              y_top   = doc.pagesize[1] - doc.topMargin - 10

                              canvas.setFont("Helvetica-Bold", 10)
                              canvas.drawRightString(x_right, y_top, str(hotel_name))

                              canvas.setFont("Helvetica", 8)
                              canvas.drawRightString(
                                    x_right,
                                    y_top - 12,
                                    f"Created by {username} on {created_str}"
                              )

                              # ----- bottom-left: confidential -----
                              canvas.setFont("Helvetica", 8)
                              canvas.drawString(
                                    doc.leftMargin,
                                    doc.bottomMargin - 14,
                                    "Confidential | 2025 Labor Pilot"
                              )

                              # ----- bottom-right: page number -----
                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.bottomMargin - 14,
                                    f"Page | {canvas.getPageNumber()}"
                              )

                              canvas.restoreState()

                        def _later_pages(canvas, doc):
                              canvas.saveState()

                              # ----- bottom-left: confidential -----
                              canvas.setFont("Helvetica", 8)
                              canvas.drawString(
                                    doc.leftMargin,
                                    doc.bottomMargin - 14,
                                    "Confidential | 2025 Labor Pilot"
                              )

                              # ----- bottom-right: page number -----
                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.bottomMargin - 14,
                                    f"Page | {canvas.getPageNumber()}"
                              )

                              canvas.restoreState()

                        # -----------------------------------------------------
                        # ✅ BUILD PDF
                        # -----------------------------------------------------
                        doc.build(elements, onFirstPage=_first_page, onLaterPages=_later_pages)

                        buffer.seek(0)
                        st.download_button(
                              "📥 Download PDF",
                              buffer.getvalue(),
                              file_name=f"Actual_Hours_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.pdf"
                        )

                  elif export_format == "CSV":

                        # -----------------------------------------------------
                        # ✅ SIMPLE CSV EXPORT (DATA TABLE ONLY)
                        # -----------------------------------------------------
                        import io

                        csv_buffer = io.StringIO()
                        export_df.to_csv(csv_buffer, index=False)

                        st.download_button(
                              "📥 Download CSV",
                              data=csv_buffer.getvalue(),
                              file_name=f"Actual_Hours_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.csv",
                              mime="text/csv"
                        )

      if report_type == "Forecast Variance":
            st.markdown("### Forecast Variance Report Filters")

            if "forecast_var_date" not in st.session_state:
                  st.session_state.forecast_var_date = date.today()

            sel_date = st.date_input("Select any date in the week", value=st.session_state.forecast_var_date, key="forecast_var_date")
            week_start = sel_date + relativedelta(weekday=MO(-1))
            week_end   = week_start + timedelta(days=6)

            st.markdown(f"📅 **Selected Week:** {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}")

            generate_forecast_var = st.button("📊 Generate Forecast Variance Report")

            if generate_forecast_var:
                  def pull_week_kpi_totals(Model, label):
                        rows = (
                              session.query(Model)
                                     .filter(Model.date.between(week_start, week_end))
                                     .all()
                        )
                        data = defaultdict(float)
                        for r in rows:
                              data[r.kpi] += r.value
                        return pd.DataFrame([(k, v) for k, v in data.items()], columns=["KPI", label])

                  df_forecast = pull_week_kpi_totals(db.RoomForecast, "Forecast")
                  df_actual   = pull_week_kpi_totals(db.RoomActual, "Actual")
                  df_otb      = pull_week_kpi_totals(db.RoomOTBPickup, "OTB + Pickup")
                  # If no data found for the selected week, show warning and stop
                  if df_forecast.empty and df_actual.empty and df_otb.empty:
                        st.warning("⚠️ No forecast, actual, or OTB + pickup data found for the selected week.")
                        st.stop()

                  merged = df_forecast.merge(df_actual, on="KPI", how="outer")\
                                      .merge(df_otb, on="KPI", how="outer").fillna(0)

                  merged = merged[["KPI", "Actual", "Forecast", "OTB + Pickup"]]
                  merged["Δ Actual - Forecast"] = merged["Actual"] - merged["Forecast"]
                  merged["Δ OTB - Forecast"]    = merged["OTB + Pickup"] - merged["Forecast"]

                  def add_arrow(val):
                        if val > 0:
                              return f"▲ {val}"
                        elif val < 0:
                              return f"▼ {abs(val)}"
                        else:
                              return "0"

                  merged["Δ Actual - Forecast"] = merged["Δ Actual - Forecast"].apply(add_arrow)
                  merged["Δ OTB - Forecast"]    = merged["Δ OTB - Forecast"].apply(add_arrow)

                  export_df = merged.copy()

                  if export_format == "Excel":
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:

                              # Write table starting at row 6 (index 5), column B (index 1)
                              export_df.to_excel(
                                    writer,
                                    sheet_name="Forecast Variance",
                                    index=False,
                                    startrow=5,
                                    startcol=1,   # column B
                                    header=False
                              )

                              workbook  = writer.book
                              worksheet = writer.sheets["Forecast Variance"]

                              # ──────────────────────────────────────────────
                              # SOLID WHITE BACKGROUND
                              # ──────────────────────────────────────────────
                              solid_bg_fmt = workbook.add_format({'bg_color': '#FFFFFF', 'pattern': 1})
                              worksheet.set_column(0, 51, None, solid_bg_fmt)
                              for r in range(0, 2000):
                                    worksheet.set_row(r, None, solid_bg_fmt)

                              worksheet.set_column(0, 0, 2)   # gutter column A

                              # ──────────────────────────────────────────────
                              # FORMATS — Standardized (NO ITALICS, NO DECIMALS)
                              # ──────────────────────────────────────────────
                              title_fmt  = workbook.add_format({'bold': True, 'font_size': 14})

                              borderless_label_fmt = workbook.add_format({
                                    'bold': True,
                                    'border': 0,
                                    'bg_color': '#FFFFFF'
                              })
                              borderless_value_fmt = workbook.add_format({
                                    'border': 0,
                                    'bg_color': '#FFFFFF'
                              })

                              header_fmt = workbook.add_format({
                                    'bold': True,
                                    'bg_color': '#16365C',
                                    'font_color': '#FFFFFF',
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'align': 'center',
                                    'pattern': 1
                              })

                              kpi_fmt = workbook.add_format({
                                    'bold': True,
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'align': 'center',
                                    'num_format': '#,##0'
                              })

                              default_fmt = workbook.add_format({
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'align': 'center',
                                    'num_format': '#,##0'
                              })

                              variance_fmt = workbook.add_format({
                                    'bg_color': '#FEF6F0',
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'align': 'center',
                                    'pattern': 1,
                                    'num_format': '#,##0'
                              })

                              # ──────────────────────────────────────────────
                              # REPORT TITLE + METADATA (BORDERLESS)
                              # ──────────────────────────────────────────────
                              worksheet.write("B1", "Forecast Variance Report", title_fmt)

                              worksheet.write("B3", "Week:", borderless_label_fmt)
                              worksheet.write("C3", f"{week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}", borderless_value_fmt)

                              # ──────────────────────────────────────────────
                              # HEADER ROW AT ROW 6 (INDEX 5)
                              # ──────────────────────────────────────────────
                              header_row = 5
                              header_col = 1
                              for col_idx, col in enumerate(export_df.columns):
                                    worksheet.write(header_row, header_col + col_idx, col, header_fmt)

                              # ──────────────────────────────────────────────
                              # DATA ROWS — START ROW 7 (INDEX 6)
                              # ──────────────────────────────────────────────
                              data_start_row = 6
                              data_start_col = 1

                              for row_idx, row in export_df.iterrows():
                                    for col_idx, col in enumerate(export_df.columns):
                                          val = row[col]

                                          if col == "KPI":
                                                fmt = kpi_fmt
                                          elif "Δ" in col:
                                                fmt = variance_fmt
                                          else:
                                                fmt = default_fmt

                                          worksheet.write(data_start_row + row_idx, data_start_col + col_idx, val, fmt)

                              # ──────────────────────────────────────────────
                              # AUTO-FIT COLUMN WIDTHS
                              # ──────────────────────────────────────────────
                              for i, col in enumerate(export_df.columns):
                                    try:
                                          max_width = max(len(str(col)), export_df[col].astype(str).str.len().max())
                                    except Exception:
                                          max_width = len(str(col))

                                    worksheet.set_column(
                                          1 + i,
                                          1 + i,
                                          max(12, min(30, (max_width or 10) + 2)),
                                          workbook.add_format({'align': 'center'})
                                    )

                              # ──────────────────────────────────────────────
                              # TOP-RIGHT HOTEL NAME + CREATED BY (ROW 1 & 2)
                              # ──────────────────────────────────────────────
                              user_obj = st.session_state.get("user") or {}
                              username = user_obj.get("username") or user_obj.get("email") or "User"
                              hotel_name = user_obj.get("hotel_name") or "Hotel"
                              from datetime import datetime
                              created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                              rightmost_col = 1 + len(export_df.columns) - 1

                              borderless_right_bold = workbook.add_format({
                                    'bold': True,
                                    'align': 'right',
                                    'border': 0,
                                    'bg_color': '#FFFFFF'
                              })

                              borderless_right = workbook.add_format({
                                    'align': 'right',
                                    'border': 0,
                                    'bg_color': '#FFFFFF'
                              })

                              worksheet.write(
                                    0, rightmost_col,
                                    hotel_name,
                                    borderless_right_bold
                              )

                              worksheet.write(
                                    1, rightmost_col,
                                    f"Created by {username} on {created_str}",
                                    borderless_right
                              )

                              # ──────────────────────────────────────────────
                              # CONFIDENTIAL FOOTER UNDER TABLE
                              # ──────────────────────────────────────────────
                              bottom_row = data_start_row + len(export_df) + 2
                              worksheet.write(
                                    bottom_row, 1,
                                    "Confidential | 2025 Labor Pilot",
                                    workbook.add_format({'align': 'left'})
                              )

                        st.download_button(
                              "⬇️ Download Excel",
                              data=output.getvalue(),
                              file_name=f"Forecast_Variance_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                  elif export_format == "PDF":
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib.pagesizes import landscape, letter
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors
                        from datetime import datetime
                        import io

                        buffer = io.BytesIO()
                        doc = SimpleDocTemplate(
                              buffer,
                              pagesize=landscape(letter),
                              leftMargin=28,
                              rightMargin=28,
                              topMargin=24,
                              bottomMargin=24
                        )

                        styles = getSampleStyleSheet()
                        elements = []

                        # ---------- TITLE (match Labor Variance styling) ----------
                        title_style = styles["Heading2"]   # correct font + spacing
                        elements.append(Paragraph("<b>Forecast Variance Report</b>", title_style))
                        elements.append(Spacer(1, 8))

                        # ---------- METADATA ----------
                        elements.append(Paragraph(f"<b>Week:</b> {week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}", styles["Normal"]))
                        elements.append(Spacer(1, 12))

                        # ---------- TABLE CONTENT ----------
                        pdf_data = [export_df.columns.tolist()]

                        for _, row in export_df.iterrows():
                              formatted_row = []
                              for col in export_df.columns:
                                    val = row[col]

                                    # Format numbers → comma, NO decimals
                                    if col != "KPI" and "Δ" not in col:
                                          try:
                                                formatted_row.append(f"{int(val):,}")
                                          except:
                                                formatted_row.append(val)
                                    elif "Δ" in col:
                                          try:
                                                formatted_row.append(f"{int(val):,}")
                                          except:
                                                formatted_row.append(val)
                                    else:
                                          formatted_row.append(val)

                              pdf_data.append(formatted_row)

                        table = Table(pdf_data)

                        # ---------- TABLE STYLE (match Labor Variance) ----------
                        style = TableStyle([
                              ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("FONTSIZE", (0, 0), (-1, -1), 9),
                              ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                              ("TOPPADDING", (0, 0), (-1, -1), 4),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ])

                        # ---------- ROW FORMATTING ----------
                        for r in range(1, len(pdf_data)):
                              # KPI column normal font (no italics)
                              style.add("FONTNAME", (0, r), (0, r), "Helvetica")

                              # Variance columns shading
                              for c, name in enumerate(export_df.columns):
                                    if "Δ" in name:
                                          style.add("BACKGROUND", (c, r), (c, r), colors.HexColor("#FEF6F0"))

                        table.setStyle(style)
                        elements.append(table)

                        # ---------- TOP-RIGHT HOTEL + CREATED BY ONLY ON PAGE 1 ----------
                        user_obj = st.session_state.get("user") or {}
                        username = (
                              user_obj.get("username")
                              or user_obj.get("email")
                              or "User"
                        )
                        hotel_name = user_obj.get("hotel_name") or "Hotel"
                        created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                        def _first_page(canvas, doc):
                              canvas.saveState()

                              # Top-right: hotel
                              canvas.setFont("Helvetica-Bold", 10)
                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.pagesize[1] - doc.topMargin - 10,
                                    hotel_name
                              )

                              # Created by
                              canvas.setFont("Helvetica", 8)
                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.pagesize[1] - doc.topMargin - 22,
                                    f"Created by {username} on {created_str}"
                              )

                              # Footer
                              canvas.setFont("Helvetica", 8)
                              canvas.drawString(
                                    doc.leftMargin,
                                    doc.bottomMargin - 14,
                                    "Confidential | 2025 Labor Pilot"
                              )

                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.bottomMargin - 14,
                                    f"Page | {canvas.getPageNumber()}"
                              )

                              canvas.restoreState()

                        def _later_pages(canvas, doc):
                              canvas.saveState()

                              canvas.setFont("Helvetica", 8)
                              canvas.drawString(
                                    doc.leftMargin,
                                    doc.bottomMargin - 14,
                                    "Confidential | 2025 Labor Pilot"
                              )

                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.bottomMargin - 14,
                                    f"Page | {canvas.getPageNumber()}"
                              )

                              canvas.restoreState()

                        # ---------- BUILD PDF ----------
                        doc.build(elements, onFirstPage=_first_page, onLaterPages=_later_pages)

                        buffer.seek(0)
                        st.download_button(
                              "📥 Download PDF",
                              data=buffer.getvalue(),
                              file_name=f"Forecast_Variance_Report_{week_start:%Y%m%d}.pdf",
                              mime="application/pdf"
                        )

                  elif export_format == "CSV":
                        csv_data = export_df.to_csv(index=False).encode("utf-8")
                        st.download_button(
                              "⬇️ Download CSV",
                              data=csv_data,
                              file_name=f"Forecast_Variance_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.csv",
                              mime="text/csv"
                        )
      if report_type == "Department Schedule":
            st.markdown("### Department Schedule Report Filters")

            import os
            from datetime import date, timedelta
            from dateutil.relativedelta import relativedelta, MO

            fmt_day = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"

            if "dept_sched_date" not in st.session_state:
                  st.session_state.dept_sched_date = date.today()

            sel_date = st.date_input("Select any date in the week",
                                     value=st.session_state.dept_sched_date,
                                     key="dept_sched_date")
            week_start = sel_date + relativedelta(weekday=MO(-1))
            week_end   = week_start + timedelta(days=6)
            week_dates = [week_start + timedelta(i) for i in range(7)]
            day_cols   = [d.strftime(fmt_day) for d in week_dates]

            st.session_state["dept_schedule_week_start"] = week_start
            st.session_state["dept_schedule_week_end"]   = week_end
            st.session_state["dept_schedule_day_cols"]   = day_cols

            st.markdown(f"📅 **Selected Week:** {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}")

            # ---------- Optional KPI totals header (independent KPI picker) ----------
            with st.expander("Optional: Selected KPI Totals header", expanded=False):
                  show_kpi_header   = st.checkbox("Show KPI header above schedule", value=False, key="dept_show_kpi_header")

                  # 🔒 Always Forecast – no selection
                  st.markdown("**KPI data source:** Forecast")
                  kpi_source = "Forecast"

                  kpi_limit_to_week = st.checkbox("List KPIs from selected week only", value=True, key="dept_kpi_week_only")
                  kpi_choices_placeholder = st.empty()
            # ------------------------------------------------------------------------

            # ---------- 12-hour display helpers (storage/math stay 24h) ----------
            from datetime import datetime as _dt

            def _to12(hhmm: str) -> str:
                  try:
                        h, m = map(int, str(hhmm).split(":"))
                        t = _dt(2000, 1, 1, h, m)
                        return t.strftime("%I:%M %p").lstrip("0")
                  except Exception:
                        return str(hhmm)

            def _to12_range(rng: str) -> str:
                  try:
                        a, b = str(rng).split("-")
                        return f"{_to12(a)} - {_to12(b)}"
                  except Exception:
                        return str(rng)

            def make_display_copy_12h(df, day_cols):
                  d = df.copy()
                  for c in day_cols:
                        if c in d.columns:
                              d[c] = d[c].apply(
                                    lambda x: _to12_range(x) if isinstance(x, str) and "-" in x and ":" in x else x
                              )
                  return d
            # --------------------------------------------------------------------

            # ---------- hours helper (subtract 0.5 hr per shift; supports overnight) ----------
            def _hours_with_break(cell: str) -> float:
                  try:
                        if not cell: return 0.0
                        s = str(cell).strip()
                        if not s or s.upper() == "OFF": return 0.0
                        if "-" not in s or ":" not in s: return 0.0
                        start, end = [t.strip() for t in s.split("-", 1)]
                        h0, m0 = map(int, start.split(":"))
                        h1, m1 = map(int, end.split(":"))
                        start_min = h0*60 + m0
                        end_min   = h1*60 + m1
                        diff_min  = end_min - start_min
                        if diff_min <= 0: diff_min += 1440  # cross-midnight
                        hrs = diff_min/60.0
                        hrs = max(0.0, hrs - 0.5)          # 30-min break
                        return round(hrs, 2)
                  except Exception:
                        return 0.0
            # -----------------------------------------------------------------------------------

            emp_df = refresh(db.Employee)
            emp_df = emp_df.loc[
                  emp_df["emp_type"].fillna("").str.strip().str.casefold() != "terminated"
            ].copy()

            # 🔒 APPLY MANAGER SCOPE FIRST
            emp_df = apply_manager_scope(
                  emp_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})

            dept_list = sorted(
                  emp_df["department"]
                  .dropna()
                  .unique()
                  .tolist()
            )

            if not dept_list:
                  st.warning("⚠️ You do not have access to any departments.")
                  st.stop()

            sel_dept  = st.selectbox(
                  "Department*",
                  dept_list,
                  key="sched_dept_filter"
            )

            pos_opts = (
                  emp_df.loc[
                        emp_df["department"] == sel_dept,
                        "role"
                  ]
                  .dropna()
                  .unique()
                  .tolist()
            )

            if len(pos_opts) == 0:
                  st.warning("⚠️ No positions found under the selected department.")
                  st.stop()

            sel_pos = st.selectbox(
                  "Position*",
                  sorted(pos_opts),
                  key="sched_pos_filter"
            )

            # Populate KPI multi-select from the chosen source (independent of position)
            if show_kpi_header:
                  # always Forecast now
                  src_tbl = refresh(db.RoomForecast)

                  df_src = src_tbl.copy()
                  if kpi_limit_to_week:
                        df_src = df_src[df_src["date"].between(week_start, week_end)]
                  kpi_options = sorted(df_src["kpi"].dropna().unique().tolist())
                  if not kpi_options:
                        kpi_options = sorted(src_tbl["kpi"].dropna().unique().tolist())

                  selected_kpis = kpi_choices_placeholder.multiselect(
                        "Choose KPIs to show",
                        options=kpi_options,
                        default=[],                  # ✅ no preselected KPIs
                        key="dept_selected_kpis"
                  )
            else:
                  selected_kpis = []

            generate_schedule = st.button("📊 Generate Schedule Report")

            # =============== KPI Totals header (computed DF kept for exports) ===============
            kpi_proj_df = None   # columns: ["KPI"] + day_cols + ["Total"]
            if generate_schedule and show_kpi_header and selected_kpis:
                  # choose the source table (Forecast only)
                  kpi_tbl = refresh(db.RoomForecast)

                  # restrict to the selected week and selected KPIs
                  kdf = kpi_tbl[kpi_tbl["date"].between(week_start, week_end)].copy()
                  kdf = kdf[kdf["kpi"].isin(selected_kpis)]

                  # build a totals table: rows=KPI, cols=Mon..Sun + Total
                  rows = []
                  for metric in selected_kpis:
                        row = {"KPI": metric}
                        total = 0.0
                        for d, col in zip(week_dates, day_cols):
                              day_total = float(kdf[(kdf["kpi"] == metric) & (kdf["date"] == d)]["value"].sum() or 0.0)
                              row[col] = round(day_total, 2)
                              total += day_total
                        row["Total"] = round(total, 2)
                        rows.append(row)

                  if rows:
                        import pandas as pd
                        kpi_proj_df = pd.DataFrame(rows, columns=["KPI"] + day_cols + ["Total"])
                        # TOTAL row (sum across selected KPIs)
                        total_row = {"KPI": "TOTAL"}
                        for col in day_cols:
                              total_row[col] = round(float(pd.to_numeric(kpi_proj_df[col], errors="coerce").fillna(0).sum()), 2)
                        total_row["Total"] = round(float(pd.to_numeric(kpi_proj_df["Total"], errors="coerce").fillna(0).sum()), 2)
                        kpi_proj_df = pd.concat([kpi_proj_df, pd.DataFrame([total_row])], ignore_index=True)
            # ================================================================================

            if generate_schedule:
                  if not sel_dept or not sel_pos:
                        st.warning("Please select both a department and a position.")
                        st.stop()

                  emp_sub = emp_df[(emp_df["department"] == sel_dept) & (emp_df["role"] == sel_pos)]
                  if emp_sub.empty:
                        st.warning("⚠️ No matching employees for selected filters.")
                        st.stop()

                  ids     = emp_sub["name"].str.extract(r"(\d+)$")[0].fillna("")
                  firsts  = emp_sub["name"].str.extract(r",\s*([^\d]+)")[0].str.strip()
                  lasts   = emp_sub["name"].str.extract(r"^\s*([^,]+)")[0].str.strip()

                  sched_df = pd.DataFrame({
                        "ID": ids,
                        "First Name": firsts,
                        "Last Name": lasts,
                        "emp_id": emp_sub["id"]
                  })
                  for dc in day_cols:
                        sched_df[dc] = ""

                  sched_rows = session.query(db.Schedule).filter(
                        db.Schedule.emp_id.in_(emp_sub["id"]),
                        db.Schedule.day.in_(week_dates)
                  ).all()
                  if not sched_rows:
                        st.warning("⚠️ No schedule data found for this position and week.")
                        st.stop()

                  for r in sched_rows:
                        col = r.day.strftime(fmt_day)
                        idx = sched_df.index[sched_df["emp_id"] == r.emp_id]
                        if not idx.empty:
                              sched_df.at[idx[0], col] = r.shift_type  # keep raw HH:MM-HH:MM or OFF

                  sched_df.drop(columns=["emp_id"], inplace=True)
                  st.session_state["dept_schedule_df"] = sched_df

            if "dept_schedule_df" in st.session_state and generate_schedule:
                  import pandas as pd
                  df         = st.session_state["dept_schedule_df"]
                  week_start = st.session_state["dept_schedule_week_start"]
                  week_end   = st.session_state["dept_schedule_week_end"]

                  # ===== build raw + display once =====
                  df_raw = df.copy()
                  df_raw["Total"] = df_raw.iloc[:, 3:].apply(
                        lambda row: round(sum(_hours_with_break(x) for x in row), 2), axis=1
                  )
                  df_disp      = make_display_copy_12h(df_raw, day_cols)
                  df_to_write  = df_disp.copy()
                  df_to_write["Total"] = df_raw["Total"]

                  # per-day sums (with 0.5h break) + weekly total sum for the TOTAL row
                  day_sums = {c: round(sum(_hours_with_break(x) for x in df_raw[c]), 2) for c in day_cols}
                  weekly_total_sum = round(float(pd.to_numeric(df_raw["Total"], errors="coerce").fillna(0).sum()), 2)


                  # ========================= EXPORTS ==============================
                  if export_format == "Excel":
                        import io
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                        from openpyxl.utils import get_column_letter

                        buffer = io.BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = f"Schedule – {sel_pos}"

                        # styles
                        bold_font     = Font(bold=True)
                        left_align    = Alignment(horizontal="left", vertical="center")
                        center_align  = Alignment(horizontal="center", vertical="center")
                        header_fill   = PatternFill("solid", fgColor="16365C")
                        kpi_header_fill = PatternFill("solid", fgColor="444444")
                        header_font   = Font(color="FFFFFF", bold=True)
                        border_side   = Side(style="thin",  color="A6A6A6")
                        border        = Border(left=border_side, right=border_side,
                                               top=border_side,  bottom=border_side)
                        off_fill      = PatternFill("solid", fgColor="FEBFBA")
                        total_fill    = PatternFill("solid", fgColor="FEF6F0")
                        blue_fill     = PatternFill("solid", fgColor="E6EEF8")
                        red_font      = Font(color="FF0000")
                        bold_italic   = Font(bold=True, italic=False)
                        solid_bg_fill = PatternFill("solid", fgColor="FFFFFF")

                        # layout
                        ws.column_dimensions["A"].width = 2
                        start_col_idx = 2  # column B
                        base_start_row = 7

                        # ========== METADATA (LEFT ALIGNED, NO BORDERS) ==========
                        ws.merge_cells(start_row=1, start_column=start_col_idx, end_row=1, end_column=start_col_idx+3)
                        title_cell = ws.cell(row=1, column=start_col_idx, value="Department Schedule Report")
                        title_cell.font = Font(bold=True, size=14)
                        title_cell.alignment = left_align

                        # Department
                        ws.cell(row=3, column=start_col_idx,     value="Department:").font = bold_font
                        ws.cell(row=3, column=start_col_idx,     ).alignment = left_align
                        ws.cell(row=3, column=start_col_idx + 1, value=sel_dept).alignment = left_align

                        # Position
                        ws.cell(row=4, column=start_col_idx,     value="Position:").font = bold_font
                        ws.cell(row=4, column=start_col_idx,     ).alignment = left_align
                        ws.cell(row=4, column=start_col_idx + 1, value=sel_pos).alignment = left_align

                        # Week
                        ws.cell(row=5, column=start_col_idx,     value="Week:").font = bold_font
                        ws.cell(row=5, column=start_col_idx,     ).alignment = left_align
                        ws.cell(row=5, column=start_col_idx + 1, value=f"{week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}").alignment = left_align

                        # ========== REMOVE borders + background from metadata ==========
                        for r in (3, 4, 5):
                              for c in (start_col_idx, start_col_idx + 1):
                                    cell = ws.cell(row=r, column=c)
                                    cell.border = Border()  # no border
                                    cell.fill = solid_bg_fill
                                    cell.alignment = left_align

                        # -------------------------------------------------------------------
                        # WIDTH HELPER
                        def _col_widths_from(df_like):
                              cols = list(df_like.columns)
                              vals = df_like.astype(str).values.tolist()
                              widths = []
                              for i, c in enumerate(cols):
                                    series = [c] + [str(row[i]) for row in vals]
                                    widths.append(max(12, min(60, int(max(len(s) for s in series) * 1.1) + 2)))
                              return widths

                        sched_widths = _col_widths_from(df_to_write)

                        # -------------------------------------------------------------------
                        # KPI TOTALS HEADER (unchanged)
                        schedule_start_row = base_start_row
                        if kpi_proj_df is not None and not kpi_proj_df.empty:
                              ws.merge_cells(start_row=base_start_row-1, start_column=start_col_idx,
                                             end_row=base_start_row-1, end_column=start_col_idx+len(df_to_write.columns)-1)
                              ws.cell(row=base_start_row-1, column=start_col_idx,
                                      value=f"Selected KPI Totals — {kpi_source}").font = Font(bold=True, size=12)

                              def _write_merge(r, c1, c2, val, fill=None, font=None, align=center_align, brd=border):
                                    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
                                    for cc in range(c1, c2+1):
                                          cell = ws.cell(row=r, column=cc, value=val if cc == c1 else None)
                                          cell.alignment = align
                                          cell.border = brd
                                          if fill: cell.fill = fill
                                          if font: cell.font = font

                              header_r = base_start_row
                              _write_merge(header_r, start_col_idx, start_col_idx+2, "KPI",
                                           fill=kpi_header_fill, font=header_font)

                              for i_day, col in enumerate(day_cols):
                                    c = ws.cell(row=header_r, column=start_col_idx+3+i_day, value=col)
                                    c.font = header_font
                                    c.fill = kpi_header_fill
                                    c.border = border
                                    c.alignment = center_align

                              c = ws.cell(row=header_r, column=start_col_idx+3+len(day_cols), value="Total")
                              c.font = header_font
                              c.fill = kpi_header_fill
                              c.border = border
                              c.alignment = center_align

                              from openpyxl.utils import get_column_letter as _gcl
                              ws.column_dimensions[_gcl(start_col_idx)].width = max(26, sched_widths[0])
                              ws.column_dimensions[_gcl(start_col_idx + 1)].width = max(22, sched_widths[1])
                              ws.column_dimensions[_gcl(start_col_idx + 2)].width = sched_widths[2]

                              for i_day in range(len(day_cols)):
                                    ws.column_dimensions[_gcl(start_col_idx + 3 + i_day)].width = sched_widths[3 + i_day]

                              ws.column_dimensions[_gcl(start_col_idx + 3 + len(day_cols))].width = sched_widths[3 + len(day_cols)]

                              for ridx, row in enumerate(kpi_proj_df.values.tolist(), start=header_r + 1):
                                    is_total = (str(row[0]).strip().upper() == "TOTAL")
                                    row_font = bold_italic if is_total else None
                                    row_fill = blue_fill if is_total else None

                                    _write_merge(ridx, start_col_idx, start_col_idx+2, row[0],
                                                 fill=row_fill, font=row_font)

                                    for i_day, val in enumerate(row[1:1+len(day_cols)]):
                                          c = ws.cell(row=ridx, column=start_col_idx+3+i_day, value=val)
                                          c.border = border
                                          c.alignment = center_align
                                          if row_fill: c.fill = row_fill
                                          if row_font: c.font = row_font

                                    c = ws.cell(row=ridx, column=start_col_idx+3+len(day_cols), value=row[-1])
                                    c.border = border
                                    c.alignment = center_align
                                    if row_fill: c.fill = row_fill
                                    if row_font: c.font = row_font

                              schedule_start_row = header_r + len(kpi_proj_df) + 2

                        # -------------------------------------------------------------------
                        # SCHEDULE TABLE HEADER
                        for i, col in enumerate(df_to_write.columns):
                              c = ws.cell(row=schedule_start_row, column=start_col_idx + i, value=col)
                              c.font = header_font
                              c.fill = header_fill
                              c.border = border
                              c.alignment = center_align

                        # -------------------------------------------------------------------
                        # SCHEDULE DATA ROWS
                        for r_idx, row in enumerate(df_to_write.values.tolist(), start=schedule_start_row + 1):
                              for c_off, val in enumerate(row):
                                    c = ws.cell(row=r_idx, column=start_col_idx + c_off, value=val)
                                    c.alignment = center_align
                                    c.border = border

                                    if c_off <= 2:
                                          c.font = bold_italic

                                    if isinstance(val, str) and val.strip().upper() == "OFF":
                                          c.fill = off_fill

                                    if df_to_write.columns[c_off] == "Total":
                                          c.fill = total_fill
                                          c.number_format = "0.00"
                                          try:
                                                if float(val) > 40:
                                                      c.font = red_font
                                          except:
                                                pass

                        # -------------------------------------------------------------------
                        # TOTAL ROW
                        total_row_idx = schedule_start_row + 1 + len(df_to_write)
                        c = ws.cell(row=total_row_idx, column=start_col_idx, value="Total")
                        c.font = bold_italic
                        c.border = border
                        c.fill = blue_fill
                        c.alignment = center_align

                        for c_off in (1, 2):
                              cc = ws.cell(row=total_row_idx, column=start_col_idx + c_off, value="")
                              cc.border = border
                              cc.fill = blue_fill
                              cc.alignment = center_align

                        # -------------------------------------------------------------------
                        # TOP-RIGHT HOTEL NAME + CREATED BY (NO BORDERS)
                        user_obj = st.session_state.get("user") or {}
                        username = user_obj.get("username") or user_obj.get("email") or "User"
                        hotel_name = user_obj.get("hotel_name") or "Hotel"
                        from datetime import datetime
                        created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                        rightmost_col = start_col_idx + len(df_to_write.columns) - 1

                        cell_hotel = ws.cell(row=1, column=rightmost_col, value=str(hotel_name))
                        cell_hotel.font = Font(bold=True)
                        cell_hotel.alignment = Alignment(horizontal="right")
                        cell_hotel.border = Border()
                        cell_hotel.fill = solid_bg_fill

                        cell_created = ws.cell(row=2, column=rightmost_col,
                                               value=f"Created by {username} on {created_str}")
                        cell_created.alignment = Alignment(horizontal="right")
                        cell_created.border = Border()
                        cell_created.fill = solid_bg_fill

                        # ---------- BOTTOM-LEFT CONFIDENTIAL ----------
                        bottom_confidential_row = total_row_idx + 1
                        ws.cell(row=bottom_confidential_row, column=start_col_idx, value="Confidential | © 2025 Labor Pilot")
                        for i, col_name in enumerate(day_cols, start=3):
                              v = day_sums[col_name]
                              cell = ws.cell(row=total_row_idx, column=start_col_idx + i)
                              if v == 0:
                                    cell.value = "-"
                                    cell.font = bold_italic
                              else:
                                    cell.value = float(v)
                                    cell.number_format = "0.00"
                                    cell.font = bold_italic
                              cell.alignment = center_align; cell.border = border; cell.fill = blue_fill
                        last_col_index = start_col_idx + len(df_to_write.columns) - 1
                        cc = ws.cell(row=total_row_idx, column=last_col_index, value=float(weekly_total_sum))
                        cc.number_format = "0.00"; cc.font = bold_italic; cc.alignment = center_align; cc.border = border; cc.fill = blue_fill

                        # background fill
                        max_rows = max(total_row_idx + 5, 200)
                        max_cols = max(start_col_idx + len(df_to_write.columns) + 5, 52)
                        for r in range(1, max_rows + 1):
                              for k in range(1, max_cols + 1):
                                    cell = ws.cell(row=r, column=k)
                                    if cell.fill is None or cell.fill.fill_type is None:
                                          cell.fill = solid_bg_fill

                        # schedule column widths (apply after writing so both tables use the same)
                        from openpyxl.utils import get_column_letter
                        for c_off, w in enumerate(sched_widths):
                              letter  = get_column_letter(start_col_idx + c_off)
                              if c_off == 0: w = max(w, 26)  # ID min
                              if c_off == 1: w = max(w, 22)  # First min
                              ws.column_dimensions[letter].width = w

                        st.download_button(
                              "📥 Download Excel",
                              data=(lambda wb=wb: (wb.save(buffer), buffer.seek(0), buffer.getvalue()))()[-1],
                              file_name=f"Dept_Schedule_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.xlsx"
                        )

                  elif export_format == "PDF":
                        import io, re
                        from datetime import datetime
                        from reportlab.lib.pagesizes import landscape, letter
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors

                        buffer = io.BytesIO()
                        # keep margins modest; width calculation will use doc.width
                        doc = SimpleDocTemplate(
                              buffer,
                              pagesize=landscape(letter),
                              leftMargin=28, rightMargin=28, topMargin=24, bottomMargin=24
                        )
                        styles = getSampleStyleSheet()
                        elements = []

                        # ----- header meta text (dates numeric) -----
                        elements.append(Paragraph("Department Schedule Report", styles["Heading2"]))
                        elements.append(Spacer(1, 6))
                        elements.append(Paragraph(f"<b>Department:</b> {sel_dept}", styles["Normal"]))
                        elements.append(Paragraph(f"<b>Position:</b> {sel_pos}", styles["Normal"]))
                        elements.append(Paragraph(f"<b>Week:</b> {week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}", styles["Normal"]))
                        elements.append(Spacer(1, 10))

                        # ---------- width helpers ----------
                        def _widths_chars(df_like):
                              cols = list(df_like.columns)
                              vals = df_like.astype(str).values.tolist()
                              widths = []
                              for i, c in enumerate(cols):
                                    series = [c] + [str(row[i]) for row in vals]
                                    widths.append(max(12, min(60, int(max(len(s) for s in series) * 1.1) + 2)))
                              return widths

                        def _fit_to_available(widths_pts, available_pts):
                              total = sum(widths_pts)
                              if total <= available_pts or total == 0:
                                    return widths_pts
                              scale = available_pts / total
                              return [w * scale for w in widths_pts]
                        # ------------------------------------

                        # 1) Compute schedule widths in points and FIT them to page width
                        sched_widths_chars = _widths_chars(df_to_write)
                        PT_PER_CHAR = 6.5
                        sched_widths_pts = [w * PT_PER_CHAR for w in sched_widths_chars]
                        sched_widths_pts = _fit_to_available(sched_widths_pts, doc.width)

                        # 2) KPI header (aligned & also fitted)
                        if kpi_proj_df is not None and not kpi_proj_df.empty:
                              elements.append(Paragraph(f"Selected KPI Totals — {kpi_source}", styles["Heading4"]))
                              elements.append(Spacer(1, 4))

                              # Build KPI widths directly from the *scaled* schedule widths so they align perfectly.
                              kpi_colwidths_pts = []
                              kpi_first_pts = sum(sched_widths_pts[0:3])  # ID + First + Last
                              kpi_colwidths_pts.append(kpi_first_pts)
                              for i_day, _ in enumerate(day_cols):
                                    kpi_colwidths_pts.append(sched_widths_pts[3 + i_day])
                              kpi_colwidths_pts.append(sched_widths_pts[3 + len(day_cols)])  # Total
                              # Safety: if rounding pushed us over the page width, fit again.
                              kpi_colwidths_pts = _fit_to_available(kpi_colwidths_pts, doc.width)

                              data_kpi = [kpi_proj_df.columns.tolist()] + kpi_proj_df.astype(str).values.tolist()
                              tbl_kpi = Table(
                                    data_kpi,
                                    repeatRows=1,
                                    colWidths=kpi_colwidths_pts,
                                    hAlign="LEFT",
                              )
                              stl = TableStyle([
                                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#444444")),   # KPI header color
                                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                              ])
                              total_r = len(data_kpi) - 1
                              stl.add("BACKGROUND", (0, total_r), (-1, total_r), colors.HexColor("#E6EEF8"))
                              # total row bold (no italics)
                              stl.add("FONTNAME", (0, total_r), (-1, total_r), "Helvetica-Bold")
                              tbl_kpi.setStyle(stl)
                              elements.append(tbl_kpi)
                              elements.append(Spacer(1, 10))

                        # 3) Schedule table using the same scaled widths
                        df_disp2 = df_to_write.copy()

                        # append "Total" row to PDF table
                        total_row = {"ID": "Total", "First Name": "", "Last Name": ""}
                        for c in day_cols:
                              v = day_sums[c]
                              total_row[c] = "-" if v == 0 else f"{v:.2f}"
                        total_row["Total"] = f"{weekly_total_sum:.2f}"
                        df_pdf = pd.concat([df_disp2, pd.DataFrame([total_row])], ignore_index=True)

                        def _squash_time_str(x: str):
                              if not isinstance(x, str): return x
                              s = x.strip()
                              if not s or s.upper()=="OFF": return s
                              s = re.sub(r"\s*-\s*", "-", s)
                              s = re.sub(r"\s*(AM|PM)\b", r"\1", s, flags=re.I)
                              return s
                        for col in day_cols:
                              if col in df_pdf.columns:
                                    df_pdf[col] = df_pdf[col].map(_squash_time_str)

                        data = [df_pdf.columns.tolist()] + df_pdf.values.tolist()
                        table = Table(
                              data,
                              repeatRows=1,
                              colWidths=sched_widths_pts,
                              hAlign="LEFT",
                        )
                        style = TableStyle([
                              ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                              ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("FONTSIZE", (0, 0), (-1, -1), 9),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                              ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ])
                        n_rows = len(data) - 1
                        # first 3 columns bold (no italics)
                        for r in range(1, n_rows+1):
                              for c_idx in range(0, 3):
                                    style.add("FONTNAME", (c_idx, r), (c_idx, r), "Helvetica-Bold")
                        # total row background + bold (no italics)
                        style.add("BACKGROUND", (0, n_rows), (-1, n_rows), colors.HexColor("#E6EEF8"))
                        for c_idx in range(0, len(df_pdf.columns)):
                              style.add("FONTNAME", (c_idx, n_rows), (c_idx, n_rows), "Helvetica-Bold")
                        total_col_idx = df_pdf.columns.get_loc("Total")
                        for r in range(1, n_rows):
                              row_vals = data[r]
                              for c_idx, val in enumerate(row_vals):
                                    if isinstance(val, str) and val.strip().upper()=="OFF":
                                          style.add("BACKGROUND", (c_idx, r), (c_idx, r), colors.HexColor("#FEBFBA"))
                                    if c_idx == total_col_idx:
                                          style.add("BACKGROUND", (c_idx, r), (c_idx, r), colors.HexColor("#FEF6F0"))
                        table.setStyle(style)
                        elements.append(table)

                        # ---------- header/footer + first-page-only top-right block ----------
                        user_obj = st.session_state.get("user") or {}
                        username = user_obj.get("username") or user_obj.get("email") or st.session_state.get("username") or "User"
                        hotel_name = user_obj.get("hotel_name") or st.session_state.get("hotel_name") or "Hotel"
                        created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                        def _first_page(canvas, doc):
                              canvas.saveState()

                              # ----- top-right: hotel + created by (ONLY FIRST PAGE) -----
                              # Position it similar to Excel export layout
                              x_right = doc.pagesize[0] - doc.rightMargin
                              y_top   = doc.pagesize[1] - doc.topMargin - 10

                              canvas.setFont("Helvetica-Bold", 10)
                              canvas.drawRightString(x_right, y_top, str(hotel_name))

                              canvas.setFont("Helvetica", 8)
                              canvas.drawRightString(x_right, y_top - 12,
                                    f"Created by {username} on {created_str}"
                              )

                              # ----- bottom-left: confidential -----
                              canvas.setFont("Helvetica", 8)
                              canvas.drawString(
                                    doc.leftMargin,
                                    doc.bottomMargin - 14,
                                    "Confidential | 2025 Labor Pilot"
                              )

                              # ----- bottom-right: page number -----
                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.bottomMargin - 14,
                                    f"Page | {canvas.getPageNumber()}"
                              )

                              canvas.restoreState()

                        def _later_pages(canvas, doc):
                              canvas.saveState()

                              # ----- NO TOP-RIGHT on later pages -----

                              # ----- bottom-left: confidential -----
                              canvas.setFont("Helvetica", 8)
                              canvas.drawString(
                                    doc.leftMargin,
                                    doc.bottomMargin - 14,
                                    "Confidential | 2025 Labor Pilot"
                              )

                              # ----- bottom-right: page number -----
                              canvas.setFont("Helvetica", 8)
                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.bottomMargin - 14,
                                    f"Page | {canvas.getPageNumber()}"
                              )

                              canvas.restoreState()

                        # Build PDF (different functions for first vs later pages)
                        doc.build(elements, onFirstPage=_first_page, onLaterPages=_later_pages)

                        buffer.seek(0)
                        st.download_button(
                              "📥 Download PDF",
                              buffer.getvalue(),
                              file_name=f"Dept_Schedule_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.pdf"
                        )
                  elif export_format == "CSV":
                        df_out = df_to_write.copy()
                        total_row = {"ID": "Total", "First Name": "", "Last Name": ""}
                        for c in day_cols:
                              v = day_sums[c]
                              total_row[c] = "" if v == 0 else f"{v:.2f}"
                        total_row["Total"] = f"{weekly_total_sum:.2f}"
                        df_out = pd.concat([df_out, pd.DataFrame([total_row])], ignore_index=True)

                        if kpi_proj_df is not None and not kpi_proj_df.empty:
                              kpi_csv   = kpi_proj_df.to_csv(index=False)
                              sched_csv = df_out.to_csv(index=False)
                              combined  = "Selected KPI Totals\n" + kpi_csv + "\n\nSchedule\n" + sched_csv
                              st.download_button(
                                    "📥 Download CSV",
                                    combined,
                                    file_name=f"Dept_Schedule_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.csv",
                                    mime="text/csv"
                              )
                        else:
                              st.download_button(
                                    "📥 Download CSV",
                                    df_out.to_csv(index=False),
                                    file_name=f"Dept_Schedule_Report_{week_start:%Y-%m-%d}_to_{week_end:%Y-%m-%d}.csv"
                              )
      elif report_type == "Productivity Index":
            st.markdown("### Productivity Index Report Filters")

            if "prod_index_start" not in st.session_state:
                  st.session_state.prod_index_start = date.today()
            if "prod_index_end" not in st.session_state:
                  st.session_state.prod_index_end = date.today()

            col1, col2 = st.columns(2)
            with col1:
                  sel_start = st.date_input("From Date", value=st.session_state.prod_index_start, key="prod_index_start")
            with col2:
                  sel_end = st.date_input("To Date", value=st.session_state.prod_index_end, key="prod_index_end")
            week_start = sel_start
            week_end   = sel_end

            dept_df = refresh(db.Department).rename(columns={"id": "dept_id", "name": "dept"})
            pos_df  = refresh(db.Position).rename(columns={"id": "position_id", "name": "position"})

            # 🔒 APPLY MANAGER SCOPE USING EMPLOYEES (SOURCE OF TRUTH)
            emp_df = refresh(db.Employee)
            emp_df = apply_manager_scope(
                  emp_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})

            # ───────── Department (SCOPED) ─────────
            dept_options = sorted(
                  emp_df["department"]
                  .dropna()
                  .unique()
                  .tolist()
            )

            sel_dept = st.selectbox(
                  "Select Department",
                  dept_options,
                  key="prod_dept"
            )

            # ───────── Position (SCOPED + DEPT FILTERED) ─────────
            allowed_positions = (
                  emp_df.loc[
                        emp_df["department"] == sel_dept,
                        "role"
                  ]
                  .dropna()
                  .unique()
                  .tolist()
            )

            sel_pos = st.selectbox(
                  "Select Position",
                  ["All Positions"] + sorted(allowed_positions),
                  key="prod_pos"
            )

            generate_btn = st.button("📊 Generate Productivity Report")

            if generate_btn:
                  # ---------- Standards filtered by dept/position ----------
                  std_df = refresh(db.LaborStandard)
                  std_df = std_df.merge(pos_df[["position_id", "position", "department_id"]], on="position_id", how="left")
                  std_df = std_df.merge(dept_df[["dept_id", "dept"]], left_on="department_id", right_on="dept_id", how="left")
                  std_df = std_df[std_df["dept"] == sel_dept]
                  if sel_pos != "All Positions":
                        std_df = std_df[std_df["position"] == sel_pos]

                  # ---------- Actual hours (manual + contract; Reg + OT) ----------
                  ah_df = refresh(db.Actual)
                  ah_df = ah_df[
                        (ah_df["source"].isin(["manual", "contract"])) &
                        (ah_df["date"].between(week_start, week_end))
                  ]
                  ah_df["total_hours"] = ah_df[["hours", "ot_hours"]].sum(axis=1)
                  hours_summary = (
                        ah_df.groupby("position_id")["total_hours"]
                        .sum()
                        .reset_index()
                        .rename(columns={"total_hours": "actual_hours"})
                  )
                  hours_summary = hours_summary.merge(pos_df[["position_id", "position"]], on="position_id", how="left")

                  # ---------- KPI outputs (RoomActual) ----------
                  kpi_df = refresh(db.RoomActual)
                  kpi_df = kpi_df[kpi_df["date"].between(week_start, week_end)]
                  kpi_summary = (
                        kpi_df.groupby("kpi")["value"]
                        .sum()
                        .reset_index()
                        .rename(columns={"value": "output"})
                  )

                  # ---------- One row per position (no KPI column) ----------
                  rows = []
                  if not std_df.empty:
                        target_positions = (
                              [sel_pos] if sel_pos != "All Positions"
                              else sorted(std_df["position"].dropna().unique().tolist())
                        )

                        for position in target_positions:
                              pos_std = std_df[std_df["position"] == position].copy()
                              if pos_std.empty:
                                    continue

                              actual_hours = hours_summary.loc[
                                    hours_summary["position"] == position, "actual_hours"
                              ].sum()

                              tmp = pos_std.merge(
                                    kpi_summary.rename(columns={"kpi": "metric"}),
                                    on="metric",
                                    how="left"
                              )
                              tmp["output"] = tmp["output"].fillna(0.0)

                              # Convert raw standard (units per FTE) -> hrs/unit
                              tmp["std_hrs_per_unit"] = tmp["standard"].apply(
                                    lambda s: (8.0 / s) if (s is not None and s not in [0, ""]) else None
                              )

                              total_output = float(tmp["output"].sum())
                              if total_output > 0 and tmp["std_hrs_per_unit"].notna().any():
                                    tmp["weighted"] = tmp["output"] * tmp["std_hrs_per_unit"].fillna(0.0)
                                    std_weighted = tmp["weighted"].sum() / total_output
                              else:
                                    std_weighted = None

                              productivity = (actual_hours / total_output) if total_output > 0 else 0.0
                              if std_weighted is not None:
                                    variance = std_weighted - productivity
                                    arrow = "▲" if variance > 0 else ("▼" if variance < 0 else "")
                                    variance_str = f"{round(variance, 2)} {arrow}"
                              else:
                                    variance_str = ""

                              rows.append({
                                    "Position": position,
                                    "Output": round(total_output, 2),
                                    "Hours": round(float(actual_hours or 0), 2),
                                    "Productivity (hrs/unit)": round(productivity, 2),
                                    "Standard (hrs/unit)": round(std_weighted, 2) if std_weighted is not None else "",
                                    "Variance": variance_str
                              })

                  final_df = pd.DataFrame(rows, columns=[
                        "Position", "Output", "Hours",
                        "Productivity (hrs/unit)", "Standard (hrs/unit)", "Variance"
                  ])

                  if final_df.empty:
                        st.warning("⚠️ No productivity data available for the selected period.")
                  else:
                        # ---------- TOTAL row (weighted) ----------
                        import io

                        total_output = final_df["Output"].sum()
                        total_hours = final_df["Hours"].sum()
                        weighted_productivity = (total_hours / total_output) if total_output else 0.0

                        merged_all = std_df.merge(
                              kpi_summary.rename(columns={"kpi": "metric"}),
                              on="metric",
                              how="left"
                        )
                        merged_all["output"] = merged_all["output"].fillna(0.0)
                        merged_all["std_hrs_per_unit"] = merged_all["standard"].apply(
                              lambda s: (8.0 / s) if (s is not None and s not in [0, ""]) else None
                        )
                        if total_output and merged_all["std_hrs_per_unit"].notna().any():
                              merged_all["weighted"] = merged_all["output"] * merged_all["std_hrs_per_unit"].fillna(0.0)
                              total_weighted_std = merged_all["weighted"].sum() / total_output
                        else:
                              total_weighted_std = None

                        variance_total = (total_weighted_std - weighted_productivity) if total_weighted_std is not None else None
                        arrow_total = "▲" if (variance_total is not None and variance_total > 0) else ("▼" if (variance_total is not None and variance_total < 0) else "")

                        total_row = {
                              "Position": "TOTAL",
                              "Output": round(total_output, 2),
                              "Hours": round(total_hours, 2),
                              "Productivity (hrs/unit)": round(weighted_productivity, 2),
                              "Standard (hrs/unit)": round(total_weighted_std, 2) if total_weighted_std is not None else "",
                              "Variance": f"{round(variance_total, 2)} {arrow_total}" if variance_total is not None else ""
                        }
                        final_df = pd.concat([final_df, pd.DataFrame([total_row])], ignore_index=True)

                        # ---------- EXPORTS ----------
                        if export_format == "Excel":
                              buffer = io.BytesIO()
                              with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
                                    workbook = writer.book
                                    ws_name = "Productivity"

                                    # --- Write table beginning at row 6 (remove extra blank row) ---
                                    final_df.to_excel(
                                          writer,
                                          index=False,
                                          sheet_name=ws_name,
                                          startrow=6,
                                          startcol=1,
                                          header=False
                                    )
                                    ws = writer.sheets[ws_name]

                                    # ===== Background =====
                                    solid_bg_fmt = workbook.add_format({"bg_color": "#FFFFFF", "pattern": 1})
                                    ws.set_column(0, 51, None, solid_bg_fmt)
                                    for r in range(0, 2000):
                                          ws.set_row(r, None, solid_bg_fmt)
                                    ws.set_column(0, 0, 2)

                                    # ===== User info =====
                                    user_obj = st.session_state.get("user") or {}
                                    username   = user_obj.get("username") or user_obj.get("email") or "User"
                                    hotel_name = user_obj.get("hotel_name") or "Hotel"
                                    created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                                    # ===== Title =====
                                    title_fmt = workbook.add_format({
                                          "bold": True,
                                          "font_size": 14
                                    })
                                    ws.write("B1", "Productivity Index Report", title_fmt)

                                    # ===== Top right name + created (NO BORDER) =====
                                    right_col = 1 + len(final_df.columns) - 1

                                    ws.write(0, right_col, hotel_name,
                                          workbook.add_format({
                                                "bold": True,
                                                "align": "right",
                                                "border": 0,
                                                "bg_color": "#FFFFFF"
                                          }))

                                    ws.write(1, right_col,
                                          f"Created by {username} on {created_str}",
                                          workbook.add_format({
                                                "align": "right",
                                                "border": 0,
                                                "bg_color": "#FFFFFF"
                                          }))

                                    # ===== Metadata left (NO BORDERS) =====
                                    meta_label_fmt = workbook.add_format({
                                          "bold": True,
                                          "border": 0,
                                          "bg_color": "#FFFFFF"
                                    })
                                    meta_val_fmt = workbook.add_format({
                                          "border": 0,
                                          "bg_color": "#FFFFFF"
                                    })

                                    # ---- FIXED DATE FORMAT: 11-10-2025 ----
                                    ws.write("B3", "Department:", meta_label_fmt); ws.write("C3", sel_dept, meta_val_fmt)
                                    ws.write("B4", "Position:",   meta_label_fmt); ws.write("C4", sel_pos, meta_val_fmt)
                                    ws.write("B5", "Period:",     meta_label_fmt)
                                    ws.write("C5",
                                          f"{week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}",
                                          meta_val_fmt
                                    )

                                    # ===== Formats =====
                                    border_fmt = {"border": 1, "border_color": "#A6A6A6"}

                                    header_fmt = workbook.add_format({
                                          **border_fmt,
                                          "bold": True,
                                          "font_color": "white",
                                          "bg_color": "#16365C",
                                          "align": "center",
                                          "pattern": 1
                                    })

                                    data_fmt = workbook.add_format({
                                          **border_fmt,
                                          "align": "center"
                                    })

                                    # ---- Variance column shading ----
                                    variance_fmt = workbook.add_format({
                                          **border_fmt,
                                          "align": "center",
                                          "bg_color": "#FEF6F0",
                                          "pattern": 1
                                    })

                                    # ---- TOTAL ROW FORMAT (DCE6F1) ----
                                    total_fmt = workbook.add_format({
                                          **border_fmt,
                                          "bg_color": "#DCE6F1",
                                          "bold": True,
                                          "align": "center",
                                          "pattern": 1
                                    })

                                    # ===== Header row at 6 =====
                                    header_row = 6
                                    header_col = 1
                                    for i, val in enumerate(final_df.columns):
                                          ws.write(header_row, header_col + i, val, header_fmt)

                                    # ===== Data rows (start row 7) =====
                                    data_start_row = 7
                                    variance_col_idx = list(final_df.columns).index("Variance")

                                    for ridx, r in final_df[:-1].iterrows():
                                          for cidx, v in enumerate(r):
                                                fmt = (
                                                      variance_fmt if cidx == variance_col_idx
                                                      else data_fmt
                                                )
                                                ws.write(
                                                      data_start_row + ridx,
                                                      header_col + cidx,
                                                      v,
                                                      fmt
                                                )

                                    # ===== TOTAL ROW =====
                                    total_row_idx = len(final_df) + 6
                                    for cidx, v in enumerate(final_df.iloc[-1]):
                                          ws.write(
                                                total_row_idx,
                                                header_col + cidx,
                                                v,
                                                total_fmt     # <-- NEW color
                                          )

                                    # ===== Footer =====
                                    ws.write(
                                          total_row_idx + 2,
                                          1,
                                          "Confidential | 2025 Labor Pilot",
                                          workbook.add_format({"align": "left"})
                                    )

                                    # ===== Auto width =====
                                    for i, col in enumerate(final_df.columns):
                                          w = max(len(str(col)), final_df[col].astype(str).str.len().max())
                                          ws.set_column(1 + i, 1 + i, max(12, min(30, (w or 10) + 2)))

                              st.download_button(
                                    "📥 Download Excel",
                                    buffer.getvalue(),
                                    file_name=f"Productivity_Index_{week_start:%Y%m%d}_{week_end:%Y%m%d}.xlsx"
                              )

                        elif export_format == "PDF":
                              from reportlab.lib.pagesizes import landscape, letter
                              from reportlab.platypus import (
                                    SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer
                              )
                              from reportlab.lib.styles import getSampleStyleSheet
                              from reportlab.lib import colors
                              from datetime import datetime

                              buffer = io.BytesIO()
                              doc = SimpleDocTemplate(
                                    buffer,
                                    pagesize=landscape(letter),
                                    leftMargin=28,
                                    rightMargin=28,
                                    topMargin=24,
                                    bottomMargin=24
                              )

                              styles = getSampleStyleSheet()
                              elements = []

                              # ---------- TITLE ----------
                              elements.append(Paragraph("<b>Productivity Index Report</b>", styles["Heading2"]))
                              elements.append(Spacer(1, 6))

                              # ---------- LEFT METADATA ----------
                              elements.append(Paragraph(f"<b>Department:</b> {sel_dept}", styles["Normal"]))
                              elements.append(Paragraph(f"<b>Position:</b> {sel_pos}", styles["Normal"]))
                              elements.append(Paragraph(
                                    f"<b>Period:</b> {week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}",
                                    styles["Normal"]
                              ))
                              elements.append(Spacer(1, 12))

                              # ---------- TABLE ----------
                              pdf_data = [final_df.columns.tolist()] + final_df.values.tolist()
                              table = Table(pdf_data, repeatRows=1)

                              table_style = TableStyle([
                                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                              ])

                              variance_col = pdf_data[0].index("Variance")

                              for r in range(1, len(pdf_data)):
                                    row = pdf_data[r]

                                    # ----- TOTAL ROW → Full #DCE6F1 -----
                                    if row[0] == "TOTAL":
                                          table_style.add(
                                                "BACKGROUND",
                                                (0, r), (-1, r),
                                                colors.HexColor("#DCE6F1")
                                          )
                                          continue   # do NOT apply variance shading

                                    # ----- Variance shading for normal rows -----
                                    table_style.add(
                                          "BACKGROUND",
                                          (variance_col, r),
                                          (variance_col, r),
                                          colors.HexColor("#FEF6F0")
                                    )

                              table.setStyle(table_style)
                              elements.append(table)

                              # ---------- HEADER / FOOTER ----------
                              user_obj = st.session_state.get("user") or {}
                              username   = user_obj.get("username") or user_obj.get("email") or "User"
                              hotel_name = user_obj.get("hotel_name") or "Hotel"
                              created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                              def _first_page(canvas, doc):
                                    canvas.saveState()

                                    # (FIX) Push both lines down one row
                                    canvas.setFont("Helvetica-Bold", 10)
                                    canvas.drawRightString(
                                          doc.pagesize[0] - doc.rightMargin,
                                          doc.pagesize[1] - 40,
                                          hotel_name
                                    )

                                    canvas.setFont("Helvetica", 8)
                                    canvas.drawRightString(
                                          doc.pagesize[0] - doc.rightMargin,
                                          doc.pagesize[1] - 54,
                                          f"Created by {username} on {created_str}"
                                    )

                                    canvas.setFont("Helvetica", 8)
                                    canvas.drawString(
                                          doc.leftMargin,
                                          doc.bottomMargin - 12,
                                          "Confidential | 2025 Labor Pilot"
                                    )

                                    canvas.drawRightString(
                                          doc.pagesize[0] - doc.rightMargin,
                                          doc.bottomMargin - 12,
                                          f"Page | {canvas.getPageNumber()}"
                                    )

                                    canvas.restoreState()

                              def _later_pages(canvas, doc):
                                    canvas.saveState()

                                    canvas.setFont("Helvetica", 8)
                                    canvas.drawString(
                                          doc.leftMargin,
                                          doc.bottomMargin - 12,
                                          "Confidential | 2025 Labor Pilot"
                                    )
                                    canvas.drawRightString(
                                          doc.pagesize[0] - doc.rightMargin,
                                          doc.bottomMargin - 12,
                                          f"Page | {canvas.getPageNumber()}"
                                    )

                                    canvas.restoreState()

                              doc.build(elements, onFirstPage=_first_page, onLaterPages=_later_pages)

                              buffer.seek(0)
                              st.download_button(
                                    "📥 Download PDF",
                                    buffer.getvalue(),
                                    file_name=f"Productivity_Index_{week_start:%Y%m%d}_{week_end:%Y%m%d}.pdf",
                                    mime="application/pdf"
                              )

                        elif export_format == "CSV":
                              clean_df = final_df.copy()
                              if "Variance" in clean_df.columns:
                                    clean_df["Variance"] = (
                                          clean_df["Variance"].astype(str)
                                          .str.replace("▲", "", regex=False)
                                          .str.replace("▼", "", regex=False)
                                          .str.strip()
                                    )
                              csv_buffer = io.StringIO()
                              clean_df.to_csv(csv_buffer, index=False)
                              st.download_button(
                                    "📥 Download CSV",
                                    data=csv_buffer.getvalue().encode("utf-8"),
                                    file_name=f"Productivity_Index_{week_start:%Y%m%d}_{week_end:%Y%m%d}.csv",
                                    mime="text/csv"
                              )

      elif report_type == "Labor Standards":
            st.markdown("### Labor Standards Report")

            generate_std = st.button("📊 Generate Report")

            if generate_std:
                  import pandas as pd

                  std_df = pd.read_sql("""
                        SELECT d.name AS Department, 
                               p.name AS Position, 
                               s.metric AS Metric, 
                               s.standard AS Standard, 
                               s.unit AS Unit
                        FROM labor_standards s
                        JOIN positions p ON s.position_id = p.id
                        JOIN departments d ON p.department_id = d.id
                        ORDER BY d.name, p.name, s.metric
                  """, con=ENGINE)

                  if std_df.empty:
                        st.warning("⚠️ No labor standards found in the database.")
                  else:
                        if export_format == "Excel":
                              from io import BytesIO
                              import xlsxwriter
                              from datetime import datetime

                              output = BytesIO()
                              with pd.ExcelWriter(output, engine="xlsxwriter") as writer:

                                    workbook  = writer.book
                                    worksheet = workbook.add_worksheet("Labor Standards")
                                    writer.sheets["Labor Standards"] = worksheet

                                    # ---------- Solid White Background + Gutter ----------
                                    solid_bg_fmt = workbook.add_format({'bg_color': '#FFFFFF', 'pattern': 1})
                                    worksheet.set_column(0, 51, None, solid_bg_fmt)
                                    for r in range(0, 2000):
                                          worksheet.set_row(r, None, solid_bg_fmt)
                                    worksheet.set_column(0, 0, 2)   # gutter column A

                                    # ---------- User + hotel info ----------
                                    user_obj = st.session_state.get("user") or {}
                                    username = user_obj.get("username") or user_obj.get("email") or "User"
                                    hotel_name = user_obj.get("hotel_name") or "Hotel"
                                    created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                                    hotel_fmt = workbook.add_format({
                                          'align': 'right',
                                          'valign': 'vcenter',
                                          'bg_color': '#FFFFFF',
                                          'border': 0,
                                          'bold': True
                                    })

                                    created_fmt = workbook.add_format({
                                          'align': 'right',
                                          'valign': 'vcenter',
                                          'bg_color': '#FFFFFF',
                                          'border': 0
                                    })

                                    # ---------- Header Formats ----------
                                    title_fmt = workbook.add_format({
                                          'bold': True,
                                          'font_size': 14,
                                          'border': 0,
                                          'align': 'left'
                                    })

                                    header_fmt = workbook.add_format({
                                          'bold': True,
                                          'font_color': '#FFFFFF',
                                          'bg_color': '#16365C',
                                          'align': 'center',
                                          'valign': 'vcenter',
                                          'border': 1,
                                          'border_color': '#A6A6A6',
                                          'pattern': 1
                                    })

                                    center_fmt = workbook.add_format({
                                          'align': 'center',
                                          'valign': 'vcenter',
                                          'border': 1,
                                          'border_color': '#A6A6A6'
                                    })

                                    bold_center_fmt = workbook.add_format({
                                          'bold': True,
                                          'align': 'center',
                                          'valign': 'vcenter',
                                          'border': 1,
                                          'border_color': '#A6A6A6'
                                    })

                                    pink_fmt = workbook.add_format({
                                          'align': 'center',
                                          'valign': 'vcenter',
                                          'border': 1,
                                          'border_color': '#A6A6A6',
                                          'bg_color': '#FEF6F0',
                                          'pattern': 1
                                    })

                                    # ---------- Title ----------
                                    worksheet.write("B1", "Labor Standards Report", title_fmt)

                                    # ---------- Top-right hotel name + created ----------
                                    target_col = 5   # Column F

                                    worksheet.write(0, target_col, hotel_name,      hotel_fmt)
                                    worksheet.write(1, target_col, f"Created by {username} on {created_str}", created_fmt)
                                    # ---------- Header Row ----------
                                    header_row = 3
                                    header_col = 1
                                    for col_num, value in enumerate(std_df.columns.values):
                                          worksheet.write(header_row, header_col + col_num, value, header_fmt)

                                    # ---------- Data Rows ----------
                                    data_start_row = header_row + 1
                                    for r_idx, row in enumerate(std_df.values):
                                          for c_idx, cell in enumerate(row):
                                                excel_row = data_start_row + r_idx
                                                excel_col = header_col + c_idx

                                                # First 2 columns bold
                                                if c_idx in (0, 1):
                                                      worksheet.write(excel_row, excel_col, cell, bold_center_fmt)

                                                # Last 2 columns pink
                                                elif c_idx in (3, 4):
                                                      worksheet.write(excel_row, excel_col, cell, pink_fmt)

                                                else:
                                                      worksheet.write(excel_row, excel_col, cell, center_fmt)

                                    # ---------- Auto-fit columns ----------
                                    for i, col in enumerate(std_df.columns):
                                          try:
                                                max_w = max(len(str(col)), std_df[col].astype(str).str.len().max())
                                          except:
                                                max_w = len(str(col))
                                          worksheet.set_column(1 + i, 1 + i, max(max_w + 2, 14))

                                    # ---------- Footer ----------
                                    footer_row = data_start_row + len(std_df) + 2
                                    worksheet.write(
                                          footer_row, 1,
                                          "Confidential | 2025 Labor Pilot",
                                          workbook.add_format({'align': 'left', 'border': 0})
                                    )

                              st.download_button(
                                    label="📥 Download Excel",
                                    data=output.getvalue(),
                                    file_name="Labor_Standards_Report.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                              )

                        elif export_format == "PDF":
                              from io import BytesIO
                              from reportlab.lib.pagesizes import landscape, letter
                              from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                              from reportlab.lib.styles import getSampleStyleSheet
                              from reportlab.lib import colors
                              from datetime import datetime

                              buffer = BytesIO()
                              doc = SimpleDocTemplate(
                                    buffer,
                                    pagesize=landscape(letter),
                                    leftMargin=28,
                                    rightMargin=28,
                                    topMargin=24,
                                    bottomMargin=24
                              )

                              styles = getSampleStyleSheet()
                              elements = []

                              # ---------- TITLE ----------
                              elements.append(Paragraph("<b>Labor Standards Report</b>", styles["Heading2"]))
                              elements.append(Spacer(1, 6))

                              # ---------- TABLE DATA ----------
                              data = [list(std_df.columns)] + std_df.astype(str).values.tolist()
                              table = Table(data, repeatRows=1)

                              # ---------- TABLE STYLE ----------
                              table_style = TableStyle([
                                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                              ])

                              # first 2 cols bold
                              table_style.add("FONTNAME", (0, 1), (1, -1), "Helvetica-Bold")

                              # last 2 cols pink
                              last2_start = len(std_df.columns) - 2
                              last2_end   = len(std_df.columns) - 1
                              table_style.add(
                                    "BACKGROUND",
                                    (last2_start, 1),
                                    (last2_end, len(data) - 1),
                                    colors.HexColor("#FEF6F0")
                              )

                              table.setStyle(table_style)
                              elements.append(table)

                              elements.append(Spacer(1, 18))

                              # ---------- REMOVE DUPLICATE FOOTER (this line was the problem) ----------
                              # elements.append(Paragraph("Confidential | 2025 Labor Pilot", styles["Normal"]))

                              # ---------- HEADER & FOOTER ----------
                              user_obj = st.session_state.get("user") or {}
                              username   = user_obj.get("username") or user_obj.get("email") or "User"
                              hotel_name = user_obj.get("hotel_name") or "Hotel"
                              created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                              def _first_page(canvas, doc):
                                    canvas.saveState()

                                    # Hotel name (shifted DOWN 1 row)
                                    canvas.setFont("Helvetica-Bold", 10)
                                    canvas.drawRightString(
                                          doc.pagesize[0] - doc.rightMargin,
                                          doc.pagesize[1] - 34,
                                          hotel_name
                                    )

                                    # Created by (also pushed DOWN)
                                    canvas.setFont("Helvetica", 8)
                                    canvas.drawRightString(
                                          doc.pagesize[0] - doc.rightMargin,
                                          doc.pagesize[1] - 48,
                                          f"Created by {username} on {created_str}"
                                    )

                                    # Footer left
                                    canvas.setFont("Helvetica", 8)
                                    canvas.drawString(
                                          doc.leftMargin,
                                          doc.bottomMargin - 12,
                                          "Confidential | 2025 Labor Pilot"
                                    )

                                    # Footer right
                                    canvas.drawRightString(
                                          doc.pagesize[0] - doc.rightMargin,
                                          doc.bottomMargin - 12,
                                          f"Page | {canvas.getPageNumber()}"
                                    )

                                    canvas.restoreState()

                              def _later_pages(canvas, doc):
                                    canvas.saveState()

                                    canvas.setFont("Helvetica", 8)
                                    canvas.drawString(
                                          doc.leftMargin,
                                          doc.bottomMargin - 12,
                                          "Confidential | 2025 Labor Pilot"
                                    )

                                    canvas.drawRightString(
                                          doc.pagesize[0] - doc.rightMargin,
                                          doc.bottomMargin - 12,
                                          f"Page | {canvas.getPageNumber()}"
                                    )

                                    canvas.restoreState()

                              doc.build(elements, onFirstPage=_first_page, onLaterPages=_later_pages)

                              st.download_button(
                                    label="📥 Download PDF",
                                    data=buffer.getvalue(),
                                    file_name="Labor_Standards_Report.pdf",
                                    mime="application/pdf"
                              )

                        elif export_format == "CSV":
                              csv_data = std_df.to_csv(index=False).encode("utf-8")
                              st.download_button(
                                    label="📥 Download CSV",
                                    data=csv_data,
                                    file_name="Labor_Standards_Report.csv",
                                    mime="text/csv"
                              )

      elif report_type == "Labor Variance":
            st.markdown("### Labor Variance Report Filters")

            from datetime import date, timedelta
            from dateutil.relativedelta import relativedelta, MO
            import pandas as pd
            import io

            if "labor_var_date" not in st.session_state:
                  st.session_state.labor_var_date = date.today()

            sel_date = st.date_input("Select any date in the week", value=st.session_state.labor_var_date, key="labor_var_date")
            week_start = sel_date + relativedelta(weekday=MO(-1))
            week_end   = week_start + timedelta(days=6)

            st.markdown(f"📅 **Selected Week:** {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}")

            emp_df = refresh(db.Employee)

            # 🔒 APPLY MANAGER SCOPE FIRST
            emp_df = apply_manager_scope(
                  emp_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})

            # ── Department supports "All" (SCOPED)
            dept_list = ["All"] + sorted(
                  emp_df["department"]
                  .dropna()
                  .unique()
                  .tolist()
            )

            sel_dept = st.selectbox(
                  "Department",
                  dept_list,
                  key="labor_var_dept"
            )

            # ── Position options respect department AND scope
            if sel_dept == "All":
                  pos_list = (
                        emp_df["role"]
                        .dropna()
                        .unique()
                        .tolist()
                  )
            else:
                  pos_list = (
                        emp_df.loc[
                              emp_df["department"] == sel_dept,
                              "role"
                        ]
                        .dropna()
                        .unique()
                        .tolist()
                  )

            sel_pos = st.selectbox(
                  "Position",
                  ["All"] + sorted(pos_list),
                  key="labor_var_pos"
            )

            # ───── Pre-check for Data Availability ─────
            schedule_df = refresh(db.Schedule)
            actual_df   = refresh(db.Actual)
            room_df     = refresh(db.RoomActual)

            has_schedule = not schedule_df[(schedule_df["day"] >= week_start) & (schedule_df["day"] <= week_end)].empty
            has_actual   = not actual_df[(actual_df["date"] >= week_start) & (actual_df["date"] <= week_end)].empty
            has_room     = not room_df[(room_df["date"] >= week_start) & (room_df["date"] <= week_end)].empty

            generate_labor_var = st.button("📊 Generate Labor Variance Report")
            if generate_labor_var:
                  schedule_df     = refresh(db.Schedule)
                  actual_df       = refresh(db.Actual)
                  room_actual_df  = refresh(db.RoomActual)
                  std_df          = refresh(db.LaborStandard)

                  # ⬇️ Positions joined with Department so each row knows its dept
                  pos_df_raw = refresh(db.Position).rename(columns={"id": "position_id", "name": "position", "department_id": "dept_id"})
                  dept_df    = refresh(db.Department).rename(columns={"id": "dept_id", "name": "department"})
                  pos_df     = pos_df_raw.merge(dept_df, on="dept_id", how="left")

                  has_schedule = not schedule_df[(schedule_df["day"] >= week_start) & (schedule_df["day"] <= week_end)].empty
                  has_actual   = not actual_df[(actual_df["date"] >= week_start) & (actual_df["date"] <= week_end)].empty
                  has_room     = not room_actual_df[(room_actual_df["date"] >= week_start) & (room_actual_df["date"] <= week_end)].empty

                  if not (has_schedule or has_actual or has_room):
                        st.warning("⚠️ No data available for the selected week. Labor Variance Report cannot be generated.")
                        st.session_state.labor_variance_ready = False
                        st.stop()

                  # ── Employees subset respecting selections
                  emp_filtered = emp_df.copy()
                  if sel_dept != "All":
                        emp_filtered = emp_filtered[emp_filtered["department"] == sel_dept]
                  if sel_pos != "All":
                        emp_filtered = emp_filtered[emp_filtered["role"] == sel_pos]

                  pos_names = emp_filtered["role"].dropna().unique()
                  pos_match = pos_df[pos_df["position"].isin(pos_names)]

                  if pos_match.empty:
                        st.warning("⚠️ No positions found matching employee roles.")
                        st.session_state.labor_variance_ready = False
                        st.stop()

                  # Actual hours summary by position_id for the week
                  ah_df = actual_df[
                        (actual_df["source"].isin(["manual", "contract"])) &
                        (actual_df["date"].between(week_start, week_end))
                  ]
                  ah_df["total_hours"] = ah_df[["hours", "ot_hours"]].sum(axis=1)
                  actual_summary = (
                        ah_df.groupby("position_id")["total_hours"]
                        .sum()
                        .reset_index()
                        .rename(columns={"total_hours": "actual_hours"})
                  )
                  actual_summary = actual_summary.merge(pos_df[["position_id", "position"]], on="position_id", how="left")

                  results = []
                  seen_positions = set()

                  for _, row in pos_match.iterrows():
                        pos_name  = row["position"]
                        pos_id    = row["position_id"]
                        dept_name = row["department"]  # department owning this position

                        if pos_name in seen_positions and sel_dept != "All":
                              continue
                        # when "All", the same position name can exist in many departments → keep per dept
                        if sel_dept != "All":
                              seen_positions.add(pos_name)

                        # ── Employee IDs by position & department (when "All" use row dept)
                        if sel_dept == "All":
                              emp_ids = emp_df[
                                    (emp_df["role"] == pos_name) &
                                    (emp_df["department"] == dept_name)
                              ]["id"].tolist()
                              dept_value = dept_name
                        else:
                              emp_ids = emp_df[
                                    (emp_df["department"] == sel_dept) &
                                    (emp_df["role"] == pos_name)
                              ]["id"].tolist()
                              dept_value = sel_dept

                        # Scheduled hours for the week
                        sched_rows = schedule_df[
                              (schedule_df["emp_id"].isin(emp_ids)) &
                              (schedule_df["day"] >= week_start) &
                              (schedule_df["day"] <= week_end)
                        ]
                        sched_hours = sched_rows["shift_type"].apply(
                              lambda x: 0 if str(x).strip().upper() == "OFF" else (
                                    pd.to_datetime(str(x).split("-")[1]) - pd.to_datetime(str(x).split("-")[0])
                              ).seconds / 3600
                        ).sum()

                        # Actual hours (manual/contract) by position
                        actual_hours = actual_summary.loc[
                              actual_summary["position"] == pos_name, "actual_hours"
                        ].sum()

                        # Projected hours from LaborStandard × actual outputs
                        std_pos = std_df[std_df["position_id"] == pos_id]
                        proj_hours_total = 0
                        for _, std_row in std_pos.iterrows():
                              metric   = std_row["metric"]
                              standard = std_row["standard"]
                              if not standard or standard == 0:
                                    continue
                              actual_output = room_actual_df[
                                    (room_actual_df["kpi"] == metric) &
                                    (room_actual_df["date"] >= week_start) &
                                    (room_actual_df["date"] <= week_end)
                              ]["value"].sum()
                              proj_hours_total += (actual_output / standard) * 8

                        projected_hours = proj_hours_total
                        variance        = actual_hours - projected_hours
                        variance_pct    = (variance / projected_hours * 100) if projected_hours else 0

                        results.append({
                              "Department": dept_value,
                              "Position": pos_name,
                              "Scheduled Hours": round(sched_hours, 1),
                              "Actual Hours": round(actual_hours, 1),
                              "Projected Hours": round(projected_hours, 1),
                              "Variance": round(variance, 1),
                              "Variance %": f"{'▲' if variance > 0 else '▼' if variance < 0 else ''} {abs(variance_pct):.2f}%" if projected_hours else "–"
                        })

                  report_df = pd.DataFrame(results)

                  # ───── Add Total Row After Deduplication ─────
                  total_row = {
                        "Department": sel_dept if sel_dept != "All" else "All Departments",
                        "Position": "TOTAL",
                        "Scheduled Hours": round(report_df["Scheduled Hours"].sum(), 1),
                        "Actual Hours": round(report_df["Actual Hours"].sum(), 1),
                        "Projected Hours": round(report_df["Projected Hours"].sum(), 1),
                        "Variance": round(report_df["Variance"].sum(), 1)
                  }

                  if total_row["Projected Hours"]:
                        total_var_pct = (total_row["Variance"] / total_row["Projected Hours"]) * 100
                        total_row["Variance %"] = f"{'▲' if total_var_pct > 0 else '▼' if total_var_pct < 0 else ''} {abs(total_var_pct):.2f}%"
                  else:
                        total_row["Variance %"] = "–"

                  report_df.loc[len(report_df)] = total_row

                  st.session_state.labor_variance_data = report_df
                  st.session_state.labor_variance_ready = True

                  # ─────────────── Excel Export ───────────────
                  if export_format == "Excel":
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                              # Table starts at row 7 (index 6), column B (index 1)
                              report_df.to_excel(
                                    writer,
                                    sheet_name="Labor Variance",
                                    index=False,
                                    startrow=6,
                                    startcol=1,   # column B
                                    header=False
                              )
                              workbook  = writer.book
                              worksheet = writer.sheets["Labor Variance"]

                              # ───── Solid sheet background ─────
                              solid_bg_fmt = workbook.add_format({'bg_color': '#FFFFFF', 'pattern': 1})
                              worksheet.set_column(0, 51, None, solid_bg_fmt)
                              for r in range(0, 2000):
                                    worksheet.set_row(r, None, solid_bg_fmt)

                              worksheet.set_column(0, 0, 2)  # gutter

                              # ───── Formats (NO ITALICS) ─────
                              title_fmt        = workbook.add_format({'bold': True, 'font_size': 14})

                              # Borderless formats for metadata block
                              borderless_label_fmt = workbook.add_format({
                                    'bold': True,
                                    'border': 0,
                                    'bg_color': '#FFFFFF'
                              })
                              borderless_value_fmt = workbook.add_format({
                                    'border': 0,
                                    'bg_color': '#FFFFFF'
                              })

                              header_fmt       = workbook.add_format({
                                    'bold': True,
                                    'bg_color': '#16365C',
                                    'font_color': '#FFFFFF',
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'align': 'center',
                                    'pattern': 1
                              })

                              dept_fmt         = workbook.add_format({
                                    'bold': True,
                                    'align': 'center',
                                    'border': 1,
                                    'border_color': '#A6A6A6'
                              })

                              pos_fmt          = workbook.add_format({
                                    'bold': True,
                                    'align': 'center',
                                    'border': 1,
                                    'border_color': '#A6A6A6'
                              })

                              default_fmt      = workbook.add_format({
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'num_format': '#,##0.00',
                                    'align': 'center'
                              })

                              variance_fmt     = workbook.add_format({
                                    'bg_color': '#FEF6F0',
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'num_format': '#,##0.00',
                                    'align': 'center',
                                    'pattern': 1
                              })

                              total_fmt        = workbook.add_format({
                                    'bg_color': '#DCE6F1',
                                    'bold': True,
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'num_format': '#,##0.00',
                                    'align': 'center',
                                    'pattern': 1
                              })

                              total_text_fmt   = workbook.add_format({
                                    'bg_color': '#DCE6F1',
                                    'bold': True,
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'align': 'center',
                                    'pattern': 1
                              })

                              total_varpct_fmt = workbook.add_format({
                                    'bg_color': '#DCE6F1',
                                    'bold': True,
                                    'border': 1,
                                    'border_color': '#A6A6A6',
                                    'align': 'center',
                                    'pattern': 1
                              })

                              # ───── Title + Metadata (NO BORDERS) ─────
                              worksheet.write("B1", "Labor Variance Report", title_fmt)

                              worksheet.write("B3", "Department:", borderless_label_fmt)
                              worksheet.write("C3", sel_dept, borderless_value_fmt)

                              worksheet.write("B4", "Position:", borderless_label_fmt)
                              worksheet.write("C4", sel_pos, borderless_value_fmt)

                              worksheet.write("B5", "Week:", borderless_label_fmt)
                              worksheet.write("C5", f"{week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}", borderless_value_fmt)

                              # ───── Header Row (row 7 index 6) ─────
                              header_row = 6
                              header_col = 1
                              for col_idx, col in enumerate(report_df.columns):
                                    worksheet.write(header_row, header_col + col_idx, col, header_fmt)

                              # ───── Precompute TOTAL % Variance ─────
                              total_actual    = report_df.loc[report_df["Position"] != "TOTAL", "Actual Hours"].sum()
                              total_projected = report_df.loc[report_df["Position"] != "TOTAL", "Projected Hours"].sum()
                              if total_projected:
                                    total_var    = total_actual - total_projected
                                    total_varpct = f"{'▲' if total_var > 0 else '▼' if total_var < 0 else ''} {abs(total_var / total_projected * 100):.2f}%"
                              else:
                                    total_varpct = "–"

                              # ───── Data Rows (start row 8 index 7) ─────
                              data_start_row = 7
                              data_start_col = 1
                              for row_idx, row in report_df.iterrows():
                                    is_total = row["Position"] == "TOTAL"
                                    for col_idx, col in enumerate(report_df.columns):
                                          val = row[col]
                                          if is_total and col == "Variance %":
                                                val = total_varpct

                                          if col == "Department":
                                                fmt = total_text_fmt if is_total else dept_fmt
                                          elif col == "Position":
                                                fmt = total_text_fmt if is_total else pos_fmt
                                          elif col == "Variance %":
                                                fmt = total_varpct_fmt if is_total else variance_fmt
                                          elif "Variance" in col:
                                                fmt = total_fmt if is_total else variance_fmt
                                          else:
                                                fmt = total_fmt if is_total else default_fmt

                                          worksheet.write(data_start_row + row_idx, data_start_col + col_idx, val, fmt)

                              # ───── Column widths ─────
                              dept_i = report_df.columns.get_loc("Department")
                              pos_i  = report_df.columns.get_loc("Position")

                              dept_width = int(max(len("Department"), report_df["Department"].astype(str).str.len().max()) * 1.1) + 2
                              pos_width  = int(max(len("Position"),  report_df["Position"].astype(str).str.len().max())  * 1.1) + 2

                              dept_width = max(18, min(dept_width, 60))
                              pos_width  = max(18, min(pos_width, 60))

                              worksheet.set_column(1 + dept_i, 1 + dept_i, dept_width)
                              worksheet.set_column(1 + pos_i,  1 + pos_i,  pos_width)

                              for i, _ in enumerate(report_df.columns):
                                    if i in (dept_i, pos_i):
                                          continue
                                    worksheet.set_column(1 + i, 1 + i, 18)

                              # ───── Top-right Hotel + Created by (NO BORDERS) ─────
                              user_obj = st.session_state.get("user") or {}
                              username = user_obj.get("username") or user_obj.get("email") or "User"
                              hotel_name = user_obj.get("hotel_name") or "Hotel"
                              from datetime import datetime
                              created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                              rightmost_col = 1 + len(report_df.columns) - 1

                              worksheet.write(0, rightmost_col, hotel_name,
                                    workbook.add_format({'bold': True, 'align': 'right', 'border': 0, 'bg_color': '#FFFFFF'}))

                              worksheet.write(1, rightmost_col,
                                    f"Created by {username} on {created_str}",
                                    workbook.add_format({'align': 'right', 'border': 0, 'bg_color': '#FFFFFF'}))

                              # ───── Bottom-left Confidential (NO ©) ─────
                              bottom_row = data_start_row + len(report_df) + 2
                              worksheet.write(bottom_row, 1,
                                    "Confidential | 2025 Labor Pilot",
                                    workbook.add_format({'align': 'left', 'border': 0, 'bg_color': '#FFFFFF'}))

                        st.download_button(
                              "📥 Download Excel",
                              data=output.getvalue(),
                              file_name=f"Labor_Variance_Report_{week_start:%Y%m%d}.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                  # ─────────────── PDF Export ───────────────
                  elif export_format == "PDF":
                        import io
                        from reportlab.lib.pagesizes import landscape, letter
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib import colors
                        from reportlab.lib.styles import getSampleStyleSheet
                        from datetime import datetime

                        buffer = io.BytesIO()
                        doc = SimpleDocTemplate(
                              buffer,
                              pagesize=landscape(letter),
                              leftMargin=28,
                              rightMargin=28,
                              topMargin=24,
                              bottomMargin=24
                        )
                        styles = getSampleStyleSheet()
                        elements = []

                        # -------------------- TITLE + METADATA --------------------
                        title_style = styles["Heading2"]
                        elements.append(Paragraph("<b>Labor Variance Report</b>", title_style))
                        elements.append(Spacer(1, 8))

                        elements.append(Paragraph(f"<b>Department:</b> {sel_dept}", styles["Normal"]))
                        elements.append(Paragraph(f"<b>Position:</b> {sel_pos}", styles["Normal"]))
                        elements.append(Paragraph(f"<b>Week:</b> {week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}", styles["Normal"]))
                        elements.append(Spacer(1, 12))

                        # -------------------- TABLE BUILD --------------------
                        # Convert data with formatted numbers
                        pdf_data = [report_df.columns.tolist()]

                        for _, row in report_df.iterrows():
                              formatted_row = []
                              for col in report_df.columns:
                                    val = row[col]

                                    # Format numeric cells with commas & decimals
                                    if col not in ["Department", "Position", "Variance %"]:
                                          try:
                                                formatted_row.append(f"{float(val):,.2f}")
                                          except:
                                                formatted_row.append(val)
                                    else:
                                          formatted_row.append(val)

                              pdf_data.append(formatted_row)

                        # Table setup
                        table = Table(pdf_data)
                        style = TableStyle([
                              ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("FONTSIZE", (0, 0), (-1, -1), 9),
                              ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                              ("TOPPADDING", (0, 0), (-1, -1), 4),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ])

                        # TOTAL row formatting + variance shading
                        for r in range(1, len(pdf_data)):
                              is_total = pdf_data[r][1] == "TOTAL"
                              if is_total:
                                    style.add("BACKGROUND", (0, r), (-1, r), colors.HexColor("#DCE6F1"))
                                    style.add("FONTNAME", (0, r), (-1, r), "Helvetica-Bold")
                              else:
                                    style.add("BACKGROUND", (-2, r), (-1, r), colors.HexColor("#FEF6F0"))

                        table.setStyle(style)
                        elements.append(table)

                        # -------------------- FIRST PAGE HEADER BLOCK --------------------
                        user_obj = st.session_state.get("user") or {}
                        username = user_obj.get("username") or user_obj.get("email") or "User"
                        hotel_name = user_obj.get("hotel_name") or "Hotel"
                        created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                        def _first_page(canvas, doc):
                              canvas.saveState()

                              # Top-right: hotel name
                              canvas.setFont("Helvetica-Bold", 10)
                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.pagesize[1] - doc.topMargin - 10,
                                    str(hotel_name)
                              )

                              # Created by
                              canvas.setFont("Helvetica", 8)
                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.pagesize[1] - doc.topMargin - 22,
                                    f"Created by {username} on {created_str}"
                              )

                              # Bottom-left footer
                              canvas.setFont("Helvetica", 8)
                              canvas.drawString(
                                    doc.leftMargin,
                                    doc.bottomMargin - 14,
                                    "Confidential | 2025 Labor Pilot"
                              )

                              # Bottom-right page #
                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.bottomMargin - 14,
                                    f"Page | {canvas.getPageNumber()}"
                              )

                              canvas.restoreState()

                        # -------------------- LATER PAGE FOOTER --------------------
                        def _later_pages(canvas, doc):
                              canvas.saveState()

                              canvas.setFont("Helvetica", 8)
                              canvas.drawString(
                                    doc.leftMargin,
                                    doc.bottomMargin - 14,
                                    "Confidential | 2025 Labor Pilot"
                              )

                              canvas.drawRightString(
                                    doc.pagesize[0] - doc.rightMargin,
                                    doc.bottomMargin - 14,
                                    f"Page | {canvas.getPageNumber()}"
                              )

                              canvas.restoreState()

                        # Build PDF
                        doc.build(elements, onFirstPage=_first_page, onLaterPages=_later_pages)

                        buffer.seek(0)
                        st.download_button(
                              "📥 Download PDF",
                              data=buffer.getvalue(),
                              file_name=f"Labor_Variance_Report_{week_start:%Y%m%d}.pdf",
                              mime="application/pdf"
                        )

                  # ─────────────── CSV Export ───────────────
                  elif export_format == "CSV":
                        csv_df = report_df.copy()
                        # If you prefer numeric % without arrows in CSV, adjust here
                        csv_data = csv_df.to_csv(index=False, encoding="utf-8-sig")
                        st.download_button(
                              "📥 Download CSV",
                              csv_data,
                              file_name=f"Labor_Variance_Report_{week_start:%Y%m%d}.csv",
                              mime="text/csv"
                        )

      elif report_type == "Schedule Variance":
            st.markdown("### Schedule Variance Report Filters")

            from datetime import date, timedelta
            from dateutil.relativedelta import relativedelta, MO
            import pandas as pd
            import io

            if "sched_var_date" not in st.session_state:
                  st.session_state.sched_var_date = date.today()

            sel_date = st.date_input("Select any date in the week", value=st.session_state.sched_var_date, key="sched_var_date")
            week_start = sel_date + relativedelta(weekday=MO(-1))
            week_end   = week_start + timedelta(days=6)

            st.markdown(f"📅 **Selected Week:** {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}")

            emp_df = refresh(db.Employee)

            # ─────────────── EMPLOYEE SCOPE (SOURCE OF TRUTH) ───────────────
            emp_df = refresh(db.Employee)

            emp_df = apply_manager_scope(
                  emp_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})

            # ─────────────── Department (SCOPED) ───────────────
            dept_opts = ["All"] + sorted(
                  emp_df["department"]
                  .dropna()
                  .unique()
                  .tolist()
            )

            sel_dept = st.selectbox(
                  "Department",
                  dept_opts,
                  key="sched_var_dept"
            )

            generate_sched_var = st.button("📊 Generate Schedule Variance Report")
            if generate_sched_var:
                  schedule_df     = refresh(db.Schedule)
                  room_actual_df  = refresh(db.RoomActual)
                  std_df          = refresh(db.LaborStandard)

                  # ⬇️ Positions joined with their department names
                  pos_df = refresh(db.Position).rename(columns={"id": "position_id", "name": "position", "department_id": "dept_id"})
                  dept_df = refresh(db.Department).rename(columns={"id": "dept_id", "name": "department"})
                  pos_df = pos_df.merge(dept_df, on="dept_id", how="left")

                  # ── Employee roles to consider (respect "All")
                  emp_filtered = emp_df.copy()
                  if sel_dept != "All":
                        emp_filtered = emp_filtered[emp_filtered["department"] == sel_dept]

                  pos_names = emp_filtered["role"].dropna().unique()
                  # keep department with each position
                  pos_match = pos_df[pos_df["position"].isin(pos_names)]

                  results = []
                  for _, row in pos_match.iterrows():
                        pos_name   = row["position"]
                        pos_id     = row["position_id"]
                        dept_name  = row["department"]  # department for THIS position row

                        # ── Employee IDs for this position/department
                        if sel_dept == "All":
                              emp_ids = emp_df[
                                    (emp_df["role"] == pos_name) &
                                    (emp_df["department"] == dept_name)
                              ]["id"].tolist()
                              dept_value = dept_name
                        else:
                              emp_ids = emp_df[
                                    (emp_df["department"] == sel_dept) &
                                    (emp_df["role"] == pos_name)
                              ]["id"].tolist()
                              dept_value = sel_dept

                        sched_rows = schedule_df[
                              (schedule_df["emp_id"].isin(emp_ids)) &
                              (schedule_df["day"] >= week_start) &
                              (schedule_df["day"] <= week_end)
                        ]
                        sched_hours = sched_rows["shift_type"].apply(
                              lambda x: 0 if str(x).strip().upper() == "OFF" else (
                                    pd.to_datetime(str(x).split("-")[1]) - pd.to_datetime(str(x).split("-")[0])
                              ).seconds / 3600
                        ).sum()

                        std_pos = std_df[std_df["position_id"] == pos_id]
                        proj_hours_total = 0
                        for _, std_row in std_pos.iterrows():
                              metric = std_row["metric"]
                              standard = std_row["standard"]
                              if not standard or standard == 0:
                                    continue
                              actual_output = room_actual_df[
                                    (room_actual_df["kpi"] == metric) &
                                    (room_actual_df["date"] >= week_start) &
                                    (room_actual_df["date"] <= week_end)
                              ]["value"].sum()
                              proj_hours_total += (actual_output / standard) * 8

                        projected_hours = proj_hours_total
                        variance = sched_hours - projected_hours
                        variance_pct = (variance / projected_hours * 100) if projected_hours else 0

                        results.append({
                              "Department": dept_value,          # department per row
                              "Position": pos_name,
                              "Scheduled Hours": round(sched_hours, 1),
                              "Projected Hours": round(projected_hours, 1),
                              "Variance": round(variance, 1),
                              "Variance %": f"{'▲' if variance > 0 else '▼' if variance < 0 else ''} {abs(variance_pct):.2f}%" if projected_hours else "–"
                        })

                  report_df = (
                        pd.DataFrame(results)
                        .groupby(["Department", "Position"], as_index=False)
                        .sum(numeric_only=True)
                        .sort_values(["Department", "Position"])
                  )

                  total_row = {
                        "Department": sel_dept if sel_dept != "All" else "All Departments",
                        "Position": "TOTAL",
                        "Scheduled Hours": report_df["Scheduled Hours"].sum(),
                        "Projected Hours": report_df["Projected Hours"].sum(),
                        "Variance": report_df["Variance"].sum(),
                        "Variance %": ""
                  }

                  if total_row["Projected Hours"]:
                        total_pct = (total_row["Variance"] / total_row["Projected Hours"]) * 100
                        total_row["Variance %"] = f"{'▲' if total_pct > 0 else '▼' if total_pct < 0 else ''} {abs(total_pct):.2f}%"
                  else:
                        total_row["Variance %"] = "–"

                  report_df.loc[len(report_df)] = total_row

                  st.session_state.schedule_variance_df = report_df
                  st.session_state.schedule_variance_ready = True

                  if export_format == "Excel" and st.session_state.get("schedule_variance_ready"):
                        sv_df = st.session_state.schedule_variance_df.copy()

                        # ====== FIX 1: Remove group header rows ======
                        sv_df = sv_df[~sv_df["Scheduled Hours"].isna()].copy()

                        # ====== FIX 2: Only keep rows that have variance ≠ 0 OR TOTAL ======
                        filtered_df = sv_df[
                              (sv_df["Variance"] != 0) |
                              (sv_df["Position"] == "TOTAL")
                        ].copy()

                        # ====== FIX 3: Push total to bottom ======
                        total_row = filtered_df[filtered_df["Position"] == "TOTAL"]
                        filtered_df = filtered_df[filtered_df["Position"] != "TOTAL"]
                        filtered_df = pd.concat([filtered_df, total_row], ignore_index=True)

                        import io
                        from datetime import datetime

                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:

                              # Write data starting at row 8 (index 7)
                              filtered_df.to_excel(
                                    writer,
                                    sheet_name="Schedule Variance",
                                    index=False,
                                    startrow=7,
                                    startcol=1,
                                    header=False
                              )

                              workbook  = writer.book
                              worksheet = writer.sheets["Schedule Variance"]

                              # ====== WHITE BACKGROUND ======
                              solid_bg = workbook.add_format({'bg_color': '#FFFFFF', 'pattern': 1})
                              worksheet.set_column(0, 51, None, solid_bg)
                              for r in range(0, 2000):
                                    worksheet.set_row(r, None, solid_bg)
                              worksheet.set_column(0, 0, 2)   # gutter column

                              # ====== User Info ======
                              user_obj = st.session_state.get("user") or {}
                              username = user_obj.get("username") or user_obj.get("email") or "User"
                              hotel_name = user_obj.get("hotel_name") or "Hotel"
                              created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                              right_fmt = workbook.add_format({'align': 'right','valign':'vcenter','bg_color':'#FFFFFF'})
                              right_bold = workbook.add_format({'align':'right','valign':'vcenter','bg_color':'#FFFFFF','bold':True})

                              # ====== Formats ======
                              title_fmt = workbook.add_format({'bold': True, 'font_size': 14})
                              label_fmt = workbook.add_format({'bold': True})

                              header_fmt = workbook.add_format({
                                    'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#16365C',
                                    'align': 'center', 'border': 1, 'border_color': '#A6A6A6', 'pattern': 1
                              })

                              dept_fmt = workbook.add_format({'bold': True,'align':'center','border':1,'border_color':'#A6A6A6'})
                              pos_fmt  = workbook.add_format({'bold':True,'italic':True,'align':'center','border':1,'border_color':'#A6A6A6'})

                              default_fmt = workbook.add_format({'align':'center','border':1,'border_color':'#A6A6A6','num_format':'#,##0.00'})
                              variance_fmt = workbook.add_format({'align':'center','border':1,'border_color':'#A6A6A6','bg_color':'#FEF6F0','num_format':'#,##0.00','pattern':1})

                              total_fmt = workbook.add_format({'align':'center','border':1,'border_color':'#A6A6A6','bg_color':'#DCE6F1','bold':True,'num_format':'#,##0.00'})
                              total_text_fmt = workbook.add_format({'align':'center','border':1,'border_color':'#A6A6A6','bg_color':'#DCE6F1','bold':True,'italic':True})
                              total_varpct_fmt = workbook.add_format({'align':'center','border':1,'border_color':'#A6A6A6','bg_color':'#DCE6F1','bold':True})

                              # ====== CLEAN FORMAT (NO BORDER) — FIX FOR DEPARTMENT/WEEK ======
                              clean_fmt = workbook.add_format({
                                    'border': 0, 'align': 'left', 'valign': 'vcenter', 'bg_color': '#FFFFFF'
                              })

                              # ====== Title ======
                              worksheet.write("B1", "Schedule Variance Report", title_fmt)

                              # ====== Metadata (labels bold, values clean — NO BORDERS) ======
                              worksheet.write("B3", "Department:", label_fmt)
                              worksheet.write("B4", "Week:", label_fmt)

                              worksheet.write("C3", sel_dept or "(All)", clean_fmt)
                              worksheet.write("C4", f"{week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}", clean_fmt)

                              # ====== FULL CLEAN-UP OF METADATA ROWS (remove leftover borders) ======
                              for col in range(1, 10):   # B..J
                                    worksheet.write_blank(2, col, None, clean_fmt)  # row 3
                                    worksheet.write_blank(3, col, None, clean_fmt)  # row 4

                              # Re-write labels + values AFTER cleanup
                              worksheet.write("B3", "Department:", label_fmt)
                              worksheet.write("B4", "Week:", label_fmt)
                              worksheet.write("C3", sel_dept or "(All)", clean_fmt)
                              worksheet.write("C4", f"{week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}", clean_fmt)

                              # ====== Top-right name + creation in COLUMN F ======
                              worksheet.write(0, 5, hotel_name, right_bold)
                              worksheet.write(1, 5, f"Created by {username} on {created_str}", right_fmt)

                              # ====== Header Row ======
                              header_row = 6
                              header_col = 1
                              for col_idx, col in enumerate(filtered_df.columns):
                                    worksheet.write(header_row, header_col + col_idx, col, header_fmt)

                              # ====== Data Rows ======
                              for row_idx, row in filtered_df.iterrows():
                                    is_total = (row["Position"] == "TOTAL")
                                    for col_idx, col in enumerate(filtered_df.columns):
                                          val = row[col]
                                          excel_row = 7 + row_idx
                                          excel_col = header_col + col_idx

                                          if col == "Department":
                                                fmt = total_text_fmt if is_total else dept_fmt
                                          elif col == "Position":
                                                fmt = total_text_fmt if is_total else pos_fmt
                                          elif col == "Variance":
                                                fmt = total_fmt if is_total else variance_fmt
                                          else:
                                                fmt = total_fmt if is_total else default_fmt

                                          worksheet.write(excel_row, excel_col, val, fmt)

                              # ====== Auto-size ======
                              for i, col in enumerate(filtered_df.columns):
                                    try:
                                          max_w = max(len(str(col)), filtered_df[col].astype(str).str.len().max())
                                    except:
                                          max_w = len(str(col))
                                    worksheet.set_column(1 + i, 1 + i, max(max_w + 2, 12))

                              # ====== Footer ======
                              footer_row = 7 + len(filtered_df) + 2
                              worksheet.write(footer_row, 1, "Confidential | 2025 Labor Pilot",
                                              workbook.add_format({'align':'left','border':0}))

                        st.download_button(
                              "📥 Download Excel",
                              data=output.getvalue(),
                              file_name=f"Schedule_Variance_Report_{week_start:%Y%m%d}.xlsx",
                              mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                  elif export_format == "PDF" and st.session_state.get("schedule_variance_ready"):
                        import io
                        from datetime import datetime
                        from reportlab.lib.pagesizes import landscape, letter
                        from reportlab.platypus import (
                              SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        )
                        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                        from reportlab.lib.enums import TA_RIGHT
                        from reportlab.lib import colors

                        # --- USER INFO ---
                        user_obj = st.session_state.get("user") or {}
                        username = user_obj.get("username") or user_obj.get("email") or "User"
                        hotel_name = user_obj.get("hotel_name") or "Hotel"

                        pdf_df = st.session_state.schedule_variance_df.copy()

                        # --- FILTER (same as Excel) ---
                        pdf_df = pdf_df[~pdf_df["Scheduled Hours"].isna()].copy()
                        pdf_df = pdf_df[
                              (pdf_df["Variance"] != 0) |
                              (pdf_df["Position"] == "TOTAL")
                        ].copy()
                        total_row = pdf_df[pdf_df["Position"] == "TOTAL"]
                        pdf_df = pdf_df[pdf_df["Position"] != "TOTAL"]
                        pdf_df = pd.concat([pdf_df, total_row], ignore_index=True)

                        # --- Number formatting ---
                        num_cols = ["Scheduled Hours", "Projected Hours", "Variance"]
                        for col in num_cols:
                              pdf_df[col] = pdf_df[col].apply(lambda x: f"{x:,.1f}")

                        buffer = io.BytesIO()

                        # ---------- FOOTER FIX ----------
                        def footer(canvas, doc):
                              canvas.saveState()
                              canvas.setFont("Helvetica", 8)

                              # left footer
                              canvas.drawString(36, 20, "Confidential | 2025 Labor Pilot")

                              # right footer
                              page_text = f"Page {doc.page}"
                              canvas.drawRightString(letter[1] - 36, 20, page_text)

                              canvas.restoreState()

                        # ---------- DOCUMENT ----------
                        doc = SimpleDocTemplate(
                              buffer,
                              pagesize=landscape(letter),
                              leftMargin=36, rightMargin=36,
                              topMargin=36, bottomMargin=36
                        )

                        styles = getSampleStyleSheet()
                        elements = []

                        # ========= TITLE (aligned left) + RIGHT USER INFO =========
                        title_style = styles["Heading2"]

                        right_style = ParagraphStyle(
                              "right_style",
                              parent=styles["Normal"],
                              alignment=TA_RIGHT,
                              fontSize=10
                        )

                        title_table = Table(
                              [
                                    [
                                          Paragraph("Schedule Variance Report", title_style),
                                          Paragraph(
                                                f"<b>{hotel_name}</b><br/>"
                                                f"Created by {username} on {datetime.now():%m/%d/%Y %I:%M %p}",
                                                right_style
                                          )
                                    ]
                              ],
                              colWidths=["*","*"]   # <-- PERFECT ALIGNMENT FIX
                        )

                        elements.append(title_table)
                        elements.append(Spacer(1, 12))

                        # ========= METADATA =========
                        elements.append(Paragraph(f"<b>Department:</b> {sel_dept or '(All)'}", styles["Normal"]))
                        elements.append(Paragraph(
                              f"<b>Week:</b> {week_start:%m-%d-%Y} to {week_end:%m-%d-%Y}",
                              styles["Normal"]
                        ))
                        elements.append(Spacer(1, 16))

                        # ========= TABLE =========
                        pdf_data = [list(pdf_df.columns)] + pdf_df.values.tolist()
                        table = Table(pdf_data, repeatRows=1)

                        style = TableStyle([
                              ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                              ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                              ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                              ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                              ("FONTSIZE", (0, 0), (-1, 0), 9),
                              ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                              ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                        ])

                        var_col = list(pdf_df.columns).index("Variance")

                        # Row shading
                        for r, row in enumerate(pdf_df.itertuples(index=False), start=1):
                              if row.Position == "TOTAL":
                                    style.add("BACKGROUND", (0, r), (-1, r), colors.HexColor("#DCE6F1"))
                                    style.add("FONTNAME", (0, r), (-1, r), "Helvetica-Bold")
                              else:
                                    style.add("BACKGROUND", (var_col, r), (var_col, r), colors.HexColor("#FEF6F0"))

                        table.setStyle(style)
                        elements.append(table)

                        # ---------- BUILD WITH FOOTER ----------
                        doc.build(elements, onFirstPage=footer, onLaterPages=footer)

                        st.download_button(
                              "📥 Download PDF",
                              data=buffer.getvalue(),
                              file_name=f"Schedule_Variance_Report_{week_start:%Y%m%d}.pdf",
                              mime="application/pdf"
                        )

                  elif export_format == "CSV" and st.session_state.get("schedule_variance_ready"):
                        report_df = st.session_state.schedule_variance_df.copy()
                        csv_data = report_df.to_csv(index=False)
                        st.download_button(
                              "📥 Download CSV",
                              data=csv_data,
                              file_name=f"Schedule_Variance_Report_{week_start:%Y%m%d}.csv",
                              mime="text/csv"
                        )

      # ─────────────── REPORT: COST MGMT – FTE VARIANCE ───────────────
      if report_type == "Cost Mgmt FTE Variance":

            st.markdown("### Cost Management: FTE Variance Report Filters")

            from collections import defaultdict

            # ---------------------------------------------------------
            # WEEK SELECTION (EXACT SAME PATTERN AS SCHEDULE VARIANCE)
            # ---------------------------------------------------------
            if "rpt_cost_week_date" not in st.session_state:
                  st.session_state.rpt_cost_week_date = date.today()

            sel_date = st.date_input(
                  "Select any date in the week",
                  value=st.session_state.rpt_cost_week_date,
                  key="rpt_cost_week_date"
            )

            week_start = sel_date + relativedelta(weekday=MO(-1))
            week_end   = week_start + timedelta(days=6)
            week_dates = [week_start + timedelta(days=i) for i in range(7)]

            st.markdown(
                  f"📅 **Selected Week:** {week_start:%Y-%m-%d} to {week_end:%Y-%m-%d}"
            )

            fmt_day  = "%a %#m/%#d" if os.name == "nt" else "%a %-m/%-d"
            day_cols = [d.strftime(fmt_day) for d in week_dates]

            # ---------------------------------------------------------
            # DEPARTMENT (EXACT SAME STYLE AS SCHEDULE VARIANCE)
            # ---------------------------------------------------------
            emp_df = refresh(db.Employee)

            emp_df = apply_manager_scope(
                  emp_df.rename(columns={"role": "position"})
            ).rename(columns={"position": "role"})

            dept_list = ["All"] + sorted(
                  emp_df["department"]
                  .dropna()
                  .unique()
                  .tolist()
            )

            sel_dept = st.selectbox("Department", dept_list, key="rpt_cost_dept")

            # ---------------------------------------------------------
            # GENERATE BUTTON (EXACT SAME FLOW)
            # ---------------------------------------------------------
            generate_cost_fte = st.button("📊 Generate Cost Mgmt FTE Variance Report")

            if not generate_cost_fte:
                  st.stop()

            # ─────────────── Actual Hours ───────────────
            actual_q = (
                  session.query(
                        db.Position.name.label("Position"),
                        db.Actual.date.label("Date"),
                        func.sum(db.Actual.hours).label("Actual Hours")
                  )
                  .join(db.Position, db.Actual.position_id == db.Position.id)
                  .join(db.Department, db.Position.department_id == db.Department.id)
                  .filter(db.Actual.date.between(week_start, week_end))
            )

            if sel_dept != "All":
                  actual_q = actual_q.filter(db.Department.name == sel_dept)

            actual_q = actual_q.group_by(db.Position.name, db.Actual.date)

            actual_df = pd.DataFrame(actual_q.all())
            if not actual_df.empty:
                  actual_df["FTE"] = actual_df["Actual Hours"] / 8
                  actual_df["Date Label"] = pd.to_datetime(actual_df["Date"]).dt.strftime(fmt_day)
            else:
                  actual_df = pd.DataFrame(columns=["Position", "Date", "Actual Hours", "FTE", "Date Label"])

            # ─────────────── Scheduled FTEs ───────────────
            sched_q = (
                  session.query(
                        db.Employee.role.label("Position"),
                        db.Schedule.day,
                        db.Schedule.shift_type
                  )
                  .join(db.Employee, db.Employee.id == db.Schedule.emp_id)
                  .filter(db.Schedule.day.between(week_start, week_end))
            )

            if sel_dept != "All":
                  sched_q = sched_q.filter(db.Employee.department == sel_dept)

            sched_df_raw = pd.DataFrame(sched_q.all())

            def parse_hours(shift):
                  if not shift or shift.upper() == "OFF": return 0
                  try:
                        a,b = shift.split("-")
                        t0 = datetime.strptime(a.strip(), "%H:%M")
                        t1 = datetime.strptime(b.strip(), "%H:%M")
                        diff = (t1 - t0).seconds / 3600
                        return diff if diff > 0 else diff + 24
                  except:
                        return 0

            if not sched_df_raw.empty:
                  sched_df_raw["Hours"] = sched_df_raw["shift_type"].apply(parse_hours)
                  sched_df_raw["Date Label"] = pd.to_datetime(sched_df_raw["day"]).dt.strftime(fmt_day)
                  sched_group = sched_df_raw.groupby(["Position", "Date Label"])["Hours"].sum().reset_index()
                  sched_group["FTE"] = sched_group["Hours"] / 8
            else:
                  sched_group = pd.DataFrame(columns=["Position", "Date Label", "Hours", "FTE"])

            # ─────────────── OTB FTEs ───────────────
            std_q = (
                  session.query(
                        db.Position.name.label("Position"),
                        db.LaborStandard.metric,
                        db.LaborStandard.standard
                  )
                  .join(db.Position, db.Position.id == db.LaborStandard.position_id)
            )

            if sel_dept != "All":
                  std_q = std_q.join(db.Department, db.Position.department_id == db.Department.id)\
                               .filter(db.Department.name == sel_dept)

            std_df = pd.DataFrame(std_q.all())

            otb_q = (
                  session.query(
                        db.RoomOTBPickup.date.label("date"),
                        db.RoomOTBPickup.kpi.label("kpi"),
                        db.RoomOTBPickup.value.label("value")
                  )
                  .filter(db.RoomOTBPickup.date.between(week_start, week_end))
            )

            otb_df_raw = pd.DataFrame(otb_q.all())
            if otb_df_raw.empty:
                  otb_df_raw = pd.DataFrame(columns=["date", "kpi", "value"])

            otb_rows = []
            for _, row in std_df.iterrows():
                  pos, metric, std_val = row["Position"], row["metric"], row["standard"]
                  subset = otb_df_raw[otb_df_raw["kpi"].str.lower() == metric.lower()]
                  for _, r in subset.iterrows():
                        date_lbl = pd.to_datetime(r["date"]).strftime(fmt_day)
                        fte = (r["value"] / std_val) if std_val else 0
                        otb_rows.append({
                              "Position": pos,
                              "Date Label": date_lbl,
                              "OTB FTE": fte
                        })

            otb_df = pd.DataFrame(otb_rows)
            if otb_df.empty:
                  otb_df = pd.DataFrame(columns=["Position", "Date Label", "OTB FTE"])

      # ─────────────── Final Variance Table ───────────────

      # 🛑 HARD SAFETY GUARD (PREVENT NameError ON RERUNS)
      if (
            "actual_df" not in locals()
            or "sched_group" not in locals()
            or "otb_df" not in locals()
      ):
            #st.warning("⚠️ Click **Generate Cost Mgmt FTE Variance Report** to run the report.")
            st.stop()

      pos_list = sorted(
            set(actual_df.get("Position", pd.Series()))
            .union(sched_group.get("Position", pd.Series()))
            .union(otb_df.get("Position", pd.Series()))
      )

      data = []
      export_raw_display = []   # UI table (always positive)
      export_raw_signed  = []   # ✅ TRUE SIGNED NUMBERS FOR EXCEL

      for pos in pos_list:
            row = {"Position": pos}
            export_row_display = {"Position": pos}
            export_row_signed  = {"Position": pos}

            for d in week_dates:
                  lbl = d.strftime(fmt_day)

                  otb_fte = otb_df[
                        (otb_df["Position"] == pos) &
                        (otb_df["Date Label"] == lbl)
                  ]["OTB FTE"].sum()

                  actual_fte = actual_df[
                        (actual_df["Position"] == pos) &
                        (actual_df["Date Label"] == lbl)
                  ]["FTE"].sum()

                  sched_fte = sched_group[
                        (sched_group["Position"] == pos) &
                        (sched_group["Date Label"] == lbl)
                  ]["FTE"].sum()

                  # ----- ✅ TRUE BUSINESS LOGIC -----
                  if actual_fte > 0:
                        # 🟧 ACTUAL VS OTB → UP (NEGATIVE IN EXCEL)
                        fte_var = actual_fte - otb_fte
                        signed_val = -round(fte_var, 2)   # ✅ NEGATIVE
                        arrow   = "▲"
                        bg      = "#FFE6CC"
                        color   = "red"
                  else:
                        # 🟦 SCHEDULED VS OTB → DOWN (POSITIVE)
                        fte_var = sched_fte - otb_fte
                        signed_val = round(fte_var, 2)    # ✅ POSITIVE
                        arrow   = "▼"
                        bg      = "#E6F0FF"
                        color   = "#003366"

                  fte_var = round(fte_var, 2)

                  # ✅ STORE FOR UI (POSITIVE ONLY)
                  export_row_display[lbl] = fte_var

                  # ✅ STORE FOR EXCEL (SIGNED)
                  export_row_signed[lbl] = signed_val

                  # ✅ DISPLAY CELL
                  cell = f"""
                  <div style="
                        background:{bg};
                        color:{color};
                        padding:4px 6px;
                        border-radius:6px;
                        text-align:center;
                        font-weight:600;
                  ">
                        {arrow} {fte_var:.2f}
                  </div>
                  """

                  row[lbl] = cell

            data.append(row)
            export_raw_display.append(export_row_display)
            export_raw_signed.append(export_row_signed)


      # ✅ UI EXPORT COPY (POSITIVE)
      export_df = pd.DataFrame(export_raw_display)

      # ✅ TRUE SIGNED NUMERIC DATAFRAME FOR EXCEL
      signed_numeric_df = pd.DataFrame(export_raw_signed).set_index("Position")

      # ─────────────── Store for Export ───────────────
      st.session_state["report_cost_fte_variance_df"] = export_df
      st.session_state["report_cost_fte_variance_numeric"] = signed_numeric_df
      st.session_state["report_cost_fte_metadata"] = {
            "department": sel_dept,
            "week_start": week_start,
            "week_end": week_end
      }      # ─────────────── EXPORT: COST MGMT FTE VARIANCE (EXCEL) ───────────────
      if (
            report_type == "Cost Mgmt FTE Variance"
            and export_format == "Excel"
            and st.session_state.get("report_cost_fte_variance_numeric") is not None
      ):

            # ✅ TRUE SIGNED VALUES (NO HTML, REAL NUMBERS)
            df = st.session_state["report_cost_fte_variance_numeric"].copy()
            meta = st.session_state["report_cost_fte_metadata"]

            import io
            from datetime import datetime

            output = io.BytesIO()

            # -----------------------------------------------------
            # ✅ ROUND VALUES
            # -----------------------------------------------------
            df = df.round(2)

            # -----------------------------------------------------
            # ✅ ADD TOTAL COLUMN (SIGNED)
            # -----------------------------------------------------
            df["Total"] = df.sum(axis=1).round(2)

            # -----------------------------------------------------
            # ✅ ADD TOTAL ROW (SIGNED)
            # -----------------------------------------------------
            total_row = df.sum(numeric_only=True)
            total_row.name = "TOTAL"
            df = pd.concat([df, pd.DataFrame(total_row).T])

            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                  df.to_excel(
                        writer,
                        sheet_name="Cost FTE Variance",
                        startrow=8,
                        startcol=1,
                        header=False
                  )

                  workbook  = writer.book
                  worksheet = writer.sheets["Cost FTE Variance"]

                  # ==================================================
                  # ✅ SOLID BACKGROUND (MATCH YOUR OTHER REPORTS)
                  # ==================================================
                  solid_bg_fmt = workbook.add_format({'bg_color': '#FFFFFF', 'pattern': 1})
                  worksheet.set_column(0, 60, None, solid_bg_fmt)
                  for r in range(0, 3000):
                        worksheet.set_row(r, None, solid_bg_fmt)

                  worksheet.set_column(0, 0, 2)

                  user_obj   = st.session_state.get("user") or {}
                  username   = user_obj.get("username") or user_obj.get("email") or "User"
                  hotel_name = user_obj.get("hotel_name") or "Hotel"

                  created_str = datetime.now().strftime("%m/%d/%Y %I:%M %p")

                  # ==================================================
                  # ✅ TOP RIGHT METADATA
                  # ==================================================
                  top_right_hotel_fmt = workbook.add_format({
                        'bold': True, 'align': 'right', 'border': 0, 'bg_color': '#FFFFFF'
                  })
                  top_right_created_fmt = workbook.add_format({
                        'align': 'right', 'border': 0, 'bg_color': '#FFFFFF'
                  })

                  rightmost_col = 1 + len(df.columns)
                  worksheet.write(0, rightmost_col, hotel_name, top_right_hotel_fmt)
                  worksheet.write(
                        1, rightmost_col,
                        f"Created by {username} on {created_str}",
                        top_right_created_fmt
                  )

                  # ==================================================
                  # ✅ TITLE + METADATA (LEFT)
                  # ==================================================
                  worksheet.merge_range("B1:E1", "Cost Mgmt: FTE Variance Report", workbook.add_format({
                        'bold': True,
                        'font_size': 14,
                        'align': 'left',
                        'border': 0,
                        'bg_color': '#FFFFFF'
                  }))

                  meta_label_fmt = workbook.add_format({
                        'bold': True,
                        'align': 'left',
                        'border': 0,
                        'bg_color': '#FFFFFF'
                  })

                  meta_value_fmt = workbook.add_format({
                        'align': 'left',
                        'border': 0,
                        'bg_color': '#FFFFFF'
                  })

                  worksheet.write("B3", "Department:", meta_label_fmt)
                  worksheet.write("C3", meta.get("department") or "(All)", meta_value_fmt)

                  worksheet.write("B4", "Week:", meta_label_fmt)
                  worksheet.write(
                        "C4",
                        f'{meta["week_start"]:%m-%d-%Y} to {meta["week_end"]:%m-%d-%Y}',
                        meta_value_fmt
                  )
                  # ==================================================
                  # ✅ EXPLANATION / MEMO (TWO LINES - NOT MERGED)
                  # ==================================================
                  memo_fmt = workbook.add_format({
                        'italic': True,
                        'align': 'left',
                        'border': 0,
                        'bg_color': '#FFFFFF'
                  })

                  worksheet.write(
                        "B6",
                        "**Negative values indicate FTE overage based on Actuals vs Forecast/OTB.",
                        memo_fmt
                  )

                  worksheet.write(
                        "B7",
                        "**Positive values indicate FTE shortage based on Scheduled vs Forecast/OTB.",
                        memo_fmt
                  )

                  # ==================================================
                  # ✅ FORMATS (RESTORED)
                  # ==================================================
                  header_fmt = workbook.add_format({
                        'bold': True,
                        'bg_color': '#16365C',
                        'font_color': '#FFFFFF',
                        'align': 'center',
                        'border': 1
                  })

                  default_fmt = workbook.add_format({
                        'align': 'center',
                        'border': 1
                  })

                  first_col_fmt = workbook.add_format({
                        'align': 'center',
                        'border': 1,
                        'bold': True
                  })

                  total_row_fmt = workbook.add_format({
                        'align': 'center',
                        'border': 1,
                        'bold': True,
                        'bg_color': '#E6EEF8'
                  })

                  total_col_fmt = workbook.add_format({
                        'align': 'center',
                        'border': 1,
                        'bg_color': '#FEF6F0'
                  })

                  # ==================================================
                  # ✅ HEADER ROW
                  # ==================================================
                  for col_idx, col_name in enumerate(["Position"] + list(df.columns)):
                        worksheet.write(7, 1 + col_idx, col_name, header_fmt)

                  # ==================================================
                  # ✅ DATA CELLS WITH ARROWS + PARENTHESES
                  # ==================================================
                  for row_idx, (pos, row) in enumerate(df.iterrows()):
                        excel_row = 8 + row_idx

                        # --- Position column ---
                        if pos == "TOTAL":
                              worksheet.write(excel_row, 1, pos, total_row_fmt)
                        else:
                              worksheet.write(excel_row, 1, pos, first_col_fmt)

                        for col_idx, val in enumerate(row.values):
                              excel_col = 2 + col_idx

                              is_total_row = pos == "TOTAL"
                              is_total_col = df.columns[col_idx] == "Total"

                              if is_total_row:
                                    fmt = total_row_fmt
                              elif is_total_col:
                                    fmt = total_col_fmt
                              else:
                                    fmt = default_fmt

                              # ✅ ARROWS + PARENTHESES DISPLAY
                              if isinstance(val, (int, float)) and val < 0:
                                    disp = f"▲ ({abs(val):.2f})"
                              else:
                                    disp = f"▼ {val:.2f}"

                              worksheet.write(excel_row, excel_col, disp, fmt)

                  # ==================================================
                  # ✅ AUTO COLUMN WIDTH
                  # ==================================================
                  for i in range(len(df.columns) + 1):
                        worksheet.set_column(1 + i, 1 + i, 13)

                  # ==================================================
                  # ✅ FOOTER
                  # ==================================================
                  footer_row = 8 + len(df) + 1
                  worksheet.write(
                        footer_row,
                        1,
                        "Confidential | 2025 Labor Pilot",
                        workbook.add_format({'align': 'left'})
                  )

            st.download_button(
                  "⬇️ Download Excel",
                  data=output.getvalue(),
                  file_name=f"Cost_Mgmt_FTE_Variance_{meta['week_start']:%Y-%m-%d}_to_{meta['week_end']:%Y-%m-%d}.xlsx",
                  mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
      # ─────────────── EXPORT: COST MGMT FTE VARIANCE (PDF) ───────────────
      elif (
            report_type == "Cost Mgmt FTE Variance"
            and export_format == "PDF"
            and st.session_state.get("report_cost_fte_variance_numeric") is not None
      ):
            import io
            from datetime import datetime
            from reportlab.lib.pagesizes import landscape, letter
            from reportlab.platypus import (
                  SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_RIGHT
            from reportlab.lib import colors

            # --- USER INFO ---
            user_obj = st.session_state.get("user") or {}
            username = user_obj.get("username") or user_obj.get("email") or "User"
            hotel_name = user_obj.get("hotel_name") or "Hotel"

            df = st.session_state["report_cost_fte_variance_numeric"].copy()
            meta = st.session_state["report_cost_fte_metadata"]

            # ✅ FORMAT VALUES WITH ARROWS + PARENTHESES
            formatted_df = df.copy()

            for col in formatted_df.columns:
                  formatted_df[col] = formatted_df[col].apply(
                        lambda v: f"▲ ({abs(v):.2f})" if v < 0 else f"▼ {v:.2f}"
                  )

            # ✅ ADD TOTAL COLUMN
            df["Total"] = df.sum(axis=1).round(2)
            formatted_df["Total"] = df["Total"].apply(
                  lambda v: f"▲ ({abs(v):.2f})" if v < 0 else f"▼ {v:.2f}"
            )

            # ✅ ADD TOTAL ROW
            total_row = df.sum(numeric_only=True)
            df.loc["TOTAL"] = total_row

            formatted_total = total_row.apply(
                  lambda v: f"▲ ({abs(v):.2f})" if v < 0 else f"▼ {v:.2f}"
            )
            formatted_df.loc["TOTAL"] = formatted_total

            buffer = io.BytesIO()

            # ---------- FOOTER ----------
            def footer(canvas, doc):
                  canvas.saveState()
                  canvas.setFont("Helvetica", 8)
                  canvas.drawString(36, 20, "Confidential | 2025 Labor Pilot")
                  canvas.drawRightString(letter[1] - 36, 20, f"Page {doc.page}")
                  canvas.restoreState()

            # ---------- DOCUMENT ----------
            doc = SimpleDocTemplate(
                  buffer,
                  pagesize=landscape(letter),
                  leftMargin=36, rightMargin=36,
                  topMargin=36, bottomMargin=36
            )

            styles = getSampleStyleSheet()
            elements = []

            # ========= TITLE + RIGHT USER INFO =========
            title_style = styles["Heading2"]

            right_style = ParagraphStyle(
                  "right_style",
                  parent=styles["Normal"],
                  alignment=TA_RIGHT,
                  fontSize=10
            )

            title_table = Table(
                  [
                        [
                              Paragraph("Cost Mgmt: FTE Variance Report", title_style),
                              Paragraph(
                                    f"<b>{hotel_name}</b><br/>"
                                    f"Created by {username} on {datetime.now():%m/%d/%Y %I:%M %p}",
                                    right_style
                              )
                        ]
                  ],
                  colWidths=["*","*"]
            )

            elements.append(title_table)
            elements.append(Spacer(1, 12))

            # ========= METADATA =========
            elements.append(Paragraph(f"<b>Department:</b> {meta.get('department') or '(All)'}", styles["Normal"]))
            elements.append(Paragraph(
                  f"<b>Week:</b> {meta['week_start']:%m-%d-%Y} to {meta['week_end']:%m-%d-%Y}",
                  styles["Normal"]
            ))
            elements.append(Spacer(1, 10))

            # ========= MEMO =========
            elements.append(Paragraph(
                  "**Negative values indicate FTE overage based on Actuals vs Forecast/OTB.",
                  styles["Italic"]
            ))
            elements.append(Paragraph(
                  "**Positive values indicate FTE shortage based on Scheduled vs Forecast/OTB.",
                  styles["Italic"]
            ))
            elements.append(Spacer(1, 16))

            # ========= TABLE =========
            pdf_data = [["Position"] + list(formatted_df.columns)]
            for idx, row in formatted_df.iterrows():
                  pdf_data.append([idx] + list(row.values))

            table = Table(pdf_data, repeatRows=1)

            style = TableStyle([
                  ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16365C")),
                  ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                  ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                  ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                  ("FONTSIZE", (0, 0), (-1, 0), 9),
                  ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#A6A6A6")),
                  ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ])

            total_row_idx = len(pdf_data) - 1
            total_col_idx = len(pdf_data[0]) - 1

            # ✅ TOTAL ROW = BLUE (FULL ROW)
            style.add(
                  "BACKGROUND",
                  (0, total_row_idx),
                  (-1, total_row_idx),
                  colors.HexColor("#DCE6F1")
            )
            style.add(
                  "FONTNAME",
                  (0, total_row_idx),
                  (-1, total_row_idx),
                  "Helvetica-Bold"
            )

            # ✅ TOTAL COLUMN = ORANGE (EXCEPT TOTAL ROW)
            style.add(
                  "BACKGROUND",
                  (total_col_idx, 1),
                  (total_col_idx, total_row_idx - 1),   # ✅ STOP BEFORE TOTAL ROW
                  colors.HexColor("#FEF6F0")
            )

            table.setStyle(style)
            elements.append(table)

            # ---------- BUILD ----------
            doc.build(elements, onFirstPage=footer, onLaterPages=footer)

            st.download_button(
                  "📥 Download PDF",
                  data=buffer.getvalue(),
                  file_name=f"Cost_Mgmt_FTE_Variance_{meta['week_start']:%Y%m%d}.pdf",
                  mime="application/pdf"
            )
      # ─────────────── EXPORT: COST MGMT FTE VARIANCE (CSV) ───────────────
      elif (
            report_type == "Cost Mgmt FTE Variance"
            and export_format == "CSV"
            and st.session_state.get("report_cost_fte_variance_numeric") is not None
      ):

            import io

            # ✅ Use TRUE signed numeric values
            csv_df = (
                  st.session_state["report_cost_fte_variance_numeric"]
                  .reset_index()   # bring Position back as a column
                  .copy()
            )

            buffer = io.StringIO()
            csv_df.to_csv(buffer, index=False)

            st.download_button(
                  "⬇️ Download CSV",
                  data=buffer.getvalue(),
                  file_name=f"Cost_Mgmt_FTE_Variance_{st.session_state['report_cost_fte_metadata']['week_start']:%Y-%m-%d}.csv",
                  mime="text/csv"
            )

# ─────────────── Scheduled Tasks ───────────────
elif main_choice == "Scheduled Tasks":
      from datetime import date, timedelta
      from dateutil.relativedelta import relativedelta, MO
      import pandas as pd
      import io
      from sqlalchemy import or_

      st.markdown("""
      <div class="la-title">
        <!-- Lucide: calendar-clock (Scheduled Tasks) -->
        <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
             viewBox="0 0 24 24" fill="none" stroke="currentColor"
             stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
          <path d="M21 12V7a2 2 0 0 0-2-2H5a2 2 0 0 0-2 2v12a2 2 0 0 0 2 2h7"></path>
          <line x1="16" y1="2" x2="16" y2="6"></line>
          <line x1="8" y1="2" x2="8" y2="6"></line>
          <line x1="3" y1="10" x2="21" y2="10"></line>
          <circle cx="17" cy="17" r="3"></circle>
          <path d="M17 15v3l2 1"></path>
        </svg>
        <span>Scheduled Tasks</span>
      </div>
      <style>
        .la-title{
          display:flex; align-items:center; gap:10px;
          margin:0 0 10px 0; line-height:1;
        }
        .la-title svg{ color:#111; }
        .la-title span{ font-weight:700; font-size:1.6rem; }
      </style>
      """, unsafe_allow_html=True)

      role = (st.session_state.get("user", {}).get("role") or "").strip().lower()


      # ---------------------------------------------
      # ➤ ADD TABS (Create Task / Saved Tasks)
      # ---------------------------------------------
      tab1, tab2 = st.tabs([
            "➕\uFE0E  Create Task",
            "📆\uFE0E  Saved Tasks"
      ])

      # =========================================================
      # TAB 1 — CREATE TASK
      # =========================================================
      with tab1:

            # ---------------------------
            # Report list
            # ---------------------------
            _all_reports = [
                  "Labor Variance",
                  "Forecast Variance",
                  "OT Risk",
                  "Productivity Index",
            ]

            _visible_reports = [
                  r for r in _all_reports
                  if not (role == "manager" and r == "OT Risk")
            ]

            report_type = st.selectbox(
                  "🧾 Select Report Type",
                  _visible_reports,
                  key="report_type_sel"
            )


            # ================================
            # 📆 DATE MODE (Shared by ALL reports)
            # ================================
            st.markdown("### Select Reporting Period")

            date_mode = st.radio(
                  "Select Reporting Period",
                  options=["Yesterday", "Current Week", "Last Week", "MTD"],
                  horizontal=True,
                  index=0,
                  key="shared_date_mode"
            )

            today = date.today()

            # ----- Auto-compute ACTUALS window -----
            if date_mode == "Yesterday":
                  week_start = today - timedelta(days=1)
                  week_end   = today - timedelta(days=1)

            elif date_mode == "Current Week":
                  week_start = today + relativedelta(weekday=MO(-1))
                  week_end   = week_start + timedelta(days=6)

            elif date_mode == "Last Week":
                  this_monday = today + relativedelta(weekday=MO(-1))
                  week_start  = this_monday - timedelta(days=7)
                  week_end    = this_monday - timedelta(days=1)

            elif date_mode == "MTD":
                  week_start = date(today.year, today.month, 1)
                  week_end   = today

            # ----- Schedule Week (Mon → Sun) -----
            sched_week_start = week_start + relativedelta(weekday=MO(-1))
            sched_week_end   = sched_week_start + timedelta(days=6)

            # ----- DISPLAY RANGE (Shared for ALL reports) -----
            st.markdown(f"**Selected Range (Actuals):** {week_start} to {week_end}")
            st.markdown(f"**Schedule Week:** {sched_week_start} to {sched_week_end}")

            # =================================================================
            # OT RISK REPORT (SCHEDULED TASK VERSION)
            # =================================================================
            if report_type == "OT Risk":

                  st.markdown("### Filters")

                  # ---- Department / Position filters ----
                  dept_df = refresh(db.Department).rename(columns={"name": "Department"})
                  dept_list = sorted(dept_df["Department"].dropna().unique())

                  ot_risk_dept = st.selectbox(
                        "Select Department (OT Risk)",
                        ["(All)"] + dept_list,
                        key="ot_risk_dept"
                  )

                  if ot_risk_dept == "(All)":
                        ot_risk_dept_id = None
                  else:
                        try:
                              ot_risk_dept_id = dept_df[dept_df["Department"] == ot_risk_dept]["id"].values[0]
                        except IndexError:
                              ot_risk_dept_id = None

                  pos_df = refresh(db.Position)
                  if ot_risk_dept_id is None:
                        filtered_positions = pos_df
                  else:
                        filtered_positions = pos_df[pos_df["department_id"] == ot_risk_dept_id]

                  pos_list = sorted(filtered_positions["name"].dropna().unique())

                  ot_risk_pos = st.selectbox(
                        "Select Position (OT Risk)",
                        ["(All)"] + pos_list,
                        key="ot_risk_pos"
                  )

            # =================================================================
            # FORECAST VARIANCE (SCHEDULED TASK VERSION)
            # =================================================================
            if report_type == "Forecast Variance":

                  st.markdown("### Filters")

                  # (Forecast Variance uses SHARED date_mode and shared week_start/week_end)

                  # ---- Optional Department Filter ----
                  dept_df = refresh(db.Department).rename(columns={"name": "Department"})
                  dept_list = sorted(dept_df["Department"].dropna().unique())

                  fv_dept = st.selectbox(
                        "Department (Optional)",
                        ["(All)"] + dept_list,
                        key="fv_task_dept"
                  )

            # =================================================================
            # PRODUCTIVITY INDEX (SCHEDULED TASK VERSION)
            # =================================================================
            if report_type == "Productivity Index":

                  st.markdown("### Filters")

                  # ---- Department list ----
                  dept_df = refresh(db.Department).rename(columns={"name": "Department"})
                  dept_list = sorted(dept_df["Department"].dropna().unique())

                  prod_task_dept = st.selectbox(
                        "Select Department (Productivity Index)",
                        dept_list,   # ✅ REMOVED "(All)"
                        key="prod_task_dept"
                  )

                  # ---- Position list depends on selected department ----
                  pos_df = refresh(db.Position)

                  try:
                        prod_dept_id = dept_df[dept_df["Department"] == prod_task_dept]["id"].values[0]
                        filtered_positions = pos_df[pos_df["department_id"] == prod_dept_id]
                  except:
                        filtered_positions = pos_df

                  pos_list = sorted(filtered_positions["name"].dropna().unique())

                  prod_task_pos = st.selectbox(
                        "Select Position (Optional)",
                        ["(All)"] + pos_list,
                        key="prod_task_pos"
                  )

            # =================================================================
            # LABOR VARIANCE (SCHEDULED TASK VERSION)
            # =================================================================
            if report_type == "Labor Variance":

                  st.markdown("### Filters")

                  # ---- Department supports "All" ----
                  emp_df = refresh(db.Employee)

                  dept_list = ["All"] + sorted(emp_df["department"].dropna().unique())

                  labor_task_dept = st.selectbox(
                        "Select Department (Labor Variance)",
                        dept_list,
                        key="labor_task_dept"   # ✅ UNIQUE KEY
                  )

                  # ---- Position list respects department ----
                  if labor_task_dept == "All":
                        pos_list = emp_df["role"].dropna().unique()
                  else:
                        pos_list = emp_df[emp_df["department"] == labor_task_dept]["role"].dropna().unique()

                  labor_task_pos = st.selectbox(
                        "Select Position (Optional)",
                        ["All"] + sorted(pos_list),
                        key="labor_task_pos"    # ✅ UNIQUE KEY
                  )
            # ====================================
            # DELIVERY SETTINGS (Always Email)
            # ====================================
            st.markdown("""
                <div style="
                    border: 2px solid #bfbfbf;
                    padding: 12px 18px;
                    border-radius: 6px;
                    font-size: 20px;
                    font-weight: 600;
                    background-color: #ffffff;
                    margin-top: 18px;
                    margin-bottom: 8px;
                ">
                    Delivery Settings
                </div>
            """, unsafe_allow_html=True)
            # ---- Email Recipients (required) ----
            email_list = st.text_input(
                  "Recipient Email(s) (required, comma-separated)",
                  key="ot_risk_email_list",
                  placeholder="manager@hotel.com, gm@hotel.com"
            )

            # Validate email presence
            if not email_list:
                  st.warning("Please enter at least one email address.")

            # ---- Frequency (Daily, Weekly, Bi-Weekly) ----
            frequency = st.selectbox(
                  "How Frequently Should This Task Run?",
                  ["Daily", "Weekly", "Bi-Weekly"],
                  key="ot_risk_frequency"
            )

            # ---- Delivery Time ----
            run_time = st.time_input(
                  "Run This Task At",
                  value=dt_mod.time(6, 0),
                  key="ot_risk_run_time"
            )



            # ====================================
            # 💾 SAVE SCHEDULED TASK  (ALWAYS VISIBLE)
            # ====================================
            st.markdown("---")

            save_clicked = st.button("💾 Save Scheduled Task")

            if save_clicked:

                  # ----- Validate required fields -----
                  if not email_list:
                        st.error("Please enter at least one email address.")
                  else:
                        # Convert run_time → "HH:MM"
                        run_time_str = run_time.strftime("%H:%M")

                        # Get logged-in user's hotel
                        user_obj = st.session_state.get("user") or {}
                        user_hotel = user_obj.get("hotel_name") or "Unknown"

                        # ================================
                        # SELECT DEPT / POS BASED ON REPORT
                        # ================================
                        if report_type == "OT Risk":
                              dept_val = ot_risk_dept if ot_risk_dept != "(All)" else None
                              pos_val  = ot_risk_pos if ot_risk_pos != "(All)" else None

                        elif report_type == "Forecast Variance":
                              dept_val = fv_dept if fv_dept != "(All)" else None
                              pos_val  = None   # Forecast Variance has NO position filter

                        elif report_type == "Productivity Index":
                              dept_val = prod_task_dept if prod_task_dept != "(All)" else None
                              pos_val  = prod_task_pos if prod_task_pos != "(All)" else None

                        elif report_type == "Labor Variance":
                              dept_val = labor_task_dept if labor_task_dept != "All" else None
                              pos_val  = labor_task_pos if labor_task_pos != "All" else None

                        else:
                              dept_val = None
                              pos_val  = None

                        # ================================
                        # SAVE TASK
                        # ================================
                        from db import ScheduledTask

                        new_task = ScheduledTask(
                              task_type=report_type,
                              date_mode=date_mode,
                              frequency=frequency,
                              run_time=run_time_str,
                              emails=email_list,
                              department=dept_val,
                              position=pos_val,
                              hotel_name=user_hotel
                        )

                        session.add(new_task)
                        session.commit()

                        st.success("✅ Scheduled Task Saved Successfully!")









      with tab2:

            st.markdown("### 📋 Saved Scheduled Tasks")

            from db import ScheduledTask

            # Load tasks (HotelScoped filter auto applies)
            tasks = (
                  session.query(ScheduledTask)
                        .order_by(ScheduledTask.created_at.desc())
                        .all()
            )

            if not tasks:
                  st.info("No scheduled tasks found.")
            else:
                  for t in tasks:

                        # ---------------------------
                        # COLLAPSIBLE PANEL PER TASK
                        # ---------------------------
                        with st.expander(
                              f"{t.task_type} — Created {t.created_at:%m/%d/%Y %I:%M %p}",
                              expanded=False
                        ):

                              # ---- STATUS ----
                              st.markdown(
                                    "<span style='color:#3CB371;font-weight:600;'>● Active</span>",
                                    unsafe_allow_html=True
                              )

                              # ---- MAIN METADATA ----
                              st.markdown(
                                    f"""
                                    **Type:** Reports  
                                    **Date Mode:** {t.date_mode}  
                                    **Frequency:** {t.frequency}  
                                    **Run Time:** {t.run_time}  
                                    """,
                                    unsafe_allow_html=True
                              )

                              # ---- FILTER DETAILS ----
                              st.markdown(
                                    f"""
                                    **Department:** {t.department or "(All)"}  
                                    **Position:** {t.position or "(All)"}  
                                    """,
                                    unsafe_allow_html=True
                              )

                              # ---- RECIPIENTS ----
                              st.markdown(f"**Recipients:** {t.emails}")

                              st.markdown("---")

                              # ---- ACTION BUTTONS (Send Now | Delete) ----
                              btn_col1, btn_col2 = st.columns([1, 1])

                              with btn_col1:
                                    send_key = f"send_now_{t.id}"
                                    if st.button("📤 Send Now", key=send_key, type="primary"):
                                          with st.spinner("Generating report and sending email..."):
                                                try:
                                                      resp = requests.post(
                                                            f"{API_URL}/run-task/{t.id}",
                                                            headers={"Authorization": f"Bearer {st.session_state.token}"},
                                                            timeout=60
                                                      )
                                                      if resp.status_code == 200:
                                                            msg = resp.json().get("message", "Report sent.")
                                                            st.success(f"✅ {msg}")
                                                      else:
                                                            detail = resp.json().get("detail", "Unknown error")
                                                            st.error(f"❌ {detail}")
                                                except Exception as e:
                                                      st.error(f"❌ Error: {e}")

                              with btn_col2:
                                    delete_key = f"delete_task_{t.id}"
                                    if st.button("🗑️ Delete Task", key=delete_key):
                                          session.delete(t)
                                          session.commit()
                                          st.success("Task deleted.")
                                          st.rerun()

elif main_choice == "Admin":
    if role not in ("admin", "super user"):  # role is already normalized to lowercase
        st.warning("⛔ You are not authorized to access this page.")
        st.stop()

    st.markdown("""
    <div class="la-title">
      <!-- Lucide: settings (gear) -->
      <svg xmlns="http://www.w3.org/2000/svg" width="40" height="40"
           viewBox="0 0 24 24" fill="none" stroke="currentColor"
           stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">
        <path d="M12 15a3 3 0 1 0 0-6 3 3 0 0 0 0 6z"></path>
        <path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 1 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 1 1-4 0v-.09a1.65 1.65 0 0 0-1-1.51 1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 1 1-2.83-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 1 1 0-4h.09a1.65 1.65 0 0 0 1.51-1 1.65 1.65 0 0 0-.33-1.82l-.06-.06A2 2 0 1 1 7.04 3.3l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 1 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 1 1 2.83 2.83l-.06.06a1.65 1.65 0 0 0-.33 1.82V9c0 .65.39 1.24 1 1.51.32.14.68.21 1.04.21H21a2 2 0 1 1 0 4h-.09c-.36 0-.72.07-1.04.21-.61.27-1 .86-1 1.51z"></path>
      </svg>
      <span>Admin Tools</span>
    </div>
    <style>
      .la-title{
        display:flex; align-items:center; gap:10px;
        margin:0 0 10px 0; line-height:1;
      }
      .la-title svg{ color:#111; }         /* pure B/W via currentColor */
      .la-title span{ font-weight:700; font-size:1.6rem; }
    </style>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs([
        "➕\uFE0E Create User", 
        "👥\uFE0E View Users", 
        "🔐\uFE0E View User Rights", 
        "📝\uFE0E Edit User Rights"
    ])

    # ----------------------------- CREATE USER -----------------------------
    with tab1:
        st.subheader("Create New User")

        # -------------------------------
        # LIVE selections (outside form)
        # -------------------------------
        access_control = []
        can_view_hourly = False
        selected_positions_by_dept = {}

        if role in ("admin", "super user"):
              # Cache departments once
              if "departments_cache" not in st.session_state:
                    dept_resp = requests.get(
                          f"{API_URL}/departments",
                          headers={"Authorization": f"Bearer {st.session_state.token}"}
                    )
                    st.session_state.departments_cache = dept_resp.json() if dept_resp.status_code == 200 else []

              # 1) MULTI-SELECT DEPARTMENTS
              selected_departments = st.multiselect(
                    "Select Departments",
                    st.session_state.departments_cache,
                    key="dept_multi_create_user"
              )

              # 2) For each department, fetch positions and show a multiselect
              for dept in selected_departments:
                    pos_resp = requests.get(
                          f"{API_URL}/positions",
                          params={"department": dept},  # URL-safe
                          headers={"Authorization": f"Bearer {st.session_state.token}"}
                    )
                    positions = pos_resp.json() if pos_resp.status_code == 200 else []
                    selected_positions_by_dept[dept] = st.multiselect(
                          f"Select Positions for {dept}",
                          positions,
                          key=f"pos_multi_create_user_{dept}"
                    )

              # 3) Hourly-rate visibility
              can_view_hourly = st.checkbox("Can view hourly rate?", value=False, key="rate_toggle_create_user")

        # --------------------------------
        # CREDENTIALS + SUBMIT (in a form)
        # --------------------------------
        with st.form("create_user_form_v2"):   # <-- unique key
            new_username = st.text_input("Username", key="new_username_create_user")
            new_email = st.text_input("Email Address", key="new_email_create_user")
            new_password = st.text_input("Password", type="password", key="new_password_create_user")

            if role == "admin":
                new_role = st.selectbox("Role", ["Manager", "Employee", "Night Audit", "Asset Manager"], key="role_create_user_admin")
                new_hotel = st.session_state.user["hotel_name"]
                st.text_input("Hotel Name", value=new_hotel, disabled=True, key="hotel_name_display_admin")
            else:
                new_role = st.selectbox("Role", ["Manager", "Admin", "Employee", "Night Audit", "Asset Manager", "Super User"], key="role_create_user_super")
                new_hotel = st.text_input("Hotel Name", key="hotel_name_input_super")

            submit_user = st.form_submit_button("Create User")  # <-- inside the form

        if submit_user:

            if not new_email:
                  st.error("❌ Email is required to create a user.")
                  st.stop()

            payload = {
                  "username": new_username,
                  "email": new_email,
                  "password": new_password,
                  "role": new_role,
                  "hotel_name": new_hotel
            }

            # Only attach access_control if creating a Manager
            if new_role.lower() in ("manager", "employee"):
                  access_control = []
                  for dept, pos_list in selected_positions_by_dept.items():
                        for pos in pos_list:
                              access_control.append({
                                    "department": dept,
                                    "position": pos,
                                    "can_view_hourly_rate": can_view_hourly
                              })

                  if not access_control:
                        st.warning("Please select at least one department and position for this manager.")
                        st.stop()

                  payload["access_control"] = access_control

            response = requests.post(f"{API_URL}/signup", json=payload)

            if response.status_code == 200:
                st.success("✅ User created successfully!")
            else:
                st.error(f"❌ Failed: {response.json().get('detail', 'Unknown error')}")
    # ----------------------------- VIEW USERS -----------------------------
    with tab2:
        st.subheader("All Users")

        response = requests.get(f"{API_URL}/users", headers={
            "Authorization": f"Bearer {st.session_state.token}"
        })

        if response.status_code == 200:
            users = response.json()
            df = pd.DataFrame(users)
            if "email" in df.columns:
                df = df[["username", "email", "role", "hotel_name", "id"]]

            # Admins only see their hotel; Super Users see all
            if role == "admin":
                df = df[df["hotel_name"] == st.session_state.user["hotel_name"]]

            if df.empty:
                st.info("No users found.")
            else:
                st.dataframe(df)

                st.markdown("---")
                st.markdown("**Delete a user**")

                # Build deletable list according to role
                del_df = df.copy()
                if "role" in del_df.columns:
                    # Admins cannot delete Super Users
                    if role == "admin":
                        del_df = del_df[del_df["role"].str.lower() != "super user"]

                # Guard against empty after filtering
                if del_df.empty:
                    st.info("No users available to delete.")
                else:
                    # Pick user to delete
                    del_user = st.selectbox(
                        "Select a user to delete",
                        sorted(del_df["username"].tolist())
                    )

                    # Extra safety: prevent deleting the currently logged-in user
                    is_self = (del_user == st.session_state.user.get("username"))

                    # Simple text confirmation
                    confirm_txt = st.text_input(
                        f"Type DELETE to confirm removing '{del_user}'"
                    )
                    ok_to_delete = (confirm_txt.strip().upper() == "DELETE") and not is_self

                    if is_self:
                        st.info("You cannot delete the user you're currently logged in as.")

                    if st.button("🗑️ Delete User", disabled=not ok_to_delete):
                        del_resp = requests.delete(
                            f"{API_URL}/users/{del_user}",
                            headers={"Authorization": f"Bearer {st.session_state.token}"}
                        )

                        if del_resp.status_code == 200:
                            # ---- success toast (bottom-right, auto-fade) ----
                            success_html = """
                                <div style="
                                    position: fixed;
                                    bottom: 24px;
                                    right: 24px;
                                    background-color: #4CAF50;
                                    color: white;
                                    padding: 14px 20px;
                                    border-radius: 8px;
                                    font-size: 15px;
                                    box-shadow: 0 4px 12px rgba(0,0,0,0.25);
                                    z-index: 10000;
                                    animation: la_fadeOut 5s forwards;">
                                    ✅ User <b>{USER}</b> deleted successfully.
                                </div>
                                <style>
                                    @keyframes la_fadeOut {
                                        0%   { opacity: 1; }
                                        80%  { opacity: 1; }
                                        100% { opacity: 0; display: none; }
                                    }
                                </style>
                            """
                            st.markdown(success_html.replace("{USER}", del_user), unsafe_allow_html=True)

                            # Optional refresh AFTER the toast is visible (no f-string here)
                            st.markdown("""
                                <script>
                                  setTimeout(function(){ window.location.reload(); }, 1200);
                                </script>
                            """, unsafe_allow_html=True)

                        else:
                            # Show API error detail if available
                            try:
                                detail = del_resp.json().get("detail", "Delete failed.")
                            except Exception:
                                detail = "Delete failed."

                            # ---- error toast (bottom-right, auto-fade) ----
                            error_html = """
                                <div style="
                                    position: fixed;
                                    bottom: 24px;
                                    right: 24px;
                                    background-color: #E53935;
                                    color: white;
                                    padding: 14px 20px;
                                    border-radius: 8px;
                                    font-size: 15px;
                                    box-shadow: 0 4px 12px rgba(0,0,0,0.25);
                                    z-index: 10000;
                                    animation: la_fadeOut 6s forwards;">
                                    ❌ {DETAIL}
                                </div>
                                <style>
                                    @keyframes la_fadeOut {
                                        0%   { opacity: 1; }
                                        85%  { opacity: 1; }
                                        100% { opacity: 0; display: none; }
                                    }
                                </style>
                            """
                            st.markdown(error_html.replace("{DETAIL}", str(detail)), unsafe_allow_html=True)
        else:
            st.warning(f"❌ {response.json().get('detail', 'Access denied.')}")
    # ----------------------------- EDIT USER RIGHTS -----------------------------
    with tab4:
        st.subheader("Edit User Rights")

        response = requests.get(f"{API_URL}/users", headers={
            "Authorization": f"Bearer {st.session_state.token}"
        })

        if response.status_code == 200:
            users = response.json()
            df = pd.DataFrame(users)

            if role == "admin":
                df = df[(df["hotel_name"] == st.session_state.user["hotel_name"]) & (df["role"].str.lower() != "super user")]

            if df.empty:
                st.info("No users available to edit.")
            else:
                selected_user = st.selectbox("Select User", df["username"])
                selected = df[df["username"] == selected_user].iloc[0]

                if role == "admin":
                    new_role = st.selectbox("New Role", ["Manager", "Admin", "Employee", "Night Audit", "Asset Manager"])
                    st.text_input("Hotel Name", value=selected["hotel_name"], disabled=True)
                    new_hotel = selected["hotel_name"]
                else:
                    new_role = st.selectbox("New Role", ["Manager", "Admin", "Employee", "Night Audit", "Asset Manager", "Super User"])
                    new_hotel = st.text_input("Hotel Name", value=selected["hotel_name"])

                if st.button("Update User"):
                    payload = {
                        "username": selected_user,
                        "role": new_role,
                        "hotel_name": new_hotel
                    }

                    response = requests.put(f"{API_URL}/users/{selected_user}", json=payload, headers={
                        "Authorization": f"Bearer {st.session_state.token}"
                    })

                    if response.status_code == 200:
                        st.success("✅ User updated successfully!")
                        st.rerun()
                    else:
                        st.error(f"❌ Failed: {response.json().get('detail', 'Update failed.')}")
        else:
            st.warning(f"❌ {response.json().get('detail', 'Access denied.')}")




if __name__ == "__main__":
    pass
